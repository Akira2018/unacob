from __future__ import annotations

import argparse
import re
import sys
import unicodedata
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

import models  # noqa: E402
from database import Base, SessionLocal, engine  # noqa: E402


DEFAULT_WORKBOOK = ROOT / "Previsão Orçamentária unacob 2026.xlsx"
ANO = 2026

CONTA_FALLBACKS = {
    "custas processuais consig a terceiros": "2.14",
    "energia eletria cpfl": "2.21",
    "energia eletrica cpfl": "2.21",
    "aplicacao investimentos": "2.30",
}


def normalizar(valor: str | None) -> str:
    texto = unicodedata.normalize("NFKD", valor or "").encode("ascii", "ignore").decode()
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def numero(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def carregar_previsoes(workbook_path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb.active

    previsoes = []
    for row in range(3, 35):
        classificacao = ws.cell(row=row, column=1).value
        if not classificacao or str(classificacao).strip().upper() == "TOTAIS":
            continue

        valor = round(numero(ws.cell(row=row, column=2).value), 2)
        if valor == 0:
            continue

        previsoes.append({
            "linha": row,
            "classificacao": str(classificacao).strip(),
            "classificacao_norm": normalizar(str(classificacao)),
            "valor": valor,
        })
    return previsoes


def importar(workbook_path: Path) -> None:
    Base.metadata.create_all(bind=engine)
    previsoes = carregar_previsoes(workbook_path)

    db = SessionLocal()
    try:
        if not db.query(models.PlanoConta).filter(models.PlanoConta.codigo == "2.30").first():
            db.add(models.PlanoConta(
                id=str(uuid.uuid4()),
                codigo="2.30",
                nome="Aplicação investimentos",
                tipo="saida",
                ordem=500,
                ativo=True,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            ))
            db.flush()

        contas = db.query(models.PlanoConta).filter(models.PlanoConta.tipo == "saida").all()
        contas_por_nome = {normalizar(conta.nome): conta for conta in contas}
        contas_por_codigo = {conta.codigo: conta for conta in contas}

        criados = 0
        atualizados = 0
        ignorados = []

        for item in previsoes:
            conta = contas_por_nome.get(item["classificacao_norm"])
            if not conta:
                codigo = CONTA_FALLBACKS.get(item["classificacao_norm"])
                conta = contas_por_codigo.get(codigo)

            if not conta:
                ignorados.append(item)
                continue

            previsao = db.query(models.PrevisaoOrcamentariaAnual).filter(
                models.PrevisaoOrcamentariaAnual.conta_id == conta.id,
                models.PrevisaoOrcamentariaAnual.ano == ANO,
            ).first()

            if previsao:
                previsao.valor_previsto_anual = float(item["valor"])
                previsao.observacoes = f"Importado da planilha {workbook_path.name}, linha {item['linha']}"
                previsao.updated_at = datetime.utcnow()
                atualizados += 1
            else:
                db.add(models.PrevisaoOrcamentariaAnual(
                    id=str(uuid.uuid4()),
                    user_id=None,
                    conta_id=conta.id,
                    ano=ANO,
                    valor_previsto_anual=float(item["valor"]),
                    observacoes=f"Importado da planilha {workbook_path.name}, linha {item['linha']}",
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                ))
                criados += 1

        db.commit()

        total = sum(float(item["valor"]) for item in previsoes)
        total_importado = total - sum(float(item["valor"]) for item in ignorados)
        print(f"Previsões lidas: {len(previsoes)} | criadas: {criados} | atualizadas: {atualizados}")
        print(f"Total da planilha: {total:.2f} | total importado: {total_importado:.2f}")
        if ignorados:
            print("Linhas ignoradas por falta de conta correspondente:")
            for item in ignorados:
                print(f"- linha {item['linha']}: {item['classificacao']} = {item['valor']:.2f}")
    finally:
        db.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa a previsão orçamentária anual de 2026 para o banco.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Planilha não encontrada: {args.workbook}")

    importar(args.workbook)


if __name__ == "__main__":
    main()
