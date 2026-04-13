from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Body, Request
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session

import models
import schemas

from database import engine, get_db, Base, SessionLocal
from auth import get_current_user, verify_password, get_password_hash, create_access_token, SECRET_KEY, ALGORITHM
from jose import jwt, JWTError
from typing import Optional, List
import os
import io
import re
import json
import shutil
import sqlite3
import tempfile
import unicodedata
import smtplib
import threading
import time
from pathlib import Path
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from dotenv import load_dotenv
from sqlalchemy import or_, cast, String, inspect, text
from sqlalchemy import func as sql_func
from collections import defaultdict
import calendar
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

load_dotenv(Path(__file__).resolve().parent / ".env")

BACKUP_RETENTION_COUNT = max(1, int(os.getenv("BACKUP_RETENTION_COUNT", "30")))
AUTO_BACKUP_ON_STARTUP = os.getenv("AUTO_BACKUP_ON_STARTUP", "true").strip().lower() not in {"0", "false", "no"}
SCHEDULED_BACKUP_ENABLED = os.getenv("SCHEDULED_BACKUP_ENABLED", "true").strip().lower() not in {"0", "false", "no"}
SCHEDULED_BACKUP_HOUR = min(23, max(0, int(os.getenv("SCHEDULED_BACKUP_HOUR", "20"))))
SCHEDULED_BACKUP_INTERVAL_SECONDS = max(300, int(os.getenv("SCHEDULED_BACKUP_INTERVAL_SECONDS", "900")))
BACKUP_TIMEZONE = os.getenv("BACKUP_TIMEZONE", "America/Sao_Paulo").strip() or "America/Sao_Paulo"

try:
    BACKUP_TZINFO = ZoneInfo(BACKUP_TIMEZONE)
except ZoneInfoNotFoundError:
    BACKUP_TIMEZONE = "UTC"
    BACKUP_TZINFO = timezone.utc

# Create tables
Base.metadata.create_all(bind=engine)


PLANO_CONTAS_PADRAO = [
    {"codigo": "1.1", "nome": "Mensalidades", "tipo": "entrada", "ordem": 110},
    {"codigo": "1.2", "nome": "Fundo social", "tipo": "entrada", "ordem": 120},
    {"codigo": "1.3", "nome": "Fundo para ações jurídicas", "tipo": "entrada", "ordem": 130},
    {"codigo": "1.4", "nome": "Resgate da conta de investimentos", "tipo": "entrada", "ordem": 140},
    {"codigo": "1.5", "nome": "Outras arrecadações", "tipo": "entrada", "ordem": 150},
    {"codigo": "1.6", "nome": "Estorno de tarifa bancária", "tipo": "entrada", "ordem": 160},
    {"codigo": "2.1", "nome": "Auxílio funeral emergencial", "tipo": "saida", "ordem": 210},
    {"codigo": "2.2", "nome": "Diárias/Alimentação", "tipo": "saida", "ordem": 220},
    {"codigo": "2.3", "nome": "Hospedagens", "tipo": "saida", "ordem": 230},
    {"codigo": "2.4", "nome": "Passagens aéreas e terrestres", "tipo": "saida", "ordem": 240},
    {"codigo": "2.5", "nome": "Combustível, pedágio, taxi, uber", "tipo": "saida", "ordem": 250},
    {"codigo": "2.6", "nome": "Ajuda de custo - Estacionamento/área azul", "tipo": "saida", "ordem": 260},
    {"codigo": "2.7", "nome": "Contribuição Assoc. de Classe (FAACO)", "tipo": "saida", "ordem": 270},
    {"codigo": "2.8", "nome": "Xerox/serviços gráficos", "tipo": "saida", "ordem": 280},
    {"codigo": "2.9", "nome": "Impostos e Taxas (IPTU, ISS, Cartórios)", "tipo": "saida", "ordem": 290},
    {"codigo": "2.10", "nome": "Comemorações e eventos", "tipo": "saida", "ordem": 300},
    {"codigo": "2.11", "nome": "Conservação, manutenção e limpeza", "tipo": "saida", "ordem": 310},
    {"codigo": "2.12", "nome": "Obras e reparos", "tipo": "saida", "ordem": 320},
    {"codigo": "2.13", "nome": "Despesas postais", "tipo": "saida", "ordem": 330},
    {"codigo": "2.14", "nome": "Custas processuais/consig.a terceiros - JANOT", "tipo": "saida", "ordem": 340},
    {"codigo": "2.15", "nome": "Honorários Contábeis", "tipo": "saida", "ordem": 350},
    {"codigo": "2.16", "nome": "Seguros", "tipo": "saida", "ordem": 360},
    {"codigo": "2.17", "nome": "Máquinas, Aparelhos e Equipamentos", "tipo": "saida", "ordem": 370},
    {"codigo": "2.18", "nome": "Materiais para Escritório", "tipo": "saida", "ordem": 380},
    {"codigo": "2.19", "nome": "Copa e cozinha - insumos", "tipo": "saida", "ordem": 390},
    {"codigo": "2.20", "nome": "Móveis e Utensílios", "tipo": "saida", "ordem": 400},
    {"codigo": "2.21", "nome": "Energia elétrica - CPFL", "tipo": "saida", "ordem": 410},
    {"codigo": "2.22", "nome": "Consumo de água - DAE", "tipo": "saida", "ordem": 420},
    {"codigo": "2.23", "nome": "Despesas bancárias", "tipo": "saida", "ordem": 430},
    {"codigo": "2.24", "nome": "Telefone e Internet", "tipo": "saida", "ordem": 440},
    {"codigo": "2.25", "nome": "Repasse fundo ações judiciais", "tipo": "saida", "ordem": 450},
    {"codigo": "2.26", "nome": "Repasse fundo ações sociais", "tipo": "saida", "ordem": 460},
    {"codigo": "2.27", "nome": "Sorteios e brindes", "tipo": "saida", "ordem": 470},
    {"codigo": "2.28", "nome": "Perda de associados", "tipo": "saida", "ordem": 480},
    {"codigo": "2.29", "nome": "Provisão para impostos e taxas (ISS)", "tipo": "saida", "ordem": 490},
]

PLANO_CONTAS_PADRAO_POR_CODIGO = {c["codigo"]: c for c in PLANO_CONTAS_PADRAO}


def _normalizar_codigo_conta_seed(codigo: Optional[str]) -> str:
    return (codigo or "").strip().replace(",", ".")


def _normalizar_tipo_conta_seed(tipo: Optional[str]) -> str:
    tipo_norm = unicodedata.normalize("NFKD", (tipo or "")).encode("ascii", "ignore").decode().strip().lower()
    if tipo_norm in {"entrada", "entradas"}:
        return "entrada"
    if tipo_norm in {"saida", "saidas"}:
        return "saida"
    return ""


def _tipo_conta_por_codigo(codigo: str) -> str:
    if codigo.startswith("1."):
        return "entrada"
    if codigo.startswith("2."):
        return "saida"
    return "saida"


def _cidade_corrompida(valor: Optional[str]) -> bool:
    cidade = (valor or "").strip()
    if not cidade or cidade == "-":
        return False

    if len(cidade) > 60:
        return True
    if "\n" in cidade:
        return True

    possui_pontuacao_frase = any(ch in cidade for ch in [".", ";", "!", "?"])
    palavras = [p for p in cidade.split() if p]
    if possui_pontuacao_frase and len(palavras) >= 6:
        return True

    cidade_norm = unicodedata.normalize("NFKD", cidade).encode("ascii", "ignore").decode().lower()
    marcadores = [
        "as vezes",
        "isso acontece",
        "sao apresentados",
        "sistema de classificacao",
        "estao sendo construidos",
    ]
    return any(marker in cidade_norm for marker in marcadores)


def _ensure_financeiro_columns_and_seed_contas():
    inspector = inspect(engine)

    def _has_column(table_name: str, column_name: str) -> bool:
        try:
            cols = inspector.get_columns(table_name)
        except Exception:
            return False
        return any(c.get("name") == column_name for c in cols)

    def _ensure_column(table_name: str, column_name: str, ddl_type: str) -> None:
        if _has_column(table_name, column_name) is False:
            conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {ddl_type}"))

    with engine.begin() as conn:
        _ensure_column("despesas", "conta_id", "VARCHAR(36)")
        _ensure_column("despesas", "conta_codigo", "VARCHAR(20)")
        _ensure_column("despesas", "conta_nome", "VARCHAR(255)")

        _ensure_column("outras_rendas", "conta_id", "VARCHAR(36)")
        _ensure_column("outras_rendas", "conta_codigo", "VARCHAR(20)")
        _ensure_column("outras_rendas", "conta_nome", "VARCHAR(255)")

        _ensure_column("aplicacoes_financeiras", "data_aplicacao", "DATE")
        _ensure_column("aplicacoes_financeiras", "origem_registro", "VARCHAR(50)")
        _ensure_column("aplicacoes_financeiras", "conta_origem", "VARCHAR(150)")
        _ensure_column("aplicacoes_financeiras", "arquivo_origem", "VARCHAR(255)")
        _ensure_column("aplicacoes_financeiras", "imposto_renda", "FLOAT")
        _ensure_column("aplicacoes_financeiras", "iof", "FLOAT")
        _ensure_column("aplicacoes_financeiras", "rendimento_liquido", "FLOAT")
        _ensure_column("aplicacoes_financeiras", "updated_at", "TIMESTAMP")
        _ensure_column("conciliacoes", "despesa_id", "VARCHAR(36)")
        _ensure_column("conciliacoes", "outra_renda_id", "VARCHAR(36)")

    from database import SessionLocal
    db = SessionLocal()
    try:
        contas_existentes = db.query(models.PlanoConta).order_by(models.PlanoConta.created_at.asc()).all()
        existentes = {}

        for conta in contas_existentes:
            codigo_norm = _normalizar_codigo_conta_seed(conta.codigo)
            if not codigo_norm:
                continue

            conta.codigo = codigo_norm
            tipo_norm = _normalizar_tipo_conta_seed(conta.tipo)
            conta.tipo = tipo_norm or _tipo_conta_por_codigo(codigo_norm)

            principal = existentes.get(codigo_norm)
            if not principal:
                existentes[codigo_norm] = conta
                continue

            db.query(models.Despesa).filter(models.Despesa.conta_id == conta.id).update(
                {
                    models.Despesa.conta_id: principal.id,
                    models.Despesa.conta_codigo: principal.codigo,
                    models.Despesa.conta_nome: principal.nome,
                },
                synchronize_session=False,
            )
            db.query(models.OutraRenda).filter(models.OutraRenda.conta_id == conta.id).update(
                {
                    models.OutraRenda.conta_id: principal.id,
                    models.OutraRenda.conta_codigo: principal.codigo,
                    models.OutraRenda.conta_nome: principal.nome,
                },
                synchronize_session=False,
            )
            db.delete(conta)

        for conta in existentes.values():
            padrao = PLANO_CONTAS_PADRAO_POR_CODIGO.get(conta.codigo)
            if padrao:
                conta.nome = padrao["nome"]
                conta.tipo = padrao["tipo"]
                conta.ordem = padrao["ordem"]
            else:
                conta.tipo = _normalizar_tipo_conta_seed(conta.tipo) or _tipo_conta_por_codigo(conta.codigo)

            if conta.ativo is None:
                conta.ativo = True
            conta.updated_at = datetime.utcnow()

        for conta in PLANO_CONTAS_PADRAO:
            atual = existentes.get(conta["codigo"])
            if atual:
                atual.nome = conta["nome"]
                atual.tipo = conta["tipo"]
                atual.ordem = conta["ordem"]
                if atual.ativo is None:
                    atual.ativo = True
                atual.updated_at = datetime.utcnow()
            else:
                db.add(models.PlanoConta(
                    id=str(uuid.uuid4()),
                    codigo=conta["codigo"],
                    nome=conta["nome"],
                    tipo=conta["tipo"],
                    ordem=conta["ordem"],
                    ativo=True,
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                ))

        membros = db.query(models.Membro).all()
        for membro in membros:
            cidade_atual = (membro.cidade or "").strip()
            if not cidade_atual:
                continue

            if _cidade_corrompida(cidade_atual):
                obs_atual = (membro.observacoes or "").strip()
                if cidade_atual not in obs_atual:
                    membro.observacoes = f"{obs_atual}\n{cidade_atual}".strip() if obs_atual else cidade_atual
                membro.cidade = None
                membro.updated_at = datetime.utcnow()

        db.commit()
    finally:
        db.close()


_ensure_financeiro_columns_and_seed_contas()


app = FastAPI(title="UNACOB - União dos aposentados dos correios em Bauru - SP API", version="1.0.0")

# Endpoint para upload e processamento do PDF do Banco do Brasil
import pdfplumber
import os
from fastapi import UploadFile, File

@app.post("/api/conciliacao/importar/pdf-bb")
async def importar_pdf_bb(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    nome_arquivo = (file.filename or "").strip() or "extrato_bb.pdf"
    if not nome_arquivo.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Arquivo deve ser PDF")

    temp_path = None
    banco = "DABB"

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            temp_path = tmp.name
            tmp.write(await file.read())

        with pdfplumber.open(temp_path) as pdf:
            texto = "\n".join((page.extract_text() or "") for page in pdf.pages).strip()

        if not texto:
            raise HTTPException(status_code=400, detail="Nao foi possivel extrair texto do PDF")

        membros_por_codigo_dabb = _indexar_membros_por_codigo_dabb(db)
        importados = []
        linhas_lidas = 0
        linhas_duplicadas = 0
        linhas_invalidas = 0
        total_baixas_automaticas = 0
        total_sem_membro = 0
        total_codigos_ambiguos = 0
        meses_lidos = set()
        codigos_sem_membro = {}
        codigos_ambiguos = {}
        diagnostico_dabb = {
            "linhas_detalhe_encontradas": 0,
            "linhas_detalhe_validas": 0,
            "motivos_invalidos": {},
            "exemplos_invalidos": [],
        }

        for indice, tx in enumerate(_iterar_transacoes_pdf_bb(texto), start=1):
            diagnostico_dabb["linhas_detalhe_encontradas"] += 1
            codigo_dabb = tx.get("codigo_dabb")
            data = tx.get("data")
            valor = tx.get("valor")

            if not codigo_dabb or not data or valor is None:
                linhas_invalidas += 1
                diagnostico_dabb["motivos_invalidos"]["bloco_invalido"] = diagnostico_dabb["motivos_invalidos"].get("bloco_invalido", 0) + 1
                if len(diagnostico_dabb["exemplos_invalidos"]) < 5:
                    diagnostico_dabb["exemplos_invalidos"].append({
                        "linha": indice,
                        "motivo": "bloco_invalido",
                        "conteudo": (tx.get("bloco_original") or "")[:160],
                    })
                continue

            diagnostico_dabb["linhas_detalhe_validas"] += 1
            linhas_lidas += 1
            tipo = tx["tipo"]
            descricao = tx["descricao"]
            numero_doc = tx.get("numero_documento")
            nome_pagador = tx.get("nome")
            mes_ref = data.strftime("%Y-%m")
            meses_lidos.add(mes_ref)

            existe = _buscar_duplicado_conciliacao(
                db=db,
                data_extrato=data,
                valor_extrato=valor,
                banco=banco,
                tipo=tipo,
                numero_documento=numero_doc,
                descricao_extrato=descricao,
            )

            if existe:
                linhas_duplicadas += 1
                continue

            observacoes = [
                f"Arquivo PDF BB: {nome_arquivo}",
                f"codigo_dabb={codigo_dabb}",
            ]
            if nome_pagador:
                observacoes.append(f"nome={nome_pagador}")

            c = models.Conciliacao(
                id=str(uuid.uuid4()),
                user_id=current_user.id,
                data_extrato=data,
                descricao_extrato=descricao,
                valor_extrato=valor,
                tipo=tipo,
                mes_referencia=mes_ref,
                banco=banco,
                numero_documento=numero_doc,
                conciliado=False,
                observacoes=" | ".join(observacoes),
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow()
            )
            db.add(c)
            db.flush()

            membros_match = {}
            for variante in _variantes_codigo_dabb(codigo_dabb):
                for membro in membros_por_codigo_dabb.get(variante, []):
                    membros_match[membro.id] = membro

            if len(membros_match) == 1:
                membro = next(iter(membros_match.values()))
                _baixar_pagamento_mensalidade_por_conciliacao(
                    db=db,
                    conciliacao=c,
                    membro=membro,
                    user_id=current_user.id,
                    observacao_origem=(
                        f"Baixa automática via PDF Banco do Brasil ({mes_ref}) - "
                        f"codigo_dabb {codigo_dabb}"
                    ),
                )
                total_baixas_automaticas += 1
            elif len(membros_match) > 1:
                total_codigos_ambiguos += 1
                item = codigos_ambiguos.setdefault(codigo_dabb, {
                    "codigo_dabb": codigo_dabb,
                    "quantidade": 0,
                    "valores": set(),
                    "meses": set(),
                    "registros": set(),
                })
                item["quantidade"] += 1
                item["valores"].add(round(float(valor or 0), 2))
                item["meses"].add(mes_ref)
                item["registros"].add("PDF")
                c.observacoes = (
                    (c.observacoes + "\n") if c.observacoes else ""
                ) + "Codigo DABB encontrado em mais de um membro ativo; baixa nao realizada automaticamente."
            else:
                total_sem_membro += 1
                item = codigos_sem_membro.setdefault(codigo_dabb, {
                    "codigo_dabb": codigo_dabb,
                    "quantidade": 0,
                    "valores": set(),
                    "meses": set(),
                    "registros": set(),
                })
                item["quantidade"] += 1
                item["valores"].add(round(float(valor or 0), 2))
                item["meses"].add(mes_ref)
                item["registros"].add("PDF")
                c.observacoes = (
                    (c.observacoes + "\n") if c.observacoes else ""
                ) + "Nenhum membro ativo encontrado para o codigo DABB; baixa nao realizada automaticamente."

            importados.append({
                "data": data.strftime("%Y-%m-%d"),
                "descricao": descricao,
                "valor": valor,
                "tipo": tipo,
                "numero_doc": numero_doc,
                "codigo_dabb": codigo_dabb,
                "nome": nome_pagador,
                "conciliado": c.conciliado,
            })

        db.commit()
        meses_importados = sorted({item.get("data", "")[:7] for item in importados if item.get("data")})
        codigos_sem_membro_lista = [
            {
                "codigo_dabb": item["codigo_dabb"],
                "quantidade": item["quantidade"],
                "valores": sorted(item["valores"]),
                "meses": sorted(item["meses"]),
                "registros": sorted(item["registros"]),
            }
            for item in sorted(codigos_sem_membro.values(), key=lambda x: x["codigo_dabb"])
        ]
        codigos_ambiguos_lista = [
            {
                "codigo_dabb": item["codigo_dabb"],
                "quantidade": item["quantidade"],
                "valores": sorted(item["valores"]),
                "meses": sorted(item["meses"]),
                "registros": sorted(item["registros"]),
            }
            for item in sorted(codigos_ambiguos.values(), key=lambda x: x["codigo_dabb"])
        ]

        return {
            "ok": True,
            "mensagem": (
                f"PDF processado com sucesso: {len(importados)} lançamento(s) importado(s), "
                f"{total_baixas_automaticas} baixa(s) automática(s)."
            ),
            "total_importados": len(importados),
            "linhas_lidas": linhas_lidas,
            "linhas_duplicadas": linhas_duplicadas,
            "linhas_invalidas": linhas_invalidas,
            "total_baixas_automaticas": total_baixas_automaticas,
            "total_despesas_automaticas": 0,
            "total_sem_membro": total_sem_membro,
            "total_codigos_ambiguos": total_codigos_ambiguos,
            "codigos_sem_membro": codigos_sem_membro_lista,
            "codigos_ambiguos": codigos_ambiguos_lista,
            "diagnostico_dabb": diagnostico_dabb,
            "meses_importados": meses_importados,
            "meses_lidos": sorted(meses_lidos),
            "registros": importados,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao processar PDF do Banco do Brasil: {str(e)}")
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def create_startup_backup_if_needed():
    try:
        _ensure_daily_startup_backup()
    except Exception:
        # Falha de backup automático não deve impedir a API de iniciar.
        pass

    try:
        _start_backup_scheduler()
    except Exception:
        pass


@app.on_event("shutdown")
def stop_backup_scheduler():
    _backup_scheduler_stop_event.set()


FINANCE_API_PREFIXES = (
    "/api/pagamentos",
    "/api/contas",
    "/api/previsoes-orcamentarias",
    "/api/despesas",
    "/api/outras-rendas",
    "/api/aplicacoes-financeiras",
    "/api/transacoes",
    "/api/saldo-inicial",
    "/api/fluxo-caixa",
    "/api/financeiro",
    "/api/conciliacao",
)

FINANCE_REPORT_PREFIXES = (
    "/api/relatorios/pagamentos",
    "/api/relatorios/balancete",
    "/api/relatorios/livro-diario",
    "/api/relatorios/conciliacao",
    "/api/relatorios/aplicacoes-financeiras",
    "/api/relatorios/consolidado-financeiro",
)


def _is_finance_path(path: str) -> bool:
    if any(path.startswith(prefix) for prefix in FINANCE_API_PREFIXES):
        return True
    if any(path.startswith(prefix) for prefix in FINANCE_REPORT_PREFIXES):
        return True
    return False


def _validate_password_strength(password: str):
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="A senha deve ter no mínimo 8 caracteres")
    if not re.search(r"[A-Z]", password):
        raise HTTPException(status_code=400, detail="A senha deve conter pelo menos 1 letra maiúscula")
    if not re.search(r"[a-z]", password):
        raise HTTPException(status_code=400, detail="A senha deve conter pelo menos 1 letra minúscula")
    if not re.search(r"\d", password):
        raise HTTPException(status_code=400, detail="A senha deve conter pelo menos 1 número")
    if not re.search(r"[^A-Za-z0-9]", password):
        raise HTTPException(status_code=400, detail="A senha deve conter pelo menos 1 caractere especial")


@app.middleware("http")
async def role_access_middleware(request: Request, call_next):
    path = request.url.path or ""

    if path.startswith("/api/auth") or path == "/api/health":
        return await call_next(request)

    users_path = path.startswith("/api/users")
    users_self_path = path == "/api/users/me"
    users_list_path = path == "/api/users"
    finance_path = _is_finance_path(path)
    if not users_path and not finance_path:
        return await call_next(request)

    auth_header = request.headers.get("authorization") or ""
    if not auth_header.lower().startswith("bearer "):
        return await call_next(request)

    token = auth_header.split(" ", 1)[1].strip()
    if not token:
        return await call_next(request)

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
    except JWTError:
        return await call_next(request)

    if not user_id:
        return await call_next(request)

    db = SessionLocal()
    try:
        user = db.query(models.User).filter(models.User.id == user_id).first()
        role = (user.role or "").lower() if user else ""

        if users_path and not users_self_path and not users_list_path and role != "administrador":
            return JSONResponse(status_code=403, content={"detail": "Acesso restrito a administradores"})

        if finance_path and role == "assistente":
            return JSONResponse(status_code=403, content={"detail": "Perfil assistente sem acesso ao módulo financeiro"})
    finally:
        db.close()

    return await call_next(request)

# ─── Seed initial admin user ───────────────────────────────────────────────────
def seed_admin():
    from database import SessionLocal
    db = SessionLocal()
    try:
        admin = db.query(models.User).filter(models.User.email == "admin@associacao.com").first()
        if not admin:
            admin = models.User(
                id=str(uuid.uuid4()),
                email="admin@associacao.com",
                nome_completo="Administrador",
                role="administrador",
                password=get_password_hash("admin123"),
                ativo=True,
                created_at=datetime.utcnow()
            )
            db.add(admin)
            db.commit()
    finally:
        db.close()

seed_admin()


@app.get("/api/health")
def health_check():
    return {"status": "ok", "service": "unacob-backend"}


def _assert_admin(current_user: models.User):
    if (current_user.role or "").lower() != "administrador":
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores")


def _sqlite_db_path_or_400() -> Path:
    db_url = str(engine.url)
    if not db_url.startswith("sqlite:///"):
        raise HTTPException(status_code=400, detail="Backup/restauração disponível apenas para SQLite")

    db_file = engine.url.database
    if not db_file:
        raise HTTPException(status_code=400, detail="Arquivo do banco SQLite não identificado")

    db_path = Path(db_file)
    if not db_path.is_absolute():
        db_path = (Path(__file__).resolve().parent / db_path).resolve()
    return db_path


def _sqlite_backup_dir_or_400() -> Path:
    db_path = _sqlite_db_path_or_400()
    backup_dir = db_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    return backup_dir


def _cleanup_old_backups(backup_dir: Path, keep_count: int = BACKUP_RETENTION_COUNT) -> None:
    if keep_count <= 0:
        return

    backup_files = sorted(
        backup_dir.glob("*.db"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )

    for old_file in backup_files[keep_count:]:
        try:
            old_file.unlink(missing_ok=True)
        except Exception:
            pass


def _backup_now() -> datetime:
    return datetime.now(BACKUP_TZINFO)


def _create_sqlite_backup_file(source_db_path: Path, prefix: str = "unacob_backup") -> Path:
    backup_dir = _sqlite_backup_dir_or_400()
    ts = _backup_now().strftime("%Y%m%d_%H%M%S")
    backup_file = backup_dir / f"{prefix}_{ts}.db"

    with sqlite3.connect(str(source_db_path)) as source_conn:
        with sqlite3.connect(str(backup_file)) as dest_conn:
            source_conn.backup(dest_conn)

    _cleanup_old_backups(backup_dir)
    return backup_file


def _ensure_daily_startup_backup() -> Optional[Path]:
    if not AUTO_BACKUP_ON_STARTUP:
        return None

    db_path = _sqlite_db_path_or_400()
    if not db_path.exists():
        return None

    backup_dir = _sqlite_backup_dir_or_400()
    today_stamp = _backup_now().strftime("%Y%m%d")
    existing_daily_backup = next(
        (item for item in backup_dir.glob(f"startup_backup_{today_stamp}_*.db")),
        None,
    )

    if existing_daily_backup:
        _cleanup_old_backups(backup_dir)
        return existing_daily_backup

    return _create_sqlite_backup_file(db_path, prefix="startup_backup")


def _resolve_backup_file_or_404(filename: str) -> Path:
    backup_dir = _sqlite_backup_dir_or_400()
    safe_name = Path(filename or "").name
    if not safe_name or safe_name != filename or not safe_name.lower().endswith(".db"):
        raise HTTPException(status_code=400, detail="Nome de arquivo de backup inválido")

    backup_file = (backup_dir / safe_name).resolve()
    if backup_file.parent != backup_dir.resolve() or not backup_file.exists():
        raise HTTPException(status_code=404, detail="Arquivo de backup não encontrado")

    return backup_file


def _get_backup_type(filename: str) -> str:
    name = (filename or "").lower()
    if name.startswith("startup_backup_"):
      return "startup"
    if name.startswith("scheduled_backup_"):
      return "agendado"
    if name.startswith("before_restore_"):
      return "pre_restauracao"
    if name.startswith("unacob_backup_"):
      return "manual"
    return "outro"


_backup_scheduler_stop_event = threading.Event()
_backup_scheduler_thread = None


def _ensure_scheduled_backup() -> Optional[Path]:
    if not SCHEDULED_BACKUP_ENABLED:
        return None

    db_path = _sqlite_db_path_or_400()
    if not db_path.exists():
        return None

    backup_dir = _sqlite_backup_dir_or_400()
    now = _backup_now()
    today_stamp = now.strftime("%Y%m%d")
    existing_scheduled_backup = next(
        (item for item in backup_dir.glob(f"scheduled_backup_{today_stamp}_*.db")),
        None,
    )

    if existing_scheduled_backup:
        _cleanup_old_backups(backup_dir)
        return existing_scheduled_backup

    if now.hour < SCHEDULED_BACKUP_HOUR:
        return None

    return _create_sqlite_backup_file(db_path, prefix="scheduled_backup")


def _backup_scheduler_loop() -> None:
    while not _backup_scheduler_stop_event.is_set():
        try:
            _ensure_scheduled_backup()
        except Exception:
            pass
        _backup_scheduler_stop_event.wait(SCHEDULED_BACKUP_INTERVAL_SECONDS)


def _start_backup_scheduler() -> None:
    global _backup_scheduler_thread

    if not SCHEDULED_BACKUP_ENABLED:
        return

    if _backup_scheduler_thread and _backup_scheduler_thread.is_alive():
        return

    _backup_scheduler_stop_event.clear()
    _backup_scheduler_thread = threading.Thread(
        target=_backup_scheduler_loop,
        name="backup-scheduler",
        daemon=True,
    )
    _backup_scheduler_thread.start()


@app.get("/api/admin/system/schema")
def schema_diagnostic(current_user=Depends(get_current_user)):
    _assert_admin(current_user)

    inspector = inspect(engine)
    required_columns = {
        "aplicacoes_financeiras": [
            "id",
            "user_id",
            "mes_referencia",
            "data_aplicacao",
            "instituicao",
            "produto",
            "origem_registro",
            "conta_origem",
            "arquivo_origem",
            "saldo_anterior",
            "aplicacoes",
            "rendimento_bruto",
            "imposto_renda",
            "iof",
            "impostos",
            "rendimento_liquido",
            "resgate",
            "saldo_atual",
            "observacoes",
            "created_at",
            "updated_at",
        ]
    }

    tables = {}
    has_any_issue = False

    for table_name, expected in required_columns.items():
        try:
            columns = [c.get("name") for c in inspector.get_columns(table_name)]
            missing = [c for c in expected if c not in columns]
            tables[table_name] = {
                "exists": True,
                "missing_columns": missing,
                "column_count": len(columns),
            }
            if missing:
                has_any_issue = True
        except Exception as exc:
            has_any_issue = True
            tables[table_name] = {
                "exists": False,
                "missing_columns": expected,
                "error": str(exc),
            }

    db_backend = engine.url.get_backend_name()
    response = {
        "status": "error" if has_any_issue else "ok",
        "database": {
            "backend": db_backend,
            "database": engine.url.database if db_backend == "sqlite" else None,
        },
        "tables": tables,
    }
    return response


@app.get("/api/admin/system/backup")
def backup_database(current_user=Depends(get_current_user)):
    _assert_admin(current_user)

    db_path = _sqlite_db_path_or_400()
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="Banco de dados não encontrado")

    backup_file = _create_sqlite_backup_file(db_path)

    return FileResponse(
        path=str(backup_file),
        media_type="application/octet-stream",
        filename=backup_file.name,
    )


@app.get("/api/admin/system/backups")
def list_backups(current_user=Depends(get_current_user)):
    _assert_admin(current_user)

    backup_dir = _sqlite_backup_dir_or_400()
    backups = []
    for backup_file in sorted(backup_dir.glob("*.db"), key=lambda item: item.stat().st_mtime, reverse=True):
        stat = backup_file.stat()
        backups.append({
            "filename": backup_file.name,
            "type": _get_backup_type(backup_file.name),
            "size_bytes": stat.st_size,
            "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })

    return {
        "items": backups,
        "directory": str(backup_dir),
    }


@app.get("/api/admin/system/backups/{filename}")
def download_saved_backup(filename: str, current_user=Depends(get_current_user)):
    _assert_admin(current_user)
    backup_file = _resolve_backup_file_or_404(filename)
    return FileResponse(
        path=str(backup_file),
        media_type="application/octet-stream",
        filename=backup_file.name,
    )


@app.post("/api/admin/system/backups/{filename}/restore")
def restore_saved_backup(filename: str, current_user=Depends(get_current_user)):
    _assert_admin(current_user)

    db_path = _sqlite_db_path_or_400()
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="Banco de dados atual não encontrado")

    backup_file = _resolve_backup_file_or_404(filename)

    try:
        with sqlite3.connect(str(backup_file)) as conn_test:
            integrity = conn_test.execute("PRAGMA integrity_check;").fetchone()
            if not integrity or str(integrity[0]).lower() != "ok":
                raise HTTPException(status_code=400, detail="Arquivo de backup inválido (integridade SQLite falhou)")

        backup_before_restore = _create_sqlite_backup_file(db_path, prefix="before_restore")

        try:
            engine.dispose()
            shutil.copy2(str(backup_file), str(db_path))

            with sqlite3.connect(str(db_path)) as conn_new:
                conn_new.execute("SELECT name FROM sqlite_master LIMIT 1;")

            _ensure_financeiro_columns_and_seed_contas()
        except Exception as restore_error:
            shutil.copy2(str(backup_before_restore), str(db_path))
            engine.dispose()
            raise HTTPException(status_code=500, detail=f"Falha ao restaurar backup: {str(restore_error)}")

        return {
            "ok": True,
            "detail": "Backup restaurado com sucesso",
            "backup_restaurado": backup_file.name,
            "backup_anterior": backup_before_restore.name,
        }
    finally:
        engine.dispose()


@app.delete("/api/admin/system/backups/{filename}")
def delete_saved_backup(filename: str, current_user=Depends(get_current_user)):
    _assert_admin(current_user)
    backup_file = _resolve_backup_file_or_404(filename)
    backup_file.unlink(missing_ok=True)
    return {"ok": True, "detail": "Backup removido com sucesso"}


@app.post("/api/admin/system/restore")
async def restore_database(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user)
):
    _assert_admin(current_user)

    db_path = _sqlite_db_path_or_400()
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="Banco de dados atual não encontrado")

    nome_arquivo = (file.filename or "").lower()
    if not nome_arquivo.endswith(".db"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .db válido")

    temp_upload = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".db") as tmp:
            temp_upload = Path(tmp.name)
            content = await file.read()
            tmp.write(content)

        with sqlite3.connect(str(temp_upload)) as conn_test:
            integrity = conn_test.execute("PRAGMA integrity_check;").fetchone()
            if not integrity or str(integrity[0]).lower() != "ok":
                raise HTTPException(status_code=400, detail="Arquivo de backup inválido (integridade SQLite falhou)")

        backup_before_restore = _create_sqlite_backup_file(db_path, prefix="before_restore")

        try:
            engine.dispose()
            shutil.copy2(str(temp_upload), str(db_path))

            with sqlite3.connect(str(db_path)) as conn_new:
                conn_new.execute("SELECT name FROM sqlite_master LIMIT 1;")

            _ensure_financeiro_columns_and_seed_contas()
        except Exception as restore_error:
            shutil.copy2(str(backup_before_restore), str(db_path))
            engine.dispose()
            raise HTTPException(status_code=500, detail=f"Falha ao restaurar backup: {str(restore_error)}")

        return {
            "ok": True,
            "detail": "Backup restaurado com sucesso",
            "backup_anterior": backup_before_restore.name,
        }
    finally:
        try:
            if temp_upload and temp_upload.exists():
                temp_upload.unlink(missing_ok=True)
        except Exception:
            pass
        await file.close()

# ════════════════════════════════════════════════════════════════════════════════
# AUTH
# ════════════════════════════════════════════════════════════════════════════════
@app.post("/api/auth/login", response_model=schemas.TokenResponse)
def login(req: schemas.LoginRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == req.email).first()
    if not user or not verify_password(req.password, user.password):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    token = create_access_token({"sub": user.id})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user.id, "email": user.email, "nome_completo": user.nome_completo, "role": user.role}
    }

@app.get("/api/auth/me")
def me(current_user: models.User = Depends(get_current_user)):
    return {"id": current_user.id, "email": current_user.email,
            "nome_completo": current_user.nome_completo, "role": current_user.role}


@app.get("/api/users/me", response_model=schemas.UserResponse)
def get_own_user(current_user=Depends(get_current_user)):
    return current_user


@app.put("/api/users/me", response_model=schemas.UserResponse)
def update_own_user(req: schemas.UserSelfUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    payload = req.dict(exclude_none=True)
    if not payload:
        raise HTTPException(status_code=400, detail="Nenhum campo informado para atualização")

    if "nome_completo" in payload:
        nome = (payload.get("nome_completo") or "").strip()
        if not nome:
            raise HTTPException(status_code=400, detail="Nome completo é obrigatório")
        current_user.nome_completo = nome

    if "password" in payload:
        password = payload.get("password") or ""
        current_password = payload.get("current_password") or ""
        if not current_password:
            raise HTTPException(status_code=400, detail="Informe a senha atual para alterar a senha")
        if not verify_password(current_password, current_user.password):
            raise HTTPException(status_code=400, detail="Senha atual inválida")
        _validate_password_strength(password)
        if verify_password(password, current_user.password):
            raise HTTPException(status_code=400, detail="A nova senha deve ser diferente da senha atual")
        current_user.password = get_password_hash(password)

    db.commit()
    db.refresh(current_user)
    return current_user


# ════════════════════════════════════════════════════════════════════════════════
# USERS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/users", response_model=List[schemas.UserResponse])
def list_users(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role == "administrador":
        return db.query(models.User).all()
    return [current_user]

@app.post("/api/users", response_model=schemas.UserResponse)
def create_user(req: schemas.UserCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role != "administrador":
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar usuários")
    existing = db.query(models.User).filter(models.User.email == req.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    _validate_password_strength(req.password)
    user = models.User(
        id=str(uuid.uuid4()),
        email=req.email,
        nome_completo=req.nome_completo,
        role=req.role,
        password=get_password_hash(req.password),
        ativo=True,
        created_at=datetime.utcnow()
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

@app.put("/api/users/{user_id}", response_model=schemas.UserResponse)
def update_user(user_id: str, req: schemas.UserUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role != "administrador":
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        if k == "password":
            _validate_password_strength(v)
            setattr(user, k, get_password_hash(v))
        else:
            setattr(user, k, v)
    db.commit()
    db.refresh(user)
    return user

@app.delete("/api/users/{user_id}")
def delete_user(user_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role != "administrador":
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores")
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Você não pode remover seu próprio usuário")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    db.delete(user)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# MEMBROS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/membros", response_model=List[schemas.MembroResponse])
def list_membros(
    skip: int = 0, limit: int = 500,
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Membro)
    if search:
        q = q.filter(or_(
            models.Membro.nome_completo.ilike(f"%{search}%"),
            models.Membro.cpf.ilike(f"%{search}%"),
            models.Membro.matricula.ilike(f"%{search}%"),
            models.Membro.email.ilike(f"%{search}%"),
        ))
    if status:
        q = q.filter(models.Membro.status == status)
    return q.order_by(models.Membro.nome_completo).offset(skip).limit(limit).all()

@app.post("/api/membros", response_model=schemas.MembroResponse)
def create_membro(req: schemas.MembroCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    membro = models.Membro(id=str(uuid.uuid4()), **req.dict(), created_at=datetime.utcnow(), updated_at=datetime.utcnow())
    db.add(membro)
    db.commit()
    db.refresh(membro)
    return membro

@app.get("/api/membros/{membro_id}", response_model=schemas.MembroResponse)
def get_membro(membro_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    m = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    return m

@app.put("/api/membros/{membro_id}", response_model=schemas.MembroResponse)
def update_membro(membro_id: str, req: schemas.MembroUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    m = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        setattr(m, k, v)
    m.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(m)
    return m

@app.delete("/api/membros/{membro_id}")
def delete_membro(membro_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    m = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    db.delete(m)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# PAGAMENTOS
# ════════════════════════════════════════════════════════════════════════════════
def _pagamentos_por_membro_no_mes(db: Session, mes_referencia: str):
    pagamentos = db.query(models.Pagamento).filter(
        models.Pagamento.mes_referencia == mes_referencia
    ).all()

    ultimo_por_membro = {}
    for pagamento in pagamentos:
        if not pagamento.membro_id:
            continue

        atual = ultimo_por_membro.get(pagamento.membro_id)
        if not atual:
            ultimo_por_membro[pagamento.membro_id] = pagamento
            continue

        dt_pagamento = pagamento.updated_at or pagamento.created_at or datetime.min
        dt_atual = atual.updated_at or atual.created_at or datetime.min
        if dt_pagamento >= dt_atual:
            ultimo_por_membro[pagamento.membro_id] = pagamento

    return ultimo_por_membro


def _pagamentos_pagos_membros_ativos_no_mes(db: Session, mes_referencia: str):
    membros_ativos_ids = {
        mid for (mid,) in db.query(models.Membro.id).filter(models.Membro.status == 'ativo').all()
    }
    pagamentos_mes = _pagamentos_por_membro_no_mes(db, mes_referencia)
    return [
        pagamento
        for membro_id, pagamento in pagamentos_mes.items()
        if membro_id in membros_ativos_ids and pagamento.status_pagamento == 'pago'
    ]


def _normalizar_texto(texto: Optional[str]) -> str:
    if not texto:
        return ""

    sem_acentos = "".join(
        c for c in unicodedata.normalize("NFKD", str(texto)) if not unicodedata.combining(c)
    )
    return re.sub(r"\s+", " ", sem_acentos).strip().lower()


def _pontuacao_nome_no_extrato(nome: Optional[str], descricao_extrato: Optional[str]) -> int:
    nome_norm = _normalizar_texto(nome)
    desc_norm = _normalizar_texto(descricao_extrato)
    if not nome_norm or not desc_norm:
        return 0

    tokens = [token for token in re.split(r"\W+", nome_norm) if len(token) >= 3]
    if not tokens:
        return 0

    return sum(1 for token in tokens if token in desc_norm)


def _somente_digitos(texto: Optional[str]) -> str:
    if not texto:
        return ""
    return "".join(ch for ch in str(texto) if ch.isdigit())


def _pontuacao_match_membro_extrato(
    membro_nome: Optional[str],
    membro_cpf: Optional[str],
    membro_matricula: Optional[str],
    descricao_extrato: Optional[str]
) -> dict:
    descricao_norm = _normalizar_texto(descricao_extrato)
    descricao_digitos = _somente_digitos(descricao_extrato)

    cpf_digitos = _somente_digitos(membro_cpf)
    matricula_digitos = _somente_digitos(membro_matricula)

    score_nome_tokens = _pontuacao_nome_no_extrato(membro_nome, descricao_extrato)
    nome_norm = _normalizar_texto(membro_nome)
    nome_completo_match = bool(nome_norm and descricao_norm and nome_norm in descricao_norm)

    cpf_match = bool(cpf_digitos and len(cpf_digitos) >= 6 and cpf_digitos in descricao_digitos)
    matricula_match = bool(matricula_digitos and len(matricula_digitos) >= 4 and matricula_digitos in descricao_digitos)

    score = 0
    if cpf_match:
        score += 100
    if matricula_match:
        score += 80
    if nome_completo_match:
        score += 40
    score += min(score_nome_tokens, 4) * 10

    if cpf_match or matricula_match or (nome_completo_match and score_nome_tokens >= 2):
        confianca = "alta"
    elif score_nome_tokens >= 2:
        confianca = "media"
    else:
        confianca = "baixa"

    return {
        "score": score,
        "confianca": confianca,
        "score_nome_tokens": score_nome_tokens,
        "nome_completo_match": nome_completo_match,
        "cpf_match": cpf_match,
        "matricula_match": matricula_match
    }


def _status_pagamento_pendente_ou_atrasado(status_pagamento: Optional[str]) -> bool:
    return status_pagamento in (None, "", "pendente", "atrasado")


@app.get("/api/pagamentos")
def list_pagamentos(
    mes_referencia: Optional[str] = None,
    membro_id: Optional[str] = None,
    status_pagamento: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Pagamento)
    if mes_referencia:
        q = q.filter(models.Pagamento.mes_referencia == mes_referencia)
    if membro_id:
        q = q.filter(models.Pagamento.membro_id == membro_id)
    if status_pagamento:
        q = q.filter(models.Pagamento.status_pagamento == status_pagamento)
    pagamentos = q.all()
    result = []
    for p in pagamentos:
        pd = {
            "id": p.id, "membro_id": p.membro_id, "valor_pago": float(p.valor_pago) if p.valor_pago else 0,
            "mes_referencia": p.mes_referencia, "data_pagamento": str(p.data_pagamento) if p.data_pagamento else None,
            "status_pagamento": p.status_pagamento, "forma_pagamento": p.forma_pagamento,
            "observacoes": p.observacoes, "created_at": str(p.created_at) if p.created_at else None,
            "membro_nome": None
        }
        if p.membro_id:
            m = db.query(models.Membro).filter(models.Membro.id == p.membro_id).first()
            if m:
                pd["membro_nome"] = m.nome_completo
        result.append(pd)
    return result

@app.get("/api/pagamentos/painel")
def painel_pagamentos(
    mes_referencia: str,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Membro).filter(models.Membro.status == 'ativo')

    if search:
        q = q.filter(
            or_(
                models.Membro.nome_completo.ilike(f"%{search}%"),
                models.Membro.cpf.ilike(f"%{search}%"),
                models.Membro.email.ilike(f"%{search}%"),
                models.Membro.cidade.ilike(f"%{search}%"),
                models.Membro.matricula.ilike(f"%{search}%"),
            )
        )

    membros = q.all()
    pagamentos = _pagamentos_por_membro_no_mes(db, mes_referencia)
    result = []
    for m in membros:
        p = pagamentos.get(m.id)
        result.append({
            "membro_id": m.id,
            "nome": m.nome_completo,
            "matricula": m.matricula,
            "valor_mensalidade": float(m.valor_mensalidade) if m.valor_mensalidade else 0,
            "pagamento_id": p.id if p else None,
            "valor_pago": float(p.valor_pago) if p and p.valor_pago else 0,
            "data_pagamento": str(p.data_pagamento) if p and p.data_pagamento else None,
            "status": p.status_pagamento if p else "pendente",
            "forma_pagamento": p.forma_pagamento if p else None,
        })
    return result


@app.post("/api/pagamentos/baixa-automatica-banco")
def baixa_automatica_pagamentos_banco(
    mes_referencia: str,
    tolerancia_valor: float = 0.01,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    if not mes_referencia or not re.match(r"^\d{4}-\d{2}$", mes_referencia):
        raise HTTPException(status_code=400, detail="mes_referencia deve estar no formato YYYY-MM")

    pagamentos_mes = _pagamentos_por_membro_no_mes(db, mes_referencia)

    membros_ativos = db.query(models.Membro).filter(models.Membro.status == "ativo").all()
    membros_disponiveis = []
    for membro in membros_ativos:
        valor_mensalidade = float(membro.valor_mensalidade or 0)
        if valor_mensalidade <= 0:
            continue

        pagamento_atual = pagamentos_mes.get(membro.id)
        if pagamento_atual and pagamento_atual.status_pagamento == "pago":
            continue

        membros_disponiveis.append({
            "id": membro.id,
            "nome": membro.nome_completo,
            "cpf": membro.cpf,
            "matricula": membro.matricula,
            "valor_mensalidade": valor_mensalidade,
            "pagamento_atual": pagamento_atual
        })

    conciliacoes_abertas = db.query(models.Conciliacao).filter(
        models.Conciliacao.mes_referencia == mes_referencia,
        models.Conciliacao.tipo == "credito",
        models.Conciliacao.conciliado == False
    ).order_by(models.Conciliacao.data_extrato.asc()).all()

    total_analisados = len(conciliacoes_abertas)
    total_baixados = 0
    total_sem_match = 0
    total_ambiguos = 0
    detalhes = []
    membros_ja_baixados = set()

    for conciliacao in conciliacoes_abertas:
        valor_extrato = float(conciliacao.valor_extrato or 0)
        descricao_extrato = conciliacao.descricao_extrato or ""

        candidatos = []
        for membro in membros_disponiveis:
            if membro["id"] in membros_ja_baixados:
                continue

            if abs(valor_extrato - membro["valor_mensalidade"]) > tolerancia_valor:
                continue

            metrica = _pontuacao_match_membro_extrato(
                membro_nome=membro["nome"],
                membro_cpf=membro.get("cpf"),
                membro_matricula=membro.get("matricula"),
                descricao_extrato=descricao_extrato
            )

            if metrica["confianca"] == "baixa":
                continue

            candidatos.append((membro, metrica))

        if not candidatos:
            total_sem_match += 1
            continue

        candidatos.sort(key=lambda item: item[1]["score"], reverse=True)
        melhor_score = candidatos[0][1]["score"]
        melhores = [item for item in candidatos if item[1]["score"] == melhor_score]

        if len(melhores) != 1:
            total_sem_match += 1
            total_ambiguos += 1
            continue

        if len(candidatos) > 1:
            segundo_score = candidatos[1][1]["score"]
            gap = melhor_score - segundo_score
            if gap < 15 and not (
                melhores[0][1]["cpf_match"] or melhores[0][1]["matricula_match"]
            ):
                total_sem_match += 1
                total_ambiguos += 1
                continue

        if melhor_score < 20:
            total_sem_match += 1
            continue

        membro_match = melhores[0][0]
        metrica_match = melhores[0][1]
        pagamento = membro_match["pagamento_atual"]

        if pagamento:
            pagamento.valor_pago = valor_extrato
            pagamento.status_pagamento = "pago"
            pagamento.data_pagamento = conciliacao.data_extrato
            pagamento.forma_pagamento = "transferencia"
            pagamento.observacoes = (
                (pagamento.observacoes + "\n") if pagamento.observacoes else ""
            ) + f"Baixa automática via extrato bancário ({mes_referencia}) - confiança {metrica_match['confianca']}"
            pagamento.updated_at = datetime.utcnow()
        else:
            pagamento = models.Pagamento(
                id=str(uuid.uuid4()),
                membro_id=membro_match["id"],
                valor_pago=valor_extrato,
                mes_referencia=mes_referencia,
                data_pagamento=conciliacao.data_extrato,
                status_pagamento="pago",
                forma_pagamento="transferencia",
                observacoes=f"Baixa automática via extrato bancário ({mes_referencia}) - confiança {metrica_match['confianca']}",
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow()
            )
            db.add(pagamento)
            db.flush()

        conciliacao.pagamento_id = pagamento.id
        conciliacao.conciliado = True
        conciliacao.updated_at = datetime.utcnow()

        _register_transaction(db, pagamento, current_user.id)

        membros_ja_baixados.add(membro_match["id"])
        total_baixados += 1
        detalhes.append({
            "conciliacao_id": conciliacao.id,
            "membro_id": membro_match["id"],
            "membro_nome": membro_match["nome"],
            "valor": valor_extrato,
            "data_extrato": str(conciliacao.data_extrato),
            "confianca": metrica_match["confianca"],
            "criterios": {
                "cpf_match": metrica_match["cpf_match"],
                "matricula_match": metrica_match["matricula_match"],
                "nome_completo_match": metrica_match["nome_completo_match"],
                "score_nome_tokens": metrica_match["score_nome_tokens"],
                "score_total": metrica_match["score"]
            }
        })

    db.commit()

    return {
        "ok": True,
        "mes_referencia": mes_referencia,
        "total_analisados": total_analisados,
        "total_baixados": total_baixados,
        "total_sem_match": total_sem_match,
        "total_ambiguos": total_ambiguos,
        "detalhes": detalhes
    }


@app.get("/api/pagamentos/pendencias-conciliacao-manual")
def listar_pendencias_conciliacao_manual(
    mes_referencia: str,
    tolerancia_valor: float = 0.01,
    limite_candidatos: int = 3,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    if not mes_referencia or not re.match(r"^\d{4}-\d{2}$", mes_referencia):
        raise HTTPException(status_code=400, detail="mes_referencia deve estar no formato YYYY-MM")

    pagamentos_mes = _pagamentos_por_membro_no_mes(db, mes_referencia)
    membros_ativos = db.query(models.Membro).filter(models.Membro.status == "ativo").all()

    membros_disponiveis = []
    for membro in membros_ativos:
        valor_mensalidade = float(membro.valor_mensalidade or 0)
        if valor_mensalidade <= 0:
            continue

        pagamento_atual = pagamentos_mes.get(membro.id)
        if pagamento_atual and pagamento_atual.status_pagamento == "pago":
            continue

        membros_disponiveis.append({
            "id": membro.id,
            "nome": membro.nome_completo,
            "cpf": membro.cpf,
            "matricula": membro.matricula,
            "valor_mensalidade": valor_mensalidade,
            "pagamento_atual": pagamento_atual
        })

    conciliacoes_abertas = db.query(models.Conciliacao).filter(
        models.Conciliacao.mes_referencia == mes_referencia,
        models.Conciliacao.tipo == "credito",
        models.Conciliacao.conciliado == False
    ).order_by(models.Conciliacao.data_extrato.asc()).all()

    pendencias = []

    for conciliacao in conciliacoes_abertas:
        valor_extrato = float(conciliacao.valor_extrato or 0)
        descricao_extrato = conciliacao.descricao_extrato or ""

        candidatos = []
        for membro in membros_disponiveis:
            if abs(valor_extrato - membro["valor_mensalidade"]) > tolerancia_valor:
                continue

            metrica = _pontuacao_match_membro_extrato(
                membro_nome=membro["nome"],
                membro_cpf=membro.get("cpf"),
                membro_matricula=membro.get("matricula"),
                descricao_extrato=descricao_extrato
            )

            pagamento_atual = membro.get("pagamento_atual")
            candidatos.append({
                "membro_id": membro["id"],
                "nome": membro["nome"],
                "matricula": membro.get("matricula"),
                "valor_mensalidade": membro["valor_mensalidade"],
                "pagamento_id": pagamento_atual.id if pagamento_atual else None,
                "status_pagamento": pagamento_atual.status_pagamento if pagamento_atual else "pendente",
                "score": metrica["score"],
                "confianca": metrica["confianca"],
                "cpf_match": metrica["cpf_match"],
                "matricula_match": metrica["matricula_match"],
                "nome_match": metrica["nome_completo_match"]
            })

        if not candidatos:
            continue

        candidatos.sort(
            key=lambda c: (
                c["score"],
                1 if c["cpf_match"] else 0,
                1 if c["matricula_match"] else 0,
                1 if c["nome_match"] else 0
            ),
            reverse=True
        )

        melhores = candidatos[:max(1, min(limite_candidatos, 10))]

        pendencias.append({
            "conciliacao_id": conciliacao.id,
            "data_extrato": str(conciliacao.data_extrato) if conciliacao.data_extrato else None,
            "descricao_extrato": conciliacao.descricao_extrato,
            "valor_extrato": valor_extrato,
            "banco": conciliacao.banco,
            "numero_documento": conciliacao.numero_documento,
            "candidatos": melhores
        })

    return {
        "ok": True,
        "mes_referencia": mes_referencia,
        "total_pendencias": len(pendencias),
        "pendencias": pendencias
    }


@app.post("/api/pagamentos/reprocessar-dabb")
def reprocessar_pendencias_dabb(
    mes_referencia: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    if not mes_referencia or not re.match(r"^\d{4}-\d{2}$", mes_referencia):
        raise HTTPException(status_code=400, detail="mes_referencia deve estar no formato YYYY-MM")

    membros_por_codigo_dabb = _indexar_membros_por_codigo_dabb(db)
    conciliacoes_abertas = db.query(models.Conciliacao).filter(
        models.Conciliacao.mes_referencia == mes_referencia,
        models.Conciliacao.tipo == "credito",
        models.Conciliacao.conciliado == False,
        models.Conciliacao.observacoes.isnot(None),
        models.Conciliacao.observacoes.like("Arquivo DABB%")
    ).order_by(models.Conciliacao.data_extrato.asc()).all()

    total_analisados = len(conciliacoes_abertas)
    total_reprocessados = 0
    total_sem_match = 0
    total_ambiguos = 0
    detalhes = []

    for conciliacao in conciliacoes_abertas:
        codigo_dabb = _extrair_codigo_dabb_das_observacoes(conciliacao.observacoes)
        if not codigo_dabb:
            total_sem_match += 1
            continue

        membros_match = {}
        for variante in _variantes_codigo_dabb(codigo_dabb):
            for membro in membros_por_codigo_dabb.get(variante, []):
                membros_match[membro.id] = membro

        if len(membros_match) == 1:
            membro = next(iter(membros_match.values()))
            _baixar_pagamento_mensalidade_por_conciliacao(
                db=db,
                conciliacao=conciliacao,
                membro=membro,
                user_id=current_user.id,
                observacao_origem=(
                    f"Baixa reprocessada via codigo DABB ({mes_referencia}) - "
                    f"codigo_dabb {codigo_dabb}"
                ),
            )
            total_reprocessados += 1
            detalhes.append({
                "conciliacao_id": conciliacao.id,
                "membro_id": membro.id,
                "membro_nome": membro.nome_completo,
                "codigo_dabb": codigo_dabb,
                "valor": float(conciliacao.valor_extrato or 0),
                "data_extrato": str(conciliacao.data_extrato) if conciliacao.data_extrato else None,
            })
        elif len(membros_match) > 1:
            total_ambiguos += 1
        else:
            total_sem_match += 1

    db.commit()

    return {
        "ok": True,
        "mes_referencia": mes_referencia,
        "total_analisados": total_analisados,
        "total_reprocessados": total_reprocessados,
        "total_sem_match": total_sem_match,
        "total_ambiguos": total_ambiguos,
        "detalhes": detalhes,
    }


@app.post("/api/pagamentos/pendencias-conciliacao-manual/confirmar")
def confirmar_pendencia_conciliacao_manual(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    conciliacao_id = payload.get("conciliacao_id")
    membro_id = payload.get("membro_id")

    if not conciliacao_id or not membro_id:
        raise HTTPException(status_code=400, detail="conciliacao_id e membro_id são obrigatórios")

    conciliacao = db.query(models.Conciliacao).filter(models.Conciliacao.id == conciliacao_id).first()
    if not conciliacao:
        raise HTTPException(status_code=404, detail="Lançamento de conciliação não encontrado")

    if conciliacao.conciliado:
        raise HTTPException(status_code=400, detail="Este lançamento já está conciliado")

    if conciliacao.tipo != "credito":
        raise HTTPException(status_code=400, detail="Apenas lançamentos de crédito podem dar baixa em mensalidade")

    mes_ref = conciliacao.mes_referencia or (
        conciliacao.data_extrato.strftime("%Y-%m") if conciliacao.data_extrato else None
    )
    if not mes_ref:
        raise HTTPException(status_code=400, detail="Não foi possível identificar o mês de referência")

    pagamento = db.query(models.Pagamento).filter(
        models.Pagamento.membro_id == membro_id,
        models.Pagamento.mes_referencia == mes_ref
    ).first()

    valor_extrato = float(conciliacao.valor_extrato or 0)

    if pagamento:
        pagamento.valor_pago = valor_extrato
        pagamento.status_pagamento = "pago"
        pagamento.data_pagamento = conciliacao.data_extrato
        pagamento.forma_pagamento = "transferencia"
        pagamento.observacoes = (
            (pagamento.observacoes + "\n") if pagamento.observacoes else ""
        ) + f"Baixa manual via pendência de conciliação ({mes_ref})"
        pagamento.updated_at = datetime.utcnow()
    else:
        pagamento = models.Pagamento(
            id=str(uuid.uuid4()),
            membro_id=membro_id,
            valor_pago=valor_extrato,
            mes_referencia=mes_ref,
            data_pagamento=conciliacao.data_extrato,
            status_pagamento="pago",
            forma_pagamento="transferencia",
            observacoes=f"Baixa manual via pendência de conciliação ({mes_ref})",
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        db.add(pagamento)
        db.flush()

    conciliacao.pagamento_id = pagamento.id
    conciliacao.conciliado = True
    conciliacao.updated_at = datetime.utcnow()

    _register_transaction(db, pagamento, current_user.id)

    db.commit()

    membro = db.query(models.Membro).filter(models.Membro.id == membro_id).first()

    return {
        "ok": True,
        "detail": "Baixa manual realizada com sucesso",
        "conciliacao_id": conciliacao.id,
        "pagamento_id": pagamento.id,
        "membro_nome": membro.nome_completo if membro else None,
        "mes_referencia": mes_ref
    }

@app.post("/api/pagamentos", response_model=schemas.PagamentoResponse)
def create_pagamento(req: schemas.PagamentoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    # Check if payment already exists
    existing = db.query(models.Pagamento).filter(
        models.Pagamento.membro_id == req.membro_id,
        models.Pagamento.mes_referencia == req.mes_referencia
    ).first()
    if existing:
        for k, v in req.dict(exclude_none=True).items():
            setattr(existing, k, v)
        existing.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(existing)
        # Register transaction
        _register_transaction(db, existing, current_user.id)
        return existing
    
    p = models.Pagamento(
        id=str(uuid.uuid4()),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        **req.dict()
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    _register_transaction(db, p, current_user.id)
    return p

def _register_transaction(db, pagamento, user_id):
    m = db.query(models.Membro).filter(models.Membro.id == pagamento.membro_id).first()
    nome = m.nome_completo if m else "Membro"
    # Check if transaction exists for this pagamento
    existing = db.query(models.Transacao).filter(
        models.Transacao.origem == "mensalidade",
        models.Transacao.membro_id == pagamento.membro_id,
        cast(models.Transacao.categoria, String).ilike(f"%{pagamento.mes_referencia}%")
    ).first()
    if not existing:
        t = models.Transacao(
            id=str(uuid.uuid4()),
            user_id=user_id,
            descricao=f"Mensalidade {nome} - {pagamento.mes_referencia}",
            valor=pagamento.valor_pago,
            tipo="entrada",
            categoria=f"Mensalidade {pagamento.mes_referencia}",
            data_transacao=pagamento.data_pagamento or date.today(),
            origem="mensalidade",
            membro_id=pagamento.membro_id,
            created_at=datetime.utcnow()
        )
        db.add(t)
        db.commit()


def _delete_transacoes_despesa_por_payload(db: Session, payload: dict):
    transacoes = db.query(models.Transacao).filter(
        models.Transacao.tipo == "saida",
        models.Transacao.descricao == payload.get("descricao"),
        models.Transacao.categoria == payload.get("categoria"),
        models.Transacao.data_transacao == payload.get("data_despesa"),
        models.Transacao.valor == payload.get("valor"),
        models.Transacao.origem == "despesa"
    ).all()
    for t in transacoes:
        db.delete(t)


def _sync_transacao_despesa(db: Session, despesa, user_id: str):
    transacoes = db.query(models.Transacao).filter(
        models.Transacao.tipo == "saida",
        models.Transacao.descricao == despesa.descricao,
        models.Transacao.categoria == despesa.categoria,
        models.Transacao.data_transacao == despesa.data_despesa,
        models.Transacao.valor == despesa.valor,
        models.Transacao.origem == "despesa"
    ).all()

    if transacoes:
        principal = transacoes[0]
        principal.user_id = user_id
        principal.descricao = despesa.descricao
        principal.valor = despesa.valor
        principal.tipo = "saida"
        principal.categoria = despesa.categoria
        principal.data_transacao = despesa.data_despesa
        principal.origem = "despesa"
        principal.updated_at = datetime.utcnow()
        for duplicada in transacoes[1:]:
            db.delete(duplicada)
        return

    db.add(models.Transacao(
        id=str(uuid.uuid4()),
        user_id=user_id,
        descricao=despesa.descricao,
        valor=despesa.valor,
        tipo="saida",
        categoria=despesa.categoria,
        data_transacao=despesa.data_despesa,
        origem="despesa",
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow()
    ))


def _delete_transacoes_outra_renda_por_payload(db: Session, payload: dict):
    origens_possiveis = ["outra_renda"]
    if payload.get("categoria"):
        origens_possiveis.append(payload["categoria"])

    transacoes = db.query(models.Transacao).filter(
        models.Transacao.tipo == "entrada",
        models.Transacao.descricao == payload.get("descricao"),
        models.Transacao.categoria == payload.get("categoria"),
        models.Transacao.data_transacao == payload.get("data_recebimento"),
        models.Transacao.valor == payload.get("valor"),
        models.Transacao.origem.in_(origens_possiveis)
    ).all()
    for t in transacoes:
        db.delete(t)


def _sync_transacao_outra_renda(db: Session, renda, user_id: str):
    origens_possiveis = ["outra_renda"]
    if renda.categoria:
        origens_possiveis.append(renda.categoria)

    transacoes = db.query(models.Transacao).filter(
        models.Transacao.tipo == "entrada",
        models.Transacao.descricao == renda.descricao,
        models.Transacao.categoria == renda.categoria,
        models.Transacao.data_transacao == renda.data_recebimento,
        models.Transacao.valor == renda.valor,
        models.Transacao.origem.in_(origens_possiveis)
    ).all()

    if transacoes:
        principal = transacoes[0]
        principal.user_id = user_id
        principal.descricao = renda.descricao
        principal.valor = renda.valor
        principal.tipo = "entrada"
        principal.categoria = renda.categoria
        principal.data_transacao = renda.data_recebimento
        principal.origem = "outra_renda"
        principal.updated_at = datetime.utcnow()
        for duplicada in transacoes[1:]:
            db.delete(duplicada)
        return

    db.add(models.Transacao(
        id=str(uuid.uuid4()),
        user_id=user_id,
        descricao=renda.descricao,
        valor=renda.valor,
        tipo="entrada",
        categoria=renda.categoria,
        data_transacao=renda.data_recebimento,
        origem="outra_renda",
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow()
    ))

@app.put("/api/pagamentos/{pagamento_id}", response_model=schemas.PagamentoResponse)
def update_pagamento(pagamento_id: str, req: schemas.PagamentoUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.Pagamento).filter(models.Pagamento.id == pagamento_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        setattr(p, k, v)
    p.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(p)
    return p

@app.delete("/api/pagamentos/{pagamento_id}")
def delete_pagamento(pagamento_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.Pagamento).filter(models.Pagamento.id == pagamento_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    db.delete(p)
    db.commit()
    return {"ok": True}


def _get_conta_or_400(db: Session, conta_id: Optional[str], tipo: str) -> models.PlanoConta:
    if not conta_id:
        raise HTTPException(status_code=400, detail="Selecione uma conta")

    conta = db.query(models.PlanoConta).filter(models.PlanoConta.id == conta_id).first()
    if not conta:
        raise HTTPException(status_code=400, detail="Conta não encontrada")

    if (conta.tipo or "").lower() != (tipo or "").lower():
        raise HTTPException(status_code=400, detail=f"Conta inválida para {tipo}")

    if conta.ativo is False:
        raise HTTPException(status_code=400, detail="Conta inativa")

    return conta


@app.get("/api/contas", response_model=List[schemas.PlanoContaResponse])
def list_contas(
    tipo: Optional[str] = None,
    apenas_ativas: bool = True,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.PlanoConta)
    if tipo:
        q = q.filter(models.PlanoConta.tipo == tipo)
    if apenas_ativas:
        q = q.filter(models.PlanoConta.ativo == True)
    return q.order_by(models.PlanoConta.ordem.asc(), models.PlanoConta.codigo.asc()).all()


def _validar_tipo_conta(tipo: Optional[str]) -> str:
    tipo_norm = unicodedata.normalize("NFKD", (tipo or "")).encode("ascii", "ignore").decode().strip().lower()
    if tipo_norm in {"entrada", "entradas"}:
        return "entrada"
    if tipo_norm in {"saida", "saidas"}:
        return "saida"
    if tipo_norm not in {"entrada", "saida"}:
        raise HTTPException(status_code=400, detail="Tipo da conta deve ser 'entrada' ou 'saida'")
    return tipo_norm


def _ordenar_codigo_conta(codigo: Optional[str]):
    valor = (codigo or "").strip()
    if not valor:
        return (9999,)

    out = []
    for parte in valor.split("."):
        try:
            out.append(int(parte))
        except Exception:
            out.append(9999)
    return tuple(out)


@app.post("/api/contas", response_model=schemas.PlanoContaResponse)
def create_conta(req: schemas.PlanoContaCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    codigo = (req.codigo or "").strip()
    if not codigo:
        raise HTTPException(status_code=400, detail="Código é obrigatório")

    nome = (req.nome or "").strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome é obrigatório")

    tipo = _validar_tipo_conta(req.tipo)
    existe_codigo = db.query(models.PlanoConta).filter(models.PlanoConta.codigo == codigo).first()
    if existe_codigo:
        raise HTTPException(status_code=400, detail="Já existe uma conta com este código")

    conta = models.PlanoConta(
        id=str(uuid.uuid4()),
        codigo=codigo,
        nome=nome,
        tipo=tipo,
        ordem=req.ordem or 0,
        ativo=True if req.ativo is None else bool(req.ativo),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(conta)
    db.commit()
    db.refresh(conta)
    return conta


@app.put("/api/contas/{conta_id}", response_model=schemas.PlanoContaResponse)
def update_conta(conta_id: str, req: schemas.PlanoContaUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    conta = db.query(models.PlanoConta).filter(models.PlanoConta.id == conta_id).first()
    if not conta:
        raise HTTPException(status_code=404, detail="Conta não encontrada")

    if req.codigo is not None:
        novo_codigo = req.codigo.strip()
        if not novo_codigo:
            raise HTTPException(status_code=400, detail="Código é obrigatório")
        existe_codigo = db.query(models.PlanoConta).filter(
            models.PlanoConta.codigo == novo_codigo,
            models.PlanoConta.id != conta_id
        ).first()
        if existe_codigo:
            raise HTTPException(status_code=400, detail="Já existe uma conta com este código")
        conta.codigo = novo_codigo

    if req.nome is not None:
        novo_nome = req.nome.strip()
        if not novo_nome:
            raise HTTPException(status_code=400, detail="Nome é obrigatório")
        conta.nome = novo_nome

    if req.tipo is not None:
        conta.tipo = _validar_tipo_conta(req.tipo)

    if req.ordem is not None:
        conta.ordem = req.ordem

    if req.ativo is not None:
        conta.ativo = bool(req.ativo)

    conta.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(conta)
    return conta


@app.delete("/api/contas/{conta_id}")
def delete_conta(conta_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    conta = db.query(models.PlanoConta).filter(models.PlanoConta.id == conta_id).first()
    if not conta:
        raise HTTPException(status_code=404, detail="Conta não encontrada")

    despesas_vinculadas = db.query(models.Despesa).filter(models.Despesa.conta_id == conta_id).count()
    rendas_vinculadas = db.query(models.OutraRenda).filter(models.OutraRenda.conta_id == conta_id).count()
    if despesas_vinculadas > 0 or rendas_vinculadas > 0:
        raise HTTPException(
            status_code=400,
            detail="Conta já possui lançamentos vinculados. Inative a conta em vez de excluir."
        )

    db.delete(conta)
    db.commit()
    return {"ok": True}


def _validar_ano_mes_previsao(ano: int, mes: int):
    if ano < 2000 or ano > 2100:
        raise HTTPException(status_code=400, detail="Ano inválido")
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=400, detail="Mês inválido")


@app.get("/api/previsoes-orcamentarias", response_model=List[schemas.PrevisaoOrcamentariaResponse])
def list_previsoes_orcamentarias(
    ano: int,
    mes: Optional[int] = None,
    tipo: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    if mes is not None:
        _validar_ano_mes_previsao(ano, mes)
    else:
        _validar_ano_mes_previsao(ano, 1)

    q = db.query(models.PrevisaoOrcamentaria)
    q = q.filter(models.PrevisaoOrcamentaria.ano == ano)
    if mes is not None:
        q = q.filter(models.PrevisaoOrcamentaria.mes == mes)

    previsoes = q.order_by(models.PrevisaoOrcamentaria.mes.asc()).all()
    conta_ids = [p.conta_id for p in previsoes]
    contas = {
        c.id: c for c in db.query(models.PlanoConta).filter(models.PlanoConta.id.in_(conta_ids)).all()
    } if conta_ids else {}

    result = []
    for p in previsoes:
        conta = contas.get(p.conta_id)
        if tipo and conta and (conta.tipo or "").lower() != tipo.lower():
            continue
        result.append({
            "id": p.id,
            "conta_id": p.conta_id,
            "conta_codigo": conta.codigo if conta else None,
            "conta_nome": conta.nome if conta else None,
            "ano": p.ano,
            "mes": p.mes,
            "valor_previsto": float(p.valor_previsto or 0),
            "observacoes": p.observacoes,
            "created_at": p.created_at,
        })
    return result


@app.post("/api/previsoes-orcamentarias", response_model=schemas.PrevisaoOrcamentariaResponse)
def create_previsao_orcamentaria(
    req: schemas.PrevisaoOrcamentariaCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    _validar_ano_mes_previsao(req.ano, req.mes)
    conta = db.query(models.PlanoConta).filter(models.PlanoConta.id == req.conta_id).first()
    if not conta:
        raise HTTPException(status_code=400, detail="Conta não encontrada")

    existente = db.query(models.PrevisaoOrcamentaria).filter(
        models.PrevisaoOrcamentaria.conta_id == req.conta_id,
        models.PrevisaoOrcamentaria.ano == req.ano,
        models.PrevisaoOrcamentaria.mes == req.mes,
    ).first()
    if existente:
        raise HTTPException(status_code=400, detail="Já existe previsão para esta conta no período")

    previsao = models.PrevisaoOrcamentaria(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        conta_id=req.conta_id,
        ano=req.ano,
        mes=req.mes,
        valor_previsto=float(req.valor_previsto or 0),
        observacoes=req.observacoes,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(previsao)
    db.commit()
    db.refresh(previsao)
    return {
        "id": previsao.id,
        "conta_id": previsao.conta_id,
        "conta_codigo": conta.codigo,
        "conta_nome": conta.nome,
        "ano": previsao.ano,
        "mes": previsao.mes,
        "valor_previsto": float(previsao.valor_previsto or 0),
        "observacoes": previsao.observacoes,
        "created_at": previsao.created_at,
    }


@app.put("/api/previsoes-orcamentarias/{previsao_id}", response_model=schemas.PrevisaoOrcamentariaResponse)
def update_previsao_orcamentaria(
    previsao_id: str,
    req: schemas.PrevisaoOrcamentariaUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    previsao = db.query(models.PrevisaoOrcamentaria).filter(models.PrevisaoOrcamentaria.id == previsao_id).first()
    if not previsao:
        raise HTTPException(status_code=404, detail="Previsão não encontrada")

    if req.valor_previsto is not None:
        previsao.valor_previsto = float(req.valor_previsto)
    if req.observacoes is not None:
        previsao.observacoes = req.observacoes
    previsao.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(previsao)

    conta = db.query(models.PlanoConta).filter(models.PlanoConta.id == previsao.conta_id).first()
    return {
        "id": previsao.id,
        "conta_id": previsao.conta_id,
        "conta_codigo": conta.codigo if conta else None,
        "conta_nome": conta.nome if conta else None,
        "ano": previsao.ano,
        "mes": previsao.mes,
        "valor_previsto": float(previsao.valor_previsto or 0),
        "observacoes": previsao.observacoes,
        "created_at": previsao.created_at,
    }


@app.delete("/api/previsoes-orcamentarias/{previsao_id}")
def delete_previsao_orcamentaria(
    previsao_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    previsao = db.query(models.PrevisaoOrcamentaria).filter(models.PrevisaoOrcamentaria.id == previsao_id).first()
    if not previsao:
        raise HTTPException(status_code=404, detail="Previsão não encontrada")
    db.delete(previsao)
    db.commit()
    return {"ok": True}


@app.post("/api/previsoes-orcamentarias/upsert-lote")
def upsert_previsao_orcamentaria_lote(
    itens: List[schemas.PrevisaoOrcamentariaUpsertItem],
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    atualizados = 0
    criados = 0

    for item in itens:
        _validar_ano_mes_previsao(item.ano, item.mes)
        conta = db.query(models.PlanoConta).filter(models.PlanoConta.id == item.conta_id).first()
        if not conta:
            continue

        previsao = db.query(models.PrevisaoOrcamentaria).filter(
            models.PrevisaoOrcamentaria.conta_id == item.conta_id,
            models.PrevisaoOrcamentaria.ano == item.ano,
            models.PrevisaoOrcamentaria.mes == item.mes,
        ).first()

        if previsao:
            previsao.valor_previsto = float(item.valor_previsto or 0)
            previsao.observacoes = item.observacoes
            previsao.updated_at = datetime.utcnow()
            atualizados += 1
        else:
            db.add(models.PrevisaoOrcamentaria(
                id=str(uuid.uuid4()),
                user_id=current_user.id,
                conta_id=item.conta_id,
                ano=item.ano,
                mes=item.mes,
                valor_previsto=float(item.valor_previsto or 0),
                observacoes=item.observacoes,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            ))
            criados += 1

    db.commit()
    return {"ok": True, "criados": criados, "atualizados": atualizados}


# ════════════════════════════════════════════════════════════════════════════════
# DESPESAS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/despesas", response_model=List[schemas.DespesaResponse])
def list_despesas(
    mes_referencia: Optional[str] = None,
    categoria: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Despesa)
    if mes_referencia:
        q = q.filter(models.Despesa.mes_referencia == mes_referencia)
    if categoria:
        q = q.filter(models.Despesa.categoria == categoria)
    return q.order_by(models.Despesa.data_despesa.desc()).all()

@app.post("/api/despesas", response_model=schemas.DespesaResponse)
def create_despesa(req: schemas.DespesaCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    mes_ref = req.mes_referencia or req.data_despesa.strftime("%Y-%m")
    descricao = (req.descricao or "").strip()
    conta = _get_conta_or_400(db, req.conta_id, "saida")
    categoria = (req.categoria or conta.nome or "Outros").strip()
    d = models.Despesa(id=str(uuid.uuid4()), user_id=current_user.id, mes_referencia=mes_ref,
                        created_at=datetime.utcnow(), **req.dict(exclude={"mes_referencia", "descricao", "categoria", "conta_id"}),
                        descricao=descricao, categoria=categoria,
                        conta_id=conta.id, conta_codigo=conta.codigo, conta_nome=conta.nome)
    d.mes_referencia = mes_ref
    db.add(d)
    _sync_transacao_despesa(db, d, current_user.id)
    db.commit()
    db.refresh(d)
    return d

@app.put("/api/despesas/{despesa_id}", response_model=schemas.DespesaResponse)
def update_despesa(despesa_id: str, req: schemas.DespesaUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    d = db.query(models.Despesa).filter(models.Despesa.id == despesa_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    payload_antigo = {
        "descricao": d.descricao,
        "categoria": d.categoria,
        "data_despesa": d.data_despesa,
        "valor": d.valor,
    }
    for k, v in req.dict(exclude_none=True).items():
        if k == "conta_id":
            continue
        if k == "descricao" and isinstance(v, str):
            v = v.strip()
        setattr(d, k, v)

    if req.conta_id is not None:
        conta = _get_conta_or_400(db, req.conta_id, "saida")
        d.conta_id = conta.id
        d.conta_codigo = conta.codigo
        d.conta_nome = conta.nome
        if not d.categoria:
            d.categoria = conta.nome

    if d.conta_id is None:
        raise HTTPException(status_code=400, detail="Selecione uma conta")

    if not d.categoria:
        d.categoria = d.conta_nome or "Outros"

    _delete_transacoes_despesa_por_payload(db, payload_antigo)
    _sync_transacao_despesa(db, d, current_user.id)
    db.commit()
    db.refresh(d)
    return d

@app.delete("/api/despesas/{despesa_id}")
def delete_despesa(despesa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    d = db.query(models.Despesa).filter(models.Despesa.id == despesa_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    _delete_transacoes_despesa_por_payload(db, {
        "descricao": d.descricao,
        "categoria": d.categoria,
        "data_despesa": d.data_despesa,
        "valor": d.valor,
    })
    db.delete(d)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# OUTRAS RENDAS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/outras-rendas", response_model=List[schemas.OutraRendaResponse])
def list_outras_rendas(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.OutraRenda)
    if mes_referencia:
        q = q.filter(models.OutraRenda.mes_referencia == mes_referencia)
    return q.order_by(models.OutraRenda.data_recebimento.desc()).all()

@app.post("/api/outras-rendas", response_model=schemas.OutraRendaResponse)
def create_outra_renda(req: schemas.OutraRendaCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    mes_ref = req.mes_referencia or req.data_recebimento.strftime("%Y-%m")
    conta = _get_conta_or_400(db, req.conta_id, "entrada")
    categoria = (req.categoria or conta.nome or "Outros").strip()
    r = models.OutraRenda(id=str(uuid.uuid4()), user_id=current_user.id, mes_referencia=mes_ref,
                           created_at=datetime.utcnow(), **req.dict(exclude={"mes_referencia", "categoria", "conta_id"}),
                           categoria=categoria,
                           conta_id=conta.id, conta_codigo=conta.codigo, conta_nome=conta.nome)
    r.mes_referencia = mes_ref
    db.add(r)
    _sync_transacao_outra_renda(db, r, current_user.id)
    db.commit()
    db.refresh(r)
    return r

@app.put("/api/outras-rendas/{renda_id}", response_model=schemas.OutraRendaResponse)
def update_outra_renda(renda_id: str, req: schemas.OutraRendaUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    r = db.query(models.OutraRenda).filter(models.OutraRenda.id == renda_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Renda não encontrada")
    payload_antigo = {
        "descricao": r.descricao,
        "categoria": r.categoria,
        "data_recebimento": r.data_recebimento,
        "valor": r.valor,
    }
    for k, v in req.dict(exclude_none=True).items():
        if k == "conta_id":
            continue
        setattr(r, k, v)

    if req.conta_id is not None:
        conta = _get_conta_or_400(db, req.conta_id, "entrada")
        r.conta_id = conta.id
        r.conta_codigo = conta.codigo
        r.conta_nome = conta.nome
        if not r.categoria:
            r.categoria = conta.nome

    if r.conta_id is None:
        raise HTTPException(status_code=400, detail="Selecione uma conta")

    if not r.categoria:
        r.categoria = r.conta_nome or "Outros"

    _delete_transacoes_outra_renda_por_payload(db, payload_antigo)
    _sync_transacao_outra_renda(db, r, current_user.id)
    db.commit()
    db.refresh(r)
    return r

@app.delete("/api/outras-rendas/{renda_id}")
def delete_outra_renda(renda_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    r = db.query(models.OutraRenda).filter(models.OutraRenda.id == renda_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Renda não encontrada")
    _delete_transacoes_outra_renda_por_payload(db, {
        "descricao": r.descricao,
        "categoria": r.categoria,
        "data_recebimento": r.data_recebimento,
        "valor": r.valor,
    })
    db.delete(r)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# APLICAÇÕES FINANCEIRAS
# ════════════════════════════════════════════════════════════════════════════════
def _calc_saldo_atual_aplicacao(
    saldo_anterior: Optional[float],
    aplicacoes: Optional[float],
    rendimento_bruto: Optional[float],
    impostos: Optional[float],
    resgate: Optional[float]
) -> float:
    saldo_anterior = float(saldo_anterior or 0)
    aplicacoes = float(aplicacoes or 0)
    rendimento_bruto = float(rendimento_bruto or 0)
    impostos = float(impostos or 0)
    resgate = float(resgate or 0)
    return round(saldo_anterior + aplicacoes + rendimento_bruto - impostos - resgate, 2)


def _calc_impostos_aplicacao(imposto_renda: Optional[float], iof: Optional[float], impostos: Optional[float]) -> float:
    total_detalhado = float(imposto_renda or 0) + float(iof or 0)
    if abs(total_detalhado) > 0.000001:
        return round(total_detalhado, 2)
    return round(float(impostos or 0), 2)


def _calc_rendimento_liquido_aplicacao(rendimento_bruto: Optional[float], impostos: Optional[float], rendimento_liquido: Optional[float]) -> float:
    if rendimento_liquido is not None and abs(float(rendimento_liquido or 0)) > 0.000001:
        return round(float(rendimento_liquido or 0), 2)
    return round(float(rendimento_bruto or 0) - float(impostos or 0), 2)


MESES_PT_BR = {
    "JANEIRO": "01",
    "FEVEREIRO": "02",
    "MARCO": "03",
    "MARÇO": "03",
    "ABRIL": "04",
    "MAIO": "05",
    "JUNHO": "06",
    "JULHO": "07",
    "AGOSTO": "08",
    "SETEMBRO": "09",
    "OUTUBRO": "10",
    "NOVEMBRO": "11",
    "DEZEMBRO": "12",
}


def _normalizar_texto_pdf(texto: str) -> str:
    return re.sub(r"\s+", " ", (texto or "")).strip()


def _valor_brl_para_float(valor: Optional[str]) -> float:
    bruto = str(valor or "").strip()
    if not bruto:
        return 0.0
    bruto = bruto.replace(".", "").replace(",", ".")
    try:
        return round(float(bruto), 2)
    except Exception:
        return 0.0


def _parse_data_br(data_str: Optional[str]) -> Optional[date]:
    texto = str(data_str or "").strip()
    if not texto:
        return None
    try:
        return datetime.strptime(texto, "%d/%m/%Y").date()
    except Exception:
        return None


def _extrair_texto_pdf_investimento(file_path: Path) -> str:
    if PdfReader is None:
        raise HTTPException(
            status_code=500,
            detail="Leitura de PDF indisponível no servidor. Instale a dependência 'pypdf'."
        )

    try:
        reader = PdfReader(str(file_path))
        paginas = [page.extract_text() or "" for page in reader.pages]
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Não foi possível ler o PDF enviado: {str(exc)}")

    texto = "\n".join(paginas).strip()
    if not texto:
        raise HTTPException(status_code=400, detail="O PDF não possui texto legível. Pode ser um arquivo escaneado.")
    return texto


def _find_existing_aplicacao_importada(
    db: Session,
    mes_referencia: str,
    instituicao: str,
    produto: str,
    conta_origem: Optional[str] = None,
    arquivo_origem: Optional[str] = None,
):
    arquivo = (arquivo_origem or "").strip()
    if arquivo:
        by_file = db.query(models.AplicacaoFinanceira).filter(
            models.AplicacaoFinanceira.mes_referencia == mes_referencia,
            models.AplicacaoFinanceira.arquivo_origem == arquivo
        ).order_by(models.AplicacaoFinanceira.updated_at.desc(), models.AplicacaoFinanceira.created_at.desc()).first()
        if by_file:
            return by_file

    q = db.query(models.AplicacaoFinanceira).filter(
        models.AplicacaoFinanceira.mes_referencia == mes_referencia,
        models.AplicacaoFinanceira.instituicao.ilike(instituicao),
        models.AplicacaoFinanceira.produto.ilike(produto)
    )

    conta = (conta_origem or "").strip()
    if conta:
        q = q.filter(models.AplicacaoFinanceira.conta_origem == conta)

    return q.order_by(models.AplicacaoFinanceira.updated_at.desc(), models.AplicacaoFinanceira.created_at.desc()).first()


def _extrair_dados_pdf_bb_investimento(texto_pdf: str, filename: str, db: Session) -> dict:
    texto = texto_pdf or ""
    texto_upper = texto.upper()

    mes_ref_match = re.search(r"M[ÊE]S/ANO REFER[ÊE]NCIA\s+([A-ZÇÃ]+)\s*/\s*(\d{4})", texto_upper)
    if not mes_ref_match:
        raise HTTPException(status_code=400, detail="Não foi possível identificar o mês/ano de referência no PDF")

    mes_nome = mes_ref_match.group(1).strip()
    ano_ref = mes_ref_match.group(2)
    mes_num = MESES_PT_BR.get(mes_nome)
    if not mes_num:
        raise HTTPException(status_code=400, detail=f"Mês de referência não reconhecido no PDF: {mes_nome}")
    mes_referencia = f"{ano_ref}-{mes_num}"

    produto_match = re.search(r"^\s*([^\n]+?)\s*-\s*CNPJ:", texto, re.MULTILINE | re.IGNORECASE)
    produto = (produto_match.group(1).strip() if produto_match else "Fundo BB").strip()

    conta_match = re.search(r"CONTA\s+([^\n]+)", texto_upper)
    conta = conta_match.group(1).strip() if conta_match else None

    resumo_match = re.search(
        r"RESUMO DO M[ÊE]S(?P<bloco>.*?)(?:VALOR DA COTA|RENTABILIDADE|TRANSA[ÇC][AÃ]O EFETUADA)",
        texto,
        re.IGNORECASE | re.DOTALL
    )
    if not resumo_match:
        raise HTTPException(status_code=400, detail="Não foi possível localizar o bloco 'Resumo do mês' no PDF")

    resumo_texto = resumo_match.group("bloco")

    def resumo_valor(rotulo: str) -> float:
        match = re.search(rf"{rotulo}\s+([\d\.\,]+)", resumo_texto, re.IGNORECASE)
        return _valor_brl_para_float(match.group(1) if match else None)

    saldo_anterior = resumo_valor(r"SALDO ANTERIOR")
    aplicacoes = resumo_valor(r"APLICA[ÇC][ÕO]ES\s*\(\+\)")
    resgate = resumo_valor(r"RESGATES?\s*\(-\)")
    rendimento_bruto = resumo_valor(r"RENDIMENTO BRUTO\s*\(\+\)")
    imposto_renda = resumo_valor(r"IMPOSTO DE RENDA\s*\(-\)")
    iof = resumo_valor(r"IOF\s*\(-\)")
    rendimento_liquido = resumo_valor(r"RENDIMENTO L[ÍI]QUIDO")
    saldo_atual = resumo_valor(r"SALDO ATUAL\s*=?")
    impostos = _calc_impostos_aplicacao(imposto_renda, iof, None)
    rendimento_liquido = _calc_rendimento_liquido_aplicacao(rendimento_bruto, impostos, rendimento_liquido)

    datas_encontradas = re.findall(r"\b(\d{2}/\d{2}/\d{4})\b", texto)
    data_aplicacao = _parse_data_br(datas_encontradas[-1]) if datas_encontradas else None

    observacoes_partes = [f"Importado do PDF: {filename}"]
    if conta:
        observacoes_partes.append(f"Conta {conta}")
    observacoes_partes.append(f"Referência {mes_referencia}")

    existing = _find_existing_aplicacao_importada(
        db,
        mes_referencia=mes_referencia,
        instituicao="%Banco do Brasil%",
        produto=produto,
        conta_origem=conta,
        arquivo_origem=filename,
    )

    saldo_calculado = _calc_saldo_atual_aplicacao(
        saldo_anterior,
        aplicacoes,
        rendimento_bruto,
        impostos,
        resgate
    )
    saldo_final = saldo_atual if saldo_atual else saldo_calculado

    return {
        "instituicao": "Banco do Brasil",
        "produto": produto,
        "data_aplicacao": data_aplicacao,
        "origem_registro": "importacao_pdf",
        "saldo_anterior": saldo_anterior,
        "aplicacoes": aplicacoes,
        "rendimento_bruto": rendimento_bruto,
        "imposto_renda": imposto_renda,
        "iof": iof,
        "impostos": impostos,
        "rendimento_liquido": rendimento_liquido,
        "resgate": resgate,
        "saldo_atual": saldo_final,
        "mes_referencia": mes_referencia,
        "observacoes": " | ".join(observacoes_partes),
        "conta": conta,
        "arquivo": filename,
        "conta_origem": conta,
        "arquivo_origem": filename,
        "existing_id": existing.id if existing else None,
        "existing_match": bool(existing),
    }


def _extrair_dados_pdf_bb_cdb(texto_pdf: str, filename: str, db: Session) -> dict:
    texto = texto_pdf or ""
    texto_upper = texto.upper()

    periodo_match = re.search(r"PER[ÍI]ODO\s+(\d{2}/\d{2}/\d{4})\s+A\s+(\d{2}/\d{2}/\d{4})", texto_upper)
    if not periodo_match:
        raise HTTPException(status_code=400, detail="Não foi possível identificar o período do extrato de CDB")

    data_inicio = _parse_data_br(periodo_match.group(1))
    data_fim = _parse_data_br(periodo_match.group(2))
    if not data_inicio or not data_fim:
        raise HTTPException(status_code=400, detail="Período do extrato de CDB inválido")

    mes_referencia = data_fim.strftime("%Y-%m")

    conta_match = re.search(r"CONTA\s+([^\n]+)", texto_upper)
    conta = conta_match.group(1).strip() if conta_match else None

    produto_match = re.search(r"^\s*(BB\s+CDB\s+DI)\s*$", texto, re.MULTILINE | re.IGNORECASE)
    produto = produto_match.group(1).strip() if produto_match else "BB CDB DI"

    secao_saldos = re.search(
        r"SALDO NOS [ÚU]LTIMOS 6 MESES(?P<bloco>.*?)(?:RESUMO DOS DEP[ÓO]SITOS EM SER|RENDIMENTO BRUTO NO PER[ÍI]ODO)",
        texto,
        re.IGNORECASE | re.DOTALL
    )
    if not secao_saldos:
        raise HTTPException(status_code=400, detail="Não foi possível localizar a seção 'SALDO NOS ÚLTIMOS 6 MESES'")

    linhas_saldo = re.findall(
        r"(\d{2}/\d{2}/\d{4})\s+([\d\.\,]+)\s+([\d\.\,]+)\s+([\d\.\,]+)\s+([\d\.\,]+)",
        secao_saldos.group("bloco")
    )
    if not linhas_saldo:
        raise HTTPException(status_code=400, detail="Não foi possível extrair os saldos históricos do CDB")

    linha_atual = linhas_saldo[-1]
    linha_anterior = linhas_saldo[-2] if len(linhas_saldo) > 1 else None

    capital_atual = _valor_brl_para_float(linha_atual[1])
    juros_atual = _valor_brl_para_float(linha_atual[2])
    ir_proj_atual = _valor_brl_para_float(linha_atual[3])
    liquido_atual = _valor_brl_para_float(linha_atual[4])

    if linha_anterior:
        juros_anterior = _valor_brl_para_float(linha_anterior[2])
        ir_proj_anterior = _valor_brl_para_float(linha_anterior[3])
        liquido_anterior = _valor_brl_para_float(linha_anterior[4])
        saldo_anterior = liquido_anterior
        rendimento_bruto = round(juros_atual - juros_anterior, 2)
        imposto_renda = round(ir_proj_atual - ir_proj_anterior, 2)
        rendimento_liquido = round(liquido_atual - liquido_anterior, 2)
    else:
        saldo_anterior = capital_atual
        rendimento_bruto = juros_atual
        imposto_renda = ir_proj_atual
        rendimento_liquido = liquido_atual - capital_atual

    iof = 0.0
    impostos = _calc_impostos_aplicacao(imposto_renda, iof, None)
    aplicacoes = 0.0
    resgate = 0.0
    saldo_atual = liquido_atual

    if abs(rendimento_liquido) <= 0.000001:
        rendimento_liquido = _calc_rendimento_liquido_aplicacao(rendimento_bruto, impostos, rendimento_liquido)

    observacoes_partes = [
        f"Importado do PDF CDB: {filename}",
        f"Período {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
    ]
    if conta:
        observacoes_partes.append(f"Conta {conta}")

    existing = _find_existing_aplicacao_importada(
        db,
        mes_referencia=mes_referencia,
        instituicao="%Banco do Brasil%",
        produto="%CDB%",
        conta_origem=conta,
        arquivo_origem=filename,
    )

    return {
        "instituicao": "Banco do Brasil",
        "produto": produto,
        "data_aplicacao": data_fim,
        "origem_registro": "importacao_pdf",
        "saldo_anterior": round(saldo_anterior, 2),
        "aplicacoes": aplicacoes,
        "rendimento_bruto": round(rendimento_bruto, 2),
        "imposto_renda": round(imposto_renda, 2),
        "iof": iof,
        "impostos": round(impostos, 2),
        "rendimento_liquido": round(rendimento_liquido, 2),
        "resgate": resgate,
        "saldo_atual": round(saldo_atual, 2),
        "mes_referencia": mes_referencia,
        "observacoes": " | ".join(observacoes_partes),
        "conta": conta,
        "arquivo": filename,
        "conta_origem": conta,
        "arquivo_origem": filename,
        "existing_id": existing.id if existing else None,
        "existing_match": bool(existing),
    }


def _extrair_dados_pdf_investimento(texto_pdf: str, filename: str, db: Session) -> dict:
    texto_upper = (texto_pdf or "").upper()
    if "CDB / RDB" in texto_upper or "BB REAPLIC" in texto_upper or "BB CDB DI" in texto_upper:
        return _extrair_dados_pdf_bb_cdb(texto_pdf, filename, db)
    return _extrair_dados_pdf_bb_investimento(texto_pdf, filename, db)


@app.get("/api/aplicacoes-financeiras", response_model=List[schemas.AplicacaoFinanceiraResponse])
def list_aplicacoes_financeiras(
    mes_referencia: Optional[str] = None,
    instituicao: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.AplicacaoFinanceira)
    if mes_referencia:
        q = q.filter(models.AplicacaoFinanceira.mes_referencia == mes_referencia)
    if instituicao:
        q = q.filter(models.AplicacaoFinanceira.instituicao.ilike(f"%{instituicao}%"))
    return q.order_by(models.AplicacaoFinanceira.instituicao.asc(), models.AplicacaoFinanceira.produto.asc()).all()


@app.get("/api/aplicacoes-financeiras/resumo", response_model=schemas.AplicacaoFinanceiraResumoResponse)
def resumo_aplicacoes_financeiras(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    mes_ref = mes_referencia or date.today().strftime("%Y-%m")
    registros = db.query(models.AplicacaoFinanceira).filter(
        models.AplicacaoFinanceira.mes_referencia == mes_ref
    ).all()

    totais = {
        "saldo_anterior": round(sum(float(r.saldo_anterior or 0) for r in registros), 2),
        "aplicacoes": round(sum(float(r.aplicacoes or 0) for r in registros), 2),
        "rendimento_bruto": round(sum(float(r.rendimento_bruto or 0) for r in registros), 2),
        "imposto_renda": round(sum(float(r.imposto_renda or 0) for r in registros), 2),
        "iof": round(sum(float(r.iof or 0) for r in registros), 2),
        "impostos": round(sum(float(r.impostos or 0) for r in registros), 2),
        "rendimento_liquido": round(sum(float(r.rendimento_liquido or 0) for r in registros), 2),
        "resgate": round(sum(float(r.resgate or 0) for r in registros), 2),
        "saldo_atual": round(sum(float(r.saldo_atual or 0) for r in registros), 2),
    }

    return {
        "mes_referencia": mes_ref,
        "total_registros": len(registros),
        "totais": totais
    }


@app.post("/api/aplicacoes-financeiras/importar-pdf-preview", response_model=schemas.AplicacaoFinanceiraImportPreview)
async def preview_import_aplicacao_financeira_pdf(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    nome_arquivo = (file.filename or "").strip()
    if not nome_arquivo.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Envie um arquivo PDF válido")

    temp_upload = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            temp_upload = Path(tmp.name)
            tmp.write(await file.read())

        texto_pdf = _extrair_texto_pdf_investimento(temp_upload)
        return _extrair_dados_pdf_investimento(texto_pdf, nome_arquivo, db)
    finally:
        if temp_upload and temp_upload.exists():
            temp_upload.unlink(missing_ok=True)
        await file.close()


@app.post("/api/aplicacoes-financeiras/importar-pdf-confirmar", response_model=schemas.AplicacaoFinanceiraResponse)
def confirm_import_aplicacao_financeira_pdf(
    req: schemas.AplicacaoFinanceiraImportConfirmRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    existing = _find_existing_aplicacao_importada(
        db,
        mes_referencia=req.mes_referencia,
        instituicao=f"%{(req.instituicao or '').strip()}%",
        produto=(req.produto or '').strip() or "%",
        conta_origem=req.conta_origem,
        arquivo_origem=req.arquivo_origem,
    )

    if existing and not req.existing_id:
        req.existing_id = existing.id

    if existing and req.existing_id and existing.id != req.existing_id:
        req.existing_id = existing.id

    impostos = _calc_impostos_aplicacao(req.imposto_renda, req.iof, req.impostos)
    rendimento_liquido = _calc_rendimento_liquido_aplicacao(req.rendimento_bruto, impostos, req.rendimento_liquido)
    saldo_atual = req.saldo_atual
    if saldo_atual is None:
        saldo_atual = _calc_saldo_atual_aplicacao(
            req.saldo_anterior,
            req.aplicacoes,
            req.rendimento_bruto,
            impostos,
            req.resgate
        )

    registro = None
    if req.existing_id:
        registro = db.query(models.AplicacaoFinanceira).filter(models.AplicacaoFinanceira.id == req.existing_id).first()

    if registro:
        registro.user_id = current_user.id
        registro.mes_referencia = req.mes_referencia
        registro.data_aplicacao = req.data_aplicacao
        registro.instituicao = (req.instituicao or "").strip()
        registro.produto = (req.produto or "").strip()
        registro.origem_registro = req.origem_registro or registro.origem_registro or "manual"
        registro.conta_origem = req.conta_origem
        registro.arquivo_origem = req.arquivo_origem
        registro.saldo_anterior = float(req.saldo_anterior or 0)
        registro.aplicacoes = float(req.aplicacoes or 0)
        registro.rendimento_bruto = float(req.rendimento_bruto or 0)
        registro.imposto_renda = float(req.imposto_renda or 0)
        registro.iof = float(req.iof or 0)
        registro.impostos = float(impostos or 0)
        registro.rendimento_liquido = float(rendimento_liquido or 0)
        registro.resgate = float(req.resgate or 0)
        registro.saldo_atual = float(saldo_atual or 0)
        registro.observacoes = req.observacoes
        registro.updated_at = datetime.utcnow()
    else:
        registro = models.AplicacaoFinanceira(
            id=str(uuid.uuid4()),
            user_id=current_user.id,
            mes_referencia=req.mes_referencia,
            data_aplicacao=req.data_aplicacao,
            instituicao=(req.instituicao or "").strip(),
            produto=(req.produto or "").strip(),
            origem_registro=req.origem_registro or "manual",
            conta_origem=req.conta_origem,
            arquivo_origem=req.arquivo_origem,
            saldo_anterior=float(req.saldo_anterior or 0),
            aplicacoes=float(req.aplicacoes or 0),
            rendimento_bruto=float(req.rendimento_bruto or 0),
            imposto_renda=float(req.imposto_renda or 0),
            iof=float(req.iof or 0),
            impostos=float(impostos or 0),
            rendimento_liquido=float(rendimento_liquido or 0),
            resgate=float(req.resgate or 0),
            saldo_atual=float(saldo_atual or 0),
            observacoes=req.observacoes,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        db.add(registro)

    db.commit()
    db.refresh(registro)
    return registro


@app.post("/api/aplicacoes-financeiras", response_model=schemas.AplicacaoFinanceiraResponse)
def create_aplicacao_financeira(
    req: schemas.AplicacaoFinanceiraCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    mes_ref = req.mes_referencia or date.today().strftime("%Y-%m")
    impostos = _calc_impostos_aplicacao(req.imposto_renda, req.iof, req.impostos)
    rendimento_liquido = _calc_rendimento_liquido_aplicacao(req.rendimento_bruto, impostos, req.rendimento_liquido)
    saldo_atual = _calc_saldo_atual_aplicacao(
        req.saldo_anterior,
        req.aplicacoes,
        req.rendimento_bruto,
        impostos,
        req.resgate
    )
    registro = models.AplicacaoFinanceira(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        mes_referencia=mes_ref,
        data_aplicacao=req.data_aplicacao,
        instituicao=(req.instituicao or "").strip(),
        produto=(req.produto or "").strip(),
        origem_registro=req.origem_registro or "manual",
        conta_origem=req.conta_origem,
        arquivo_origem=req.arquivo_origem,
        saldo_anterior=float(req.saldo_anterior or 0),
        aplicacoes=float(req.aplicacoes or 0),
        rendimento_bruto=float(req.rendimento_bruto or 0),
        imposto_renda=float(req.imposto_renda or 0),
        iof=float(req.iof or 0),
        impostos=float(impostos or 0),
        rendimento_liquido=float(rendimento_liquido or 0),
        resgate=float(req.resgate or 0),
        saldo_atual=saldo_atual,
        observacoes=req.observacoes,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow()
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)
    return registro


@app.put("/api/aplicacoes-financeiras/{aplicacao_id}", response_model=schemas.AplicacaoFinanceiraResponse)
def update_aplicacao_financeira(
    aplicacao_id: str,
    req: schemas.AplicacaoFinanceiraUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    registro = db.query(models.AplicacaoFinanceira).filter(models.AplicacaoFinanceira.id == aplicacao_id).first()
    if not registro:
        raise HTTPException(status_code=404, detail="Aplicação financeira não encontrada")

    for k, v in req.dict(exclude_none=True).items():
        if k in ("instituicao", "produto") and isinstance(v, str):
            v = v.strip()
        setattr(registro, k, v)

    registro.impostos = _calc_impostos_aplicacao(registro.imposto_renda, registro.iof, registro.impostos)
    registro.rendimento_liquido = _calc_rendimento_liquido_aplicacao(
        registro.rendimento_bruto,
        registro.impostos,
        registro.rendimento_liquido
    )
    registro.saldo_atual = _calc_saldo_atual_aplicacao(
        registro.saldo_anterior,
        registro.aplicacoes,
        registro.rendimento_bruto,
        registro.impostos,
        registro.resgate
    )
    registro.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(registro)
    return registro


@app.delete("/api/aplicacoes-financeiras/{aplicacao_id}")
def delete_aplicacao_financeira(
    aplicacao_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    registro = db.query(models.AplicacaoFinanceira).filter(models.AplicacaoFinanceira.id == aplicacao_id).first()
    if not registro:
        raise HTTPException(status_code=404, detail="Aplicação financeira não encontrada")

    db.delete(registro)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# TRANSAÇÕES / FLUXO DE CAIXA
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/transacoes")
def list_transacoes(
    mes_referencia: Optional[str] = None,
    tipo: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Transacao)
    if mes_referencia:
        ano, mes = mes_referencia.split("-")
        inicio = date(int(ano), int(mes), 1)
        ultimo_dia = calendar.monthrange(int(ano), int(mes))[1]
        fim = date(int(ano), int(mes), ultimo_dia)
        q = q.filter(models.Transacao.data_transacao.between(inicio, fim))
    if tipo:
        q = q.filter(models.Transacao.tipo == tipo)
    return q.order_by(models.Transacao.data_transacao.desc()).all()


def _mes_anterior(mes_ref: str) -> str:
    ano, mes = mes_ref.split("-")
    ano_i = int(ano)
    mes_i = int(mes)
    if mes_i == 1:
        return f"{ano_i - 1}-12"
    return f"{ano_i}-{str(mes_i - 1).zfill(2)}"


def _saldo_anterior_mes(db: Session, mes_ref: str) -> float:
    saldo_manual = db.query(models.SaldoMensal).filter(models.SaldoMensal.mes_referencia == mes_ref).first()
    if saldo_manual:
        return round(float(saldo_manual.valor_saldo_inicial or 0), 2)

    total_pagamentos_anteriores = db.query(sql_func.coalesce(sql_func.sum(models.Pagamento.valor_pago), 0)).filter(
        models.Pagamento.status_pagamento == "pago",
        models.Pagamento.mes_referencia.isnot(None),
        models.Pagamento.mes_referencia < mes_ref
    ).scalar()

    total_outras_rendas_anteriores = db.query(sql_func.coalesce(sql_func.sum(models.OutraRenda.valor), 0)).filter(
        models.OutraRenda.mes_referencia.isnot(None),
        models.OutraRenda.mes_referencia < mes_ref
    ).scalar()

    total_despesas_anteriores = db.query(sql_func.coalesce(sql_func.sum(models.Despesa.valor), 0)).filter(
        models.Despesa.mes_referencia.isnot(None),
        models.Despesa.mes_referencia < mes_ref
    ).scalar()

    return round(
        float(total_pagamentos_anteriores or 0) + float(total_outras_rendas_anteriores or 0) - float(total_despesas_anteriores or 0),
        2
    )


@app.get("/api/saldo-inicial", response_model=schemas.SaldoMensalResponse)
def get_saldo_inicial(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    mes_ref = mes_referencia or date.today().strftime("%Y-%m")
    saldo_manual = db.query(models.SaldoMensal).filter(models.SaldoMensal.mes_referencia == mes_ref).first()

    if saldo_manual:
        return {
            "mes_referencia": mes_ref,
            "valor_saldo_inicial": round(float(saldo_manual.valor_saldo_inicial or 0), 2),
            "origem": "manual",
            "observacoes": saldo_manual.observacoes
        }

    return {
        "mes_referencia": mes_ref,
        "valor_saldo_inicial": _saldo_anterior_mes(db, mes_ref),
        "origem": "calculado",
        "observacoes": None
    }


@app.put("/api/saldo-inicial", response_model=schemas.SaldoMensalResponse)
def upsert_saldo_inicial(
    req: schemas.SaldoMensalUpsert,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    registro = db.query(models.SaldoMensal).filter(models.SaldoMensal.mes_referencia == req.mes_referencia).first()

    if registro:
        registro.valor_saldo_inicial = float(req.valor_saldo_inicial)
        registro.observacoes = req.observacoes
        registro.user_id = current_user.id
        registro.updated_at = datetime.utcnow()
    else:
        registro = models.SaldoMensal(
            id=str(uuid.uuid4()),
            user_id=current_user.id,
            mes_referencia=req.mes_referencia,
            valor_saldo_inicial=float(req.valor_saldo_inicial),
            observacoes=req.observacoes,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.add(registro)

    db.commit()
    db.refresh(registro)

    return {
        "mes_referencia": registro.mes_referencia,
        "valor_saldo_inicial": round(float(registro.valor_saldo_inicial or 0), 2),
        "origem": "manual",
        "observacoes": registro.observacoes
    }


@app.delete("/api/saldo-inicial")
def delete_saldo_inicial(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    mes_ref = mes_referencia or date.today().strftime("%Y-%m")
    registro = db.query(models.SaldoMensal).filter(models.SaldoMensal.mes_referencia == mes_ref).first()

    if not registro:
        return {"ok": True, "deleted": False, "detail": "Saldo manual não encontrado para este mês"}

    db.delete(registro)
    db.commit()
    return {"ok": True, "deleted": True, "detail": "Saldo manual removido; cálculo automático reativado"}

@app.get("/api/fluxo-caixa")
def fluxo_caixa(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")
    pags_mes = _pagamentos_pagos_membros_ativos_no_mes(db, mes_ref)
    rendas_mes = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_ref).all()
    despesas_mes = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_ref).all()

    membros_ids = [p.membro_id for p in pags_mes if p.membro_id]
    nomes_membros = {}
    if membros_ids:
        nomes_membros = {
            m.id: m.nome_completo
            for m in db.query(models.Membro).filter(models.Membro.id.in_(membros_ids)).all()
        }

    transacoes = []

    for p in pags_mes:
        nome = nomes_membros.get(p.membro_id, "Membro")
        transacoes.append({
            "id": f"pag-{p.id}",
            "descricao": f"Mensalidade {nome} - {mes_ref}",
            "valor": float(p.valor_pago) if p.valor_pago else 0,
            "tipo": "entrada",
            "categoria": f"Mensalidade {mes_ref}",
            "data_transacao": p.data_pagamento,
            "origem": "mensalidade"
        })

    for r in rendas_mes:
        transacoes.append({
            "id": f"renda-{r.id}",
            "descricao": r.descricao,
            "valor": float(r.valor) if r.valor else 0,
            "tipo": "entrada",
            "categoria": r.categoria,
            "data_transacao": r.data_recebimento,
            "origem": r.categoria
        })

    for d in despesas_mes:
        transacoes.append({
            "id": f"desp-{d.id}",
            "descricao": d.descricao,
            "valor": float(d.valor) if d.valor else 0,
            "tipo": "saida",
            "categoria": d.categoria,
            "data_transacao": d.data_despesa,
            "origem": "despesa"
        })

    transacoes.sort(key=lambda t: t["data_transacao"] or date.min)

    total_entradas = sum(t["valor"] for t in transacoes if t["tipo"] == "entrada" and t["valor"])
    total_saidas = sum(t["valor"] for t in transacoes if t["tipo"] == "saida" and t["valor"])
    saldo = total_entradas - total_saidas
    saldo_manual_registrado = db.query(models.SaldoMensal).filter(models.SaldoMensal.mes_referencia == mes_ref).first()
    saldo_anterior = _saldo_anterior_mes(db, mes_ref)
    saldo_final = round(saldo_anterior + saldo, 2)

    # Monthly evolution (last 12 months)
    evolucao = []
    for i in range(11, -1, -1):
        ref_date = today - timedelta(days=i * 30)
        mes_iter = ref_date.strftime("%Y-%m")
        pags_i = _pagamentos_pagos_membros_ativos_no_mes(db, mes_iter)
        rendas_i = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_iter).all()
        despesas_i = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_iter).all()
        ent = sum(float(p.valor_pago) for p in pags_i if p.valor_pago) + sum(float(r.valor) for r in rendas_i if r.valor)
        sai = sum(float(d.valor) for d in despesas_i if d.valor)
        evolucao.append({"mes": mes_iter, "entradas": ent, "saidas": sai, "saldo": ent - sai})

    return {
        "mes_referencia": mes_ref,
        "mes_referencia_anterior": _mes_anterior(mes_ref),
        "saldo_anterior": saldo_anterior,
        "origem_saldo_anterior": "manual" if saldo_manual_registrado else "calculado",
        "total_entradas": total_entradas,
        "total_saidas": total_saidas,
        "saldo": saldo,
        "saldo_final": saldo_final,
        "transacoes": [{
            "id": t["id"], "descricao": t["descricao"], "valor": t["valor"],
            "tipo": t["tipo"], "categoria": t["categoria"],
            "data_transacao": str(t["data_transacao"]) if t["data_transacao"] else None,
            "origem": t["origem"]
        } for t in transacoes],
        "evolucao_mensal": evolucao
    }

@app.post("/api/transacoes", response_model=schemas.TransacaoResponse)
def create_transacao(req: schemas.TransacaoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    t = models.Transacao(id=str(uuid.uuid4()), user_id=current_user.id, created_at=datetime.utcnow(), **req.dict())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


# ════════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/dashboard")
def dashboard(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    if mes_referencia:
        mes_atual = mes_referencia
    else:
        meses_com_movimento = []

        ultimo_pagamento = db.query(sql_func.max(models.Pagamento.mes_referencia)).filter(
            models.Pagamento.mes_referencia.isnot(None),
            models.Pagamento.status_pagamento == 'pago'
        ).scalar()
        if ultimo_pagamento:
            meses_com_movimento.append(ultimo_pagamento)

        ultimo_despesa = db.query(sql_func.max(models.Despesa.mes_referencia)).filter(
            models.Despesa.mes_referencia.isnot(None)
        ).scalar()
        if ultimo_despesa:
            meses_com_movimento.append(ultimo_despesa)

        ultima_renda = db.query(sql_func.max(models.OutraRenda.mes_referencia)).filter(
            models.OutraRenda.mes_referencia.isnot(None)
        ).scalar()
        if ultima_renda:
            meses_com_movimento.append(ultima_renda)

        ultima_aplicacao = db.query(sql_func.max(models.AplicacaoFinanceira.mes_referencia)).filter(
            models.AplicacaoFinanceira.mes_referencia.isnot(None)
        ).scalar()
        if ultima_aplicacao:
            meses_com_movimento.append(ultima_aplicacao)

        mes_atual = max(meses_com_movimento) if meses_com_movimento else today.strftime("%Y-%m")
    ano, mes = mes_atual.split("-")
    ref_date = date(int(ano), int(mes), 1)

    total_membros = db.query(models.Membro).filter(models.Membro.status == 'ativo').count()
    total_inativos = db.query(models.Membro).filter(models.Membro.status == 'inativo').count()

    # Pagamentos do mês (por membro ativo, evitando duplicidade no mesmo mês)
    pags_mes = _pagamentos_pagos_membros_ativos_no_mes(db, mes_atual)
    total_arrecadado = sum(float(p.valor_pago) for p in pags_mes if p.valor_pago)
    total_pagantes = len(pags_mes)
    inadimplentes = total_membros - total_pagantes

    # Despesas do mês
    despesas_mes = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_atual).all()
    total_despesas = sum(d.valor for d in despesas_mes if d.valor)

    # Outras rendas do mês
    rendas_mes = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_atual).all()
    total_outras_rendas = sum(r.valor for r in rendas_mes if r.valor)

    saldo_mes = total_arrecadado + total_outras_rendas - total_despesas

    # Aniversariantes do mês
    aniv_mes = db.query(models.Membro).filter(
        sql_func.strftime('%m', models.Membro.data_nascimento) == str(int(mes)).zfill(2)
    ).count()

    # Aniversariantes do dia
    aniv_hoje = db.query(models.Membro).filter(
        sql_func.strftime('%m-%d', models.Membro.data_nascimento) == today.strftime('%m-%d')
    ).all()

    # Evolução últimos 6 meses
    evolucao = []
    for i in range(5, -1, -1):
        ref_iter = ref_date - timedelta(days=i * 28)
        mes_iter = ref_iter.strftime("%Y-%m")
        pags_i = db.query(models.Pagamento).filter(
            models.Pagamento.mes_referencia == mes_iter,
            models.Pagamento.status_pagamento == 'pago'
        ).all()
        desp_i = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_iter).all()
        rend_i = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_iter).all()
        ent_i = sum(float(p.valor_pago) for p in pags_i if p.valor_pago) + sum(r.valor for r in rend_i if r.valor)
        sai_i = sum(d.valor for d in desp_i if d.valor)
        evolucao.append({"mes": mes_iter, "entradas": ent_i, "saidas": sai_i, "saldo": ent_i - sai_i})

    # Por status
    status_membros = db.query(models.Membro.status, sql_func.count(models.Membro.id)).group_by(models.Membro.status).all()

    return {
        "total_membros": total_membros,
        "total_inativos": total_inativos,
        "total_pagantes": total_pagantes,
        "inadimplentes": inadimplentes,
        "total_arrecadado": total_arrecadado,
        "total_despesas": total_despesas,
        "total_outras_rendas": total_outras_rendas,
        "saldo_mes": saldo_mes,
        "aniversariantes_mes": aniv_mes,
        "aniversariantes_hoje": [{"id": m.id, "nome": m.nome_completo, "data_nascimento": str(m.data_nascimento)} for m in aniv_hoje],
        "evolucao_mensal": evolucao,
        "status_membros": [{"status": s, "total": c} for s, c in status_membros],
        "mes_atual": mes_atual,
    }


# ════════════════════════════════════════════════════════════════════════════════
# FESTAS
# ════════════════════════════════════════════════════════════════════════════════
def _get_frontend_base_url() -> str:
    return (
        os.getenv("FRONTEND_BASE_URL")
        or os.getenv("FRONTEND_URL")
        or "http://localhost:5173"
    ).rstrip("/")


def _get_festa_link_template() -> str:
    return f"{_get_frontend_base_url()}/#/festa-inscricao/{{festa_id}}"


def _build_festa_public_link(festa_id: str) -> str:
    return f"{_get_frontend_base_url()}/#/festa-inscricao/{festa_id}"


def _resolve_festa_public_link(festa: models.Festa) -> str:
    festa_id = str(festa.id)
    fallback = _build_festa_public_link(festa_id)
    template = (getattr(festa, "link_inscricao", None) or "").strip()

    if not template:
        return fallback

    link = template.replace("{festa_id}", festa_id).replace("{token}", festa_id)

    if "festa-inscricao" not in link:
        return fallback

    if "/festa-inscricao/" in link and "/#/festa-inscricao/" not in link:
        link = link.replace("/festa-inscricao/", "/#/festa-inscricao/")

    return link


def _somente_numeros(valor: Optional[str]) -> str:
    return re.sub(r"\D", "", (valor or "").strip())


def _politica_preco_default() -> dict:
    return {
        "cortesia_acompanhantes": 1,
        "idade_gratis_ate": 5,
        "idade_meia_de": 6,
        "idade_meia_ate": 10,
        "percentual_meia": 50,
    }


def _safe_int(valor, default: int, min_value: Optional[int] = None, max_value: Optional[int] = None) -> int:
    try:
        parsed = int(valor)
    except (TypeError, ValueError):
        parsed = default
    if min_value is not None:
        parsed = max(min_value, parsed)
    if max_value is not None:
        parsed = min(max_value, parsed)
    return parsed


def _obter_politica_preco_festa(festa: models.Festa) -> dict:
    politica = _politica_preco_default()
    raw = (festa.politica_precos or "").strip() if festa else ""
    if not raw:
        return politica

    try:
        data = json.loads(raw)
    except Exception:
        return politica

    if isinstance(data, dict):
        src = data.get("pricing_rules") if isinstance(data.get("pricing_rules"), dict) else data
        if isinstance(src, dict):
            if src.get("cortesia_acompanhantes") is not None:
                politica["cortesia_acompanhantes"] = _safe_int(src.get("cortesia_acompanhantes"), 1, 0)
            if src.get("idade_gratis_ate") is not None:
                politica["idade_gratis_ate"] = _safe_int(src.get("idade_gratis_ate"), 5, 0)
            if src.get("idade_meia_de") is not None:
                politica["idade_meia_de"] = _safe_int(src.get("idade_meia_de"), 6, 0)
            if src.get("idade_meia_ate") is not None:
                politica["idade_meia_ate"] = _safe_int(src.get("idade_meia_ate"), 10, 0)
            if src.get("percentual_meia") is not None:
                politica["percentual_meia"] = _safe_int(src.get("percentual_meia"), 50, 0, 100)

    return politica


def _normalizar_idade(valor: Optional[int], campo: str) -> Optional[int]:
    if valor is None:
        return None
    try:
        idade = int(valor)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"{campo} inválida")
    if idade < 0 or idade > 120:
        raise HTTPException(status_code=400, detail=f"{campo} deve estar entre 0 e 120")
    return idade


def _calcular_custo_participante_por_idade(idade: Optional[int], valor_integral: float, politica: dict) -> float:
    valor_base = float(valor_integral or 0)
    if idade is None:
        return valor_base
    idade_gratis_ate = int(politica.get("idade_gratis_ate", 5) or 5)
    idade_meia_de = int(politica.get("idade_meia_de", 6) or 6)
    idade_meia_ate = int(politica.get("idade_meia_ate", 10) or 10)
    percentual_meia = int(politica.get("percentual_meia", 50) or 50)

    if idade <= idade_gratis_ate:
        return 0.0
    if idade_meia_de <= idade <= idade_meia_ate:
        return round(valor_base * (percentual_meia / 100), 2)
    return valor_base


def _validar_membro_por_matricula_cpf(db: Session, matricula: str, cpf: str):
    matricula_limpa = (matricula or "").strip()
    cpf_limpo = _somente_numeros(cpf)

    if not matricula_limpa or not cpf_limpo:
        raise HTTPException(status_code=400, detail="Informe matrícula e CPF")

    membro = db.query(models.Membro).filter(models.Membro.matricula == matricula_limpa).first()
    if not membro:
        raise HTTPException(status_code=401, detail="Matrícula ou CPF inválidos")

    cpf_cadastrado = _somente_numeros(membro.cpf)
    if not cpf_cadastrado or cpf_cadastrado != cpf_limpo:
        raise HTTPException(status_code=401, detail="Matrícula ou CPF inválidos")

    return membro


def _normalizar_filtro_sexo(valor: Optional[str]) -> Optional[List[str]]:
    if not valor:
        return None
    sexo = valor.strip().lower()
    if sexo in ("masculino", "m"):
        return ["masculino", "m"]
    if sexo in ("feminino", "f"):
        return ["feminino", "f"]
    return [sexo]


def _email_valido(email: Optional[str]) -> bool:
    if not email:
        return False
    return re.fullmatch(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email.strip()) is not None


def _montar_payload_inscricao_publica(db: Session, festa: models.Festa, membro: models.Membro):
    politica_preco = _obter_politica_preco_festa(festa)
    titular = db.query(models.ParticipacaoFesta).filter(
        models.ParticipacaoFesta.festa_id == festa.id,
        models.ParticipacaoFesta.membro_id == membro.id,
        models.ParticipacaoFesta.tipo_participante == "titular"
    ).first()

    dependente = db.query(models.ParticipacaoFesta).filter(
        models.ParticipacaoFesta.festa_id == festa.id,
        models.ParticipacaoFesta.membro_id == membro.id,
        models.ParticipacaoFesta.tipo_participante == "dependente"
    ).first()

    convidado = db.query(models.ParticipacaoFesta).filter(
        models.ParticipacaoFesta.festa_id == festa.id,
        models.ParticipacaoFesta.membro_id == membro.id,
        models.ParticipacaoFesta.tipo_participante == "convidado"
    ).first()

    return {
        "ok": True,
        "festa": {
            "id": festa.id,
            "nome_festa": festa.nome_festa,
            "data_festa": str(festa.data_festa) if festa.data_festa else None,
            "local_festa": festa.local_festa,
            "beneficio_titular_dependente_gratis": True,
            "valor_convite": float(festa.valor_convite) if festa.valor_convite else 0,
            "valor_convite_dependente": float(festa.valor_convite_dependente) if festa.valor_convite_dependente else 0,
            "politica_preco": politica_preco,
            "descricao": festa.descricao,
            "status": festa.status,
            "capacidade": festa.capacidade,
        },
        "membro": {
            "id": membro.id,
            "nome": membro.nome_completo,
            "email": membro.email,
            "matricula": membro.matricula,
            "cpf": membro.cpf,
        },
        "inscricao": {
            "titular": {
                "id": titular.id,
                "nome_participante": titular.nome_participante,
                "pago": titular.pago,
                "observacoes": titular.observacoes,
            } if titular else None,
            "dependente": {
                "id": dependente.id,
                "nome_dependente": dependente.nome_dependente,
                "parentesco": dependente.parentesco,
                "pago": dependente.pago,
                "observacoes": dependente.observacoes,
            } if dependente else None,
            "convidado": {
                "id": convidado.id,
                "nome_convidado": convidado.nome_participante,
                "custo_convite": float(convidado.custo_convite) if convidado.custo_convite else 0,
                "pago": convidado.pago,
                "observacoes": convidado.observacoes,
            } if convidado else None,
        }
    }


def _validar_festa_link_disponivel(festa: models.Festa):
    if festa.data_festa and festa.data_festa <= date.today():
        raise HTTPException(status_code=400, detail="Link disponível apenas para festas com data futura")


def _send_html_email(to_email: str, subject: str, html_body: str, plain_text_body: str):
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from_name = os.getenv("SMTP_FROM_NAME", "UNACOB")
    smtp_from_email = os.getenv("SMTP_FROM_EMAIL", smtp_user or "")
    smtp_starttls = os.getenv("SMTP_STARTTLS", "true").lower() in ("1", "true", "yes", "on")

    if not smtp_host or not smtp_from_email:
        raise HTTPException(status_code=500, detail="Configuração de e-mail incompleta no servidor")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"{smtp_from_name} <{smtp_from_email}>"
    msg["To"] = to_email
    msg.set_content(plain_text_body)
    msg.add_alternative(html_body, subtype="html")

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as smtp:
            smtp.ehlo()
            if smtp_starttls:
                smtp.starttls()
                smtp.ehlo()
            if smtp_user and smtp_password:
                smtp.login(smtp_user, smtp_password)
            smtp.send_message(msg)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Falha ao enviar e-mail: {str(exc)}")


@app.get("/api/festas", response_model=List[schemas.FestaResponse])
def list_festas(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    return db.query(models.Festa).order_by(models.Festa.data_festa.desc()).all()

@app.post("/api/festas", response_model=schemas.FestaResponse)
def create_festa(req: schemas.FestaCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    payload = req.dict()
    if not payload.get("link_inscricao"):
        payload["link_inscricao"] = _get_festa_link_template()
    f = models.Festa(id=str(uuid.uuid4()), user_id=current_user.id, created_at=datetime.utcnow(), updated_at=datetime.utcnow(), **payload)
    db.add(f)
    db.commit()
    db.refresh(f)
    return f

@app.get("/api/festas/{festa_id}", response_model=schemas.FestaResponse)
def get_festa(festa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    return f

@app.put("/api/festas/{festa_id}", response_model=schemas.FestaResponse)
def update_festa(festa_id: str, req: schemas.FestaUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    for k, v in req.dict(exclude_none=True).items():
        setattr(f, k, v)
    if not f.link_inscricao:
        f.link_inscricao = _get_festa_link_template()
    f.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(f)
    return f


@app.post("/api/festas/{festa_id}/enviar-convites")
def enviar_convites_festa(
    festa_id: str,
    req: schemas.FestaConviteEmailRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    festa = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not festa:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    _validar_festa_link_disponivel(festa)

    q_membros = db.query(models.Membro).filter(
        models.Membro.email.isnot(None),
        models.Membro.email != ""
    )

    if req.filtro_status:
        q_membros = q_membros.filter(models.Membro.status == req.filtro_status)

    if req.filtro_nome:
        q_membros = q_membros.filter(models.Membro.nome_completo.ilike(f"%{req.filtro_nome}%"))

    if req.filtro_matricula:
        q_membros = q_membros.filter(models.Membro.matricula.ilike(f"%{req.filtro_matricula}%"))

    if req.filtro_cidade:
        q_membros = q_membros.filter(models.Membro.cidade.ilike(f"%{req.filtro_cidade}%"))

    sexos = _normalizar_filtro_sexo(req.filtro_sexo)
    if sexos:
        q_membros = q_membros.filter(sql_func.lower(models.Membro.sexo).in_(sexos))

    if req.membro_ids:
        q_membros = q_membros.filter(models.Membro.id.in_(req.membro_ids))

    membros = q_membros.order_by(models.Membro.nome_completo).all()

    if req.somente_email_valido:
        membros = [m for m in membros if _email_valido(m.email)]

    if req.somente_pendentes:
        membros_ja_confirmados = {
            mid for (mid,) in db.query(models.ParticipacaoFesta.membro_id).filter(
                models.ParticipacaoFesta.festa_id == festa.id,
                models.ParticipacaoFesta.tipo_participante == "titular",
                models.ParticipacaoFesta.membro_id.isnot(None)
            ).all()
        }
        membros = [m for m in membros if m.id not in membros_ja_confirmados]

    enviados = 0
    falhas = []

    convite_link = _resolve_festa_public_link(festa)
    data_festa_txt = festa.data_festa.strftime("%d/%m/%Y") if festa.data_festa else "-"
    valor_convite = float(festa.valor_convite or 0)
    valor_convite_txt = f"R$ {valor_convite:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    valor_dependente = float(festa.valor_convite_dependente or 0)
    valor_dependente_txt = f"R$ {valor_dependente:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    logo_url = (os.getenv("EMAIL_LOGO_URL") or "").strip()
    festa_image_url = (os.getenv("EMAIL_FESTA_IMAGE_URL") or "").strip()

    logo_block = ""
    if logo_url:
        logo_block = f"""
        <div style="margin:0 0 16px;text-align:center;">
            <img src="{logo_url}" alt="UNACOB" style="max-width:190px;width:100%;height:auto;display:inline-block;" />
        </div>
        """

    festa_image_block = """
    <div style="font-size:46px;line-height:1;margin-bottom:8px;">🎉</div>
    """
    if festa_image_url:
        festa_image_block = f"""
        <div style="margin:0 0 10px;text-align:center;">
            <img src="{festa_image_url}" alt="Celebração" style="max-width:170px;width:100%;height:auto;display:inline-block;border-radius:10px;" />
        </div>
        """

    for membro in membros:
        assunto = req.assunto or f"Convite: {festa.nome_festa}"

        mensagem_extra = ""
        if req.mensagem:
            mensagem_extra = f"""
            <div class="message-box">{req.mensagem}</div>
            """

        html = f"""
        <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 0; }}
                    .container {{ max-width: 600px; margin: 20px auto; background-color: #fef8f3; border-radius: 10px; padding: 34px; box-shadow: 0 2px 8px rgba(0,0,0,0.10); }}
                    .header {{ text-align: center; color: #d4844a; margin-bottom: 24px; }}
                    .title {{ font-size: 34px; font-weight: bold; margin: 8px 0; color: #cc7a3a; line-height: 1.2; }}
                    .subtitle {{ font-size: 14px; color: #888; margin-top: 4px; }}
                    .content {{ color: #333; line-height: 1.75; font-size: 15px; }}
                    .paragraph {{ margin-bottom: 14px; }}
                    .highlight {{ font-weight: bold; color: #cc7a3a; }}
                    .info-table {{ width: 100%; border-collapse: collapse; margin: 0 0 18px; border: 1px solid #eadfce; border-radius: 10px; overflow: hidden; background: #fff; }}
                    .info-table td {{ padding: 12px 14px; border-bottom: 1px solid #f0e8db; font-size: 15px; }}
                    .info-table tr:last-child td {{ border-bottom: 0; }}
                    .message-box {{ margin: 0 0 16px; padding: 12px 14px; border-radius: 10px; background: #fff7ed; border: 1px solid #fed7aa; color: #9a3412; font-size: 14px; line-height: 1.5; }}
                    .cta-wrap {{ margin: 16px 0 18px; text-align: center; }}
                    .cta-btn {{ display: inline-block; background: #1e3a5f; color: #fff !important; text-decoration: none; padding: 12px 20px; border-radius: 8px; font-weight: 700; font-size: 15px; }}
                    .link-box {{ padding: 10px 12px; border: 1px dashed #d8cbb8; border-radius: 10px; background: #fffaf4; margin-bottom: 12px; }}
                    .link-help {{ margin: 0 0 6px; font-size: 12px; color: #7d7468; }}
                    .link-url {{ margin: 0; font-size: 12px; word-break: break-all; color: #1e3a5f; }}
                    .signature {{ margin-top: 18px; text-align: center; color: #cc7a3a; font-style: italic; font-weight: bold; }}
                    .footer {{ text-align: center; margin-top: 18px; font-size: 12px; color: #999; border-top: 1px solid #e0d5c7; padding-top: 12px; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        {logo_block}
                        {festa_image_block}
                        <div class="title">Convite para Festa</div>
                        <div class="subtitle">Prepare-se para um momento especial com a UNACOB</div>
                    </div>

                    <div class="content">
                        <div class="paragraph">Olá, <strong>{membro.nome_completo}</strong>! Você está convidado(a) para a nossa confraternização.</div>
                        <div class="paragraph"><span class="highlight">{festa.nome_festa}</span></div>

                        <table class="info-table" role="presentation" cellspacing="0" cellpadding="0">
                            <tr><td><strong>Data:</strong> {data_festa_txt}</td></tr>
                            <tr><td><strong>Local:</strong> {festa.local_festa or '-'}</td></tr>
                            <tr><td><strong>Convite titular:</strong> {valor_convite_txt}</td></tr>
                            <tr><td><strong>Convite dependente:</strong> {valor_dependente_txt}</td></tr>
                        </table>

                        {mensagem_extra}

                        <div class="paragraph">Clique no botão abaixo para confirmar sua presença:</div>
                        <div class="cta-wrap">
                            <a href="{convite_link}" class="cta-btn">Confirmar participação</a>
                        </div>

                        <div class="link-box">
                            <p class="link-help">Se o botão não funcionar, copie e cole este link no navegador:</p>
                            <p class="link-url">{convite_link}</p>
                        </div>

                        <div class="signature">Com carinho e alegria,<br><strong>Diretoria da UNACOB</strong></div>
                    </div>

                    <div class="footer">© UNACOB - União dos aposentados dos correios em Bauru - SP | Festas e eventos</div>
                </div>
            </body>
        </html>
        """

        plain = (
            f"{festa.nome_festa}\n"
            f"Data: {data_festa_txt}\n"
            f"Local: {festa.local_festa or '-'}\n"
            f"Convite titular: {valor_convite_txt}\n"
            f"Convite dependente: {valor_dependente_txt}\n\n"
            f"Olá, {membro.nome_completo}.\n"
            "Use o link para confirmar sua participação:\n"
            f"{convite_link}\n"
        )

        try:
            _send_html_email(membro.email, assunto, html, plain)
            enviados += 1
        except Exception as exc:
            falhas.append({"membro": membro.nome_completo, "email": membro.email, "erro": str(exc)})

    return {
        "ok": True,
        "festa_id": festa.id,
        "somente_pendentes": bool(req.somente_pendentes),
        "total_membros_com_email": len(membros),
        "emails_enviados": enviados,
        "emails_com_falha": len(falhas),
        "falhas": falhas[:20]
    }


@app.get("/api/festas/{festa_id}/convite-link/{membro_id}")
def get_convite_link_individual(
    festa_id: str,
    membro_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    festa = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not festa:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    _validar_festa_link_disponivel(festa)

    membro = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
    if not membro:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    return {
        "ok": True,
        "festa_id": festa_id,
        "membro_id": membro_id,
        "link": _resolve_festa_public_link(festa)
    }

@app.delete("/api/festas/{festa_id}")
def delete_festa(festa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    db.delete(f)
    db.commit()
    return {"ok": True}

@app.get("/api/festas/{festa_id}/participantes")
def get_participantes(festa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    parts = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.festa_id == festa_id).all()
    result = []
    for p in parts:
        pd = {
            "id": p.id, "festa_id": p.festa_id, "membro_id": p.membro_id,
            "nome_participante": p.nome_participante, "tipo_participante": p.tipo_participante,
            "custo_convite": float(p.custo_convite) if p.custo_convite else 0,
            "pago": p.pago, "data_pagamento": str(p.data_pagamento) if p.data_pagamento else None,
            "nome_dependente": p.nome_dependente, "parentesco": p.parentesco,
            "observacoes": p.observacoes, "membro_nome": None
        }
        if p.membro_id:
            m = db.query(models.Membro).filter(models.Membro.id == p.membro_id).first()
            if m:
                pd["membro_nome"] = m.nome_completo
        result.append(pd)
    return result

@app.post("/api/festas/{festa_id}/participantes", response_model=schemas.ParticipacaoResponse)
def add_participante(festa_id: str, req: schemas.ParticipacaoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    req.festa_id = festa_id
    p = models.ParticipacaoFesta(id=str(uuid.uuid4()), created_at=datetime.utcnow(), updated_at=datetime.utcnow(), **req.dict())
    db.add(p)
    db.commit()
    db.refresh(p)
    return p

@app.put("/api/participantes/{part_id}", response_model=schemas.ParticipacaoResponse)
def update_participante(part_id: str, req: schemas.ParticipacaoUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.id == part_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Participante não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        setattr(p, k, v)
    p.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(p)
    return p

@app.delete("/api/participantes/{part_id}")
def delete_participante(part_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.id == part_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Participante não encontrado")
    db.delete(p)
    db.commit()
    return {"ok": True}


@app.get("/api/public/festas/{festa_id}")
def get_festa_publica(festa_id: str, db: Session = Depends(get_db)):
    festa = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not festa:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    _validar_festa_link_disponivel(festa)

    return {
        "ok": True,
        "festa": {
            "id": festa.id,
            "nome_festa": festa.nome_festa,
            "data_festa": str(festa.data_festa) if festa.data_festa else None,
            "local_festa": festa.local_festa,
            "beneficio_titular_dependente_gratis": True,
            "valor_convite": float(festa.valor_convite) if festa.valor_convite else 0,
            "valor_convite_dependente": float(festa.valor_convite_dependente) if festa.valor_convite_dependente else 0,
            "descricao": festa.descricao,
            "status": festa.status,
            "capacidade": festa.capacidade,
        },
    }


@app.post("/api/public/festas/{festa_id}/identificar")
def identificar_membro_festa_publica(
    festa_id: str,
    req: schemas.ParticipacaoPublicaAuthRequest,
    db: Session = Depends(get_db)
):
    festa = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not festa:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    _validar_festa_link_disponivel(festa)

    membro = _validar_membro_por_matricula_cpf(db, req.matricula, req.cpf)
    return _montar_payload_inscricao_publica(db, festa, membro)


@app.post("/api/public/festas/{festa_id}/confirmar", response_model=schemas.ParticipacaoPublicaResponse)
def confirmar_participacao_publica(
    festa_id: str,
    req: schemas.ParticipacaoPublicaConfirmRequest,
    db: Session = Depends(get_db)
):
    festa = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not festa:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    _validar_festa_link_disponivel(festa)

    membro = _validar_membro_por_matricula_cpf(db, req.matricula, req.cpf)

    if req.levar_dependente and not req.nome_dependente:
        raise HTTPException(status_code=400, detail="Informe o nome do dependente")

    if req.levar_convidado and not req.nome_convidado:
        raise HTTPException(status_code=400, detail="Informe o nome do convidado")

    idade_dependente = _normalizar_idade(req.idade_dependente, "Idade do dependente")
    idade_convidado = _normalizar_idade(req.idade_convidado, "Idade do convidado")

    valor_integral = float(festa.valor_convite or 0)
    politica_preco = _obter_politica_preco_festa(festa)
    custo_dependente = _calcular_custo_participante_por_idade(idade_dependente, valor_integral, politica_preco)
    custo_convidado = _calcular_custo_participante_por_idade(idade_convidado, valor_integral, politica_preco)

    cortesias = int(politica_preco.get("cortesia_acompanhantes", 1) or 0)
    if req.levar_dependente and cortesias > 0:
        custo_dependente = 0.0
        cortesias -= 1
    if req.levar_convidado and cortesias > 0:
        custo_convidado = 0.0

    titular = db.query(models.ParticipacaoFesta).filter(
        models.ParticipacaoFesta.festa_id == festa.id,
        models.ParticipacaoFesta.membro_id == membro.id,
        models.ParticipacaoFesta.tipo_participante == "titular"
    ).first()

    if not titular:
        titular = models.ParticipacaoFesta(
            id=str(uuid.uuid4()),
            festa_id=festa.id,
            membro_id=membro.id,
            nome_participante=membro.nome_completo,
            tipo_participante="titular",
            custo_convite=0,
            pago=False,
            observacoes=req.observacoes,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.add(titular)
    else:
        titular.nome_participante = membro.nome_completo
        titular.custo_convite = 0
        titular.observacoes = req.observacoes
        titular.updated_at = datetime.utcnow()

    dependente_id = None
    if req.levar_dependente and req.nome_dependente:
        dependente = db.query(models.ParticipacaoFesta).filter(
            models.ParticipacaoFesta.festa_id == festa.id,
            models.ParticipacaoFesta.membro_id == membro.id,
            models.ParticipacaoFesta.tipo_participante == "dependente"
        ).first()
        if not dependente:
            dependente = models.ParticipacaoFesta(
                id=str(uuid.uuid4()),
                festa_id=festa.id,
                membro_id=membro.id,
                nome_participante=membro.nome_completo,
                tipo_participante="dependente",
                custo_convite=custo_dependente,
                pago=False,
                nome_dependente=req.nome_dependente,
                parentesco=req.parentesco,
                observacoes=req.observacoes,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
            db.add(dependente)
        else:
            dependente.nome_participante = membro.nome_completo
            dependente.custo_convite = custo_dependente
            dependente.nome_dependente = req.nome_dependente
            dependente.parentesco = req.parentesco
            dependente.observacoes = req.observacoes
            dependente.updated_at = datetime.utcnow()
        dependente_id = dependente.id
    else:
        dependentes = db.query(models.ParticipacaoFesta).filter(
            models.ParticipacaoFesta.festa_id == festa.id,
            models.ParticipacaoFesta.membro_id == membro.id,
            models.ParticipacaoFesta.tipo_participante == "dependente"
        ).all()
        for dep in dependentes:
            db.delete(dep)

    convidado_id = None
    if req.levar_convidado and req.nome_convidado:
        convidado = db.query(models.ParticipacaoFesta).filter(
            models.ParticipacaoFesta.festa_id == festa.id,
            models.ParticipacaoFesta.membro_id == membro.id,
            models.ParticipacaoFesta.tipo_participante == "convidado"
        ).first()
        if not convidado:
            convidado = models.ParticipacaoFesta(
                id=str(uuid.uuid4()),
                festa_id=festa.id,
                membro_id=membro.id,
                nome_participante=req.nome_convidado,
                tipo_participante="convidado",
                custo_convite=custo_convidado,
                pago=False,
                observacoes=req.observacoes,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
            db.add(convidado)
        else:
            convidado.nome_participante = req.nome_convidado
            convidado.custo_convite = custo_convidado
            convidado.observacoes = req.observacoes
            convidado.updated_at = datetime.utcnow()
        convidado_id = convidado.id
    else:
        convidados = db.query(models.ParticipacaoFesta).filter(
            models.ParticipacaoFesta.festa_id == festa.id,
            models.ParticipacaoFesta.membro_id == membro.id,
            models.ParticipacaoFesta.tipo_participante == "convidado"
        ).all()
        for convidado in convidados:
            db.delete(convidado)

    db.commit()
    db.refresh(titular)

    return {
        "ok": True,
        "detail": "Participação confirmada com sucesso",
        "festa_id": festa.id,
        "membro_id": membro.id,
        "titular_participacao_id": titular.id,
        "dependente_participacao_id": dependente_id,
        "convidado_participacao_id": convidado_id,
    }


# ════════════════════════════════════════════════════════════════════════════════
# ANIVERSARIANTES
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/aniversariantes")
def aniversariantes(
    mes: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_filtro = mes or today.month

    membros = db.query(models.Membro).filter(
        models.Membro.data_nascimento.isnot(None),
        sql_func.strftime('%m', models.Membro.data_nascimento) == str(mes_filtro).zfill(2)
    ).order_by(sql_func.strftime('%d', models.Membro.data_nascimento)).all()

    result = []   # 🔥 ESTA LINHA ESTAVA FALTANDO

    for m in membros:
        nascimento = m.data_nascimento
        idade = today.year - nascimento.year if nascimento else None
        is_hoje = (
            nascimento.month == today.month and nascimento.day == today.day
        ) if nascimento else False

        result.append({
            "id": m.id,
            "nome": m.nome_completo,
            "data_nascimento": str(nascimento),
            "dia": nascimento.day if nascimento else None,
            "mes": nascimento.month if nascimento else None,
            "idade": idade,
            "email": m.email,
            "celular": m.celular,
            "aniversario_hoje": is_hoje
        })

    return result


@app.post("/api/aniversariantes/enviar-email")
def enviar_email_aniversario(
    req: schemas.AniversarioEmailRequest,
    current_user=Depends(get_current_user)
):
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from_name = os.getenv("SMTP_FROM_NAME", "UNACOB")
    smtp_from_email = os.getenv("SMTP_FROM_EMAIL", smtp_user or "")
    smtp_starttls = os.getenv("SMTP_STARTTLS", "true").lower() in ("1", "true", "yes", "on")

    if not smtp_host or not smtp_from_email:
        raise HTTPException(status_code=500, detail="Configuração de e-mail incompleta no servidor")

    subject = f"🎂 Feliz Aniversário, {req.nome}!"
    
    html_body = f"""
    <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 0; }}
                .container {{ max-width: 600px; margin: 20px auto; background-color: #fef8f3; border-radius: 10px; padding: 40px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
                .header {{ text-align: center; color: #d4844a; margin-bottom: 30px; }}
                .title {{ font-size: 36px; font-weight: bold; margin: 10px 0; color: #cc7a3a; }}
                .subtitle {{ font-size: 14px; color: #888; margin-top: 5px; }}
                .content {{ color: #333; line-height: 1.8; font-size: 15px; }}
                .paragraph {{ margin-bottom: 15px; text-align: justify; }}
                .highlight {{ font-weight: bold; color: #cc7a3a; }}
                .signature {{ margin-top: 30px; text-align: center; color: #cc7a3a; font-style: italic; font-weight: bold; }}
                .footer {{ text-align: center; margin-top: 20px; font-size: 12px; color: #999; border-top: 1px solid #e0d5c7; padding-top: 15px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <div style="font-size: 48px;">🎂</div>
                    <div class="title">Feliz Aniversário, {req.nome}!</div>
                    <div class="subtitle">Que sua data seja repleta de alegria e celebração</div>
                </div>
                
                <div class="content">
                    <div class="paragraph">
                        Em nome da diretoria da UNACOB, celebramos com alegria o seu aniversário! ✨
                    </div>
                    
                    <div class="paragraph">
                        Que esta data especial traga ainda mais <span class="highlight">saúde, paz</span> e momentos felizes ao lado de quem você ama.
                    </div>
                    
                    <div class="paragraph">
                        Sua trajetória é motivo de <span class="highlight">inspiração e orgulho</span> para todos nós. Aproveite cada instante deste novo ciclo e conte sempre com nossa amizade e apoio.
                    </div>
                    
                    <div class="signature">
                        Com carinho e celebração,<br>
                        <strong>Diretoria da UNACOB</strong>
                    </div>
                </div>
                
                <div class="footer">
                    © UNACOB - União dos aposentados dos correios em Bauru - SP | Celebrando momentos especiais
                </div>
            </div>
        </body>
    </html>
    """
    
    plain_text_body = (
        f"🎂 Feliz Aniversário, {req.nome}!\n\n"
        "Em nome da diretoria da UNACOB, celebramos com alegria o seu aniversário! ✨\n\n"
        "Que esta data especial traga ainda mais saúde, paz e momentos felizes ao lado de quem você ama.\n\n"
        "Sua trajetória é motivo de inspiração e orgulho para todos nós. Aproveite cada instante deste novo ciclo e "
        "conte sempre com nossa amizade e apoio.\n\n"
        "Com carinho,\n"
        "Diretoria da UNACOB"
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"{smtp_from_name} <{smtp_from_email}>"
    msg["To"] = req.email
    msg.set_content(plain_text_body)
    msg.add_alternative(html_body, subtype='html')

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as smtp:
            smtp.ehlo()
            if smtp_starttls:
                smtp.starttls()
                smtp.ehlo()

            if smtp_user and smtp_password:
                smtp.login(smtp_user, smtp_password)

            smtp.send_message(msg)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Falha ao enviar e-mail: {str(exc)}")

    return {"ok": True, "detail": "E-mail enviado com sucesso"}

# CONCILIAÇÃO
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/conciliacao", response_model=List[schemas.ConciliacaoResponse])
def list_conciliacao(
    mes_referencia: Optional[str] = None,
    conciliado: Optional[bool] = None,
    tipo: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Lista lançamentos bancários com opção de filtro por mês, status e tipo"""
    q = db.query(models.Conciliacao)
    if mes_referencia:
        q = q.filter(models.Conciliacao.mes_referencia == mes_referencia)
    if conciliado is not None:
        q = q.filter(models.Conciliacao.conciliado == conciliado)
    if tipo:
        q = q.filter(models.Conciliacao.tipo == tipo)
    return q.order_by(models.Conciliacao.data_extrato.desc()).all()


@app.get("/api/conciliacao/resumo")
def resumo_conciliacao(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    mes_ref = mes_referencia or date.today().strftime("%Y-%m")
    saldo_anterior = _saldo_anterior_mes(db, mes_ref)

    conciliacoes = db.query(models.Conciliacao).filter(
        models.Conciliacao.mes_referencia == mes_ref
    ).all()

    total_creditos = round(sum(float(c.valor_extrato or 0) for c in conciliacoes if c.tipo == "credito"), 2)
    total_debitos = round(sum(float(c.valor_extrato or 0) for c in conciliacoes if c.tipo == "debito"), 2)
    saldo_extrato = round(total_creditos - total_debitos, 2)
    saldo_final = round(saldo_anterior + saldo_extrato, 2)
    total_conciliados = len([c for c in conciliacoes if c.conciliado])

    return {
        "mes_referencia": mes_ref,
        "mes_referencia_anterior": _mes_anterior(mes_ref),
        "saldo_anterior": saldo_anterior,
        "total_creditos": total_creditos,
        "total_debitos": total_debitos,
        "saldo_extrato": saldo_extrato,
        "saldo_final": saldo_final,
        "total_lancamentos": len(conciliacoes),
        "total_conciliados": total_conciliados,
        "total_pendentes": len(conciliacoes) - total_conciliados
    }


def _montar_observacao_conciliacao_lancada(
    conciliacao: models.Conciliacao,
    observacoes_existentes: Optional[str],
    observacoes_formulario: Optional[str],
    tipo_lancamento: str,
) -> str:
    partes = []
    texto_existente = (observacoes_existentes or "").strip()
    if texto_existente:
        partes.append(texto_existente)

    marcador = (
        f"Lancado via conciliacao bancaria em {conciliacao.data_extrato.strftime('%d/%m/%Y')}"
        if conciliacao.data_extrato else
        "Lancado via conciliacao bancaria"
    )
    partes.append(f"{marcador} ({tipo_lancamento})")

    texto_formulario = (observacoes_formulario or "").strip()
    if texto_formulario:
        partes.append(texto_formulario)

    return "\n".join(partes)


def _get_conciliacao_or_404(db: Session, conc_id: str) -> models.Conciliacao:
    conciliacao = db.query(models.Conciliacao).filter(models.Conciliacao.id == conc_id).first()
    if not conciliacao:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    return conciliacao


def _normalizar_nome_busca_mensalidade(valor: Optional[str]) -> str:
    texto = unicodedata.normalize("NFKD", (valor or "")).encode("ascii", "ignore").decode().upper()
    texto = re.sub(r"^\s*\d{2}/\d{2}(?:/\d{2,4})?\s+\d{2}:\d{2}\s*", "", texto)
    texto = re.sub(r"^\s*\d+\s*", "", texto)
    texto = re.sub(r"[^A-Z\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _tokens_nome_busca_mensalidade(valor: Optional[str]) -> list[str]:
    ignorar = {"DA", "DE", "DI", "DO", "DOS", "DAS", "E"}
    tokens = []
    for token in _normalizar_nome_busca_mensalidade(valor).split():
        if len(token) <= 1 or token in ignorar:
            continue
        tokens.append(token)
    return tokens


def _sugerir_membros_para_conciliacao_credito(db: Session, conciliacao: models.Conciliacao) -> dict:
    tokens = _tokens_nome_busca_mensalidade(conciliacao.descricao_extrato)
    termo_busca = " ".join(tokens[:3]).strip()
    valor_conc = float(conciliacao.valor_extrato or 0)
    if not tokens:
        return {"termo_busca": "", "membros": []}

    membros = db.query(models.Membro).filter(models.Membro.status == "ativo").all()
    candidatos = []

    for membro in membros:
        nome_norm = _normalizar_nome_busca_mensalidade(membro.nome_completo)
        hits = sum(1 for token in tokens if token in nome_norm)
        if hits == 0:
            continue

        pagamentos = db.query(models.Pagamento).filter(
            models.Pagamento.membro_id == membro.id,
            models.Pagamento.status_pagamento.in_(["pendente", "atrasado"])
        ).order_by(models.Pagamento.mes_referencia.asc()).all()

        menor_diferenca = min(
            [abs(valor_conc - float(p.valor_pago or 0)) for p in pagamentos] or [0.0]
        )
        score = (hits * 100) - menor_diferenca
        candidatos.append({
            "membro_id": membro.id,
            "nome": membro.nome_completo,
            "email": membro.email,
            "cpf": membro.cpf,
            "quantidade_pendente": len(pagamentos) or 1,
            "total_pendente": round(sum(float(p.valor_pago or 0) for p in pagamentos), 2) or round(valor_conc, 2),
            "menor_diferenca": round(menor_diferenca, 2),
            "score": round(score, 2),
            "hits_nome": hits,
        })

    candidatos.sort(key=lambda item: (-item["score"], item["menor_diferenca"], item["nome"] or ""))
    return {"termo_busca": termo_busca, "membros": candidatos[:10]}


@app.post("/api/conciliacao", response_model=schemas.ConciliacaoResponse)
def create_conciliacao(
    req: schemas.ConciliacaoCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Cria novo lançamento bancário manual"""
    mes_ref = req.mes_referencia or req.data_extrato.strftime("%Y-%m")
    c = models.Conciliacao(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        mes_referencia=mes_ref,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        **req.dict(exclude={"mes_referencia"})
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


# ⭐ ENDPOINTS ESPECÍFICOS (devem vir antes de /{conc_id})
@app.get("/api/conciliacao/membros/buscar")
def buscar_membros_com_pagamentos(
    q: str = "",
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Busca membros com pagamentos pendentes/atrasados"""
    try:
        query = db.query(models.Membro).filter(
            models.Membro.nome_completo.ilike(f"%{q}%")
        ).all()

        resultado = []
        for membro in query:
            pagamentos_pendentes = db.query(models.Pagamento).filter(
                models.Pagamento.membro_id == membro.id,
                models.Pagamento.status_pagamento.in_(["pendente", "atrasado"])
            ).all()

            resultado.append({
                "membro_id": membro.id,
                "nome": membro.nome_completo,
                "email": membro.email,
                "cpf": membro.cpf,
                "total_pendente": sum(float(p.valor_pago or 0) for p in pagamentos_pendentes) or float(membro.valor_mensalidade or 0),
                "quantidade_pendente": len(pagamentos_pendentes) or 1
            })

        return resultado[:20]  # Limita a 20 resultados
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao buscar membros: {str(e)}")


@app.get("/api/conciliacao/membro/{membro_id}/pagamentos-pendentes")
def listar_pagamentos_pendentes_membro(
    membro_id: str,
    mes_referencia: Optional[str] = None,
    valor: Optional[float] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Lista todos os pagamentos pendentes/atrasados de um membro"""
    try:
        membro = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
        if not membro:
            raise HTTPException(status_code=404, detail="Membro não encontrado")

        pagamentos = db.query(models.Pagamento).filter(
            models.Pagamento.membro_id == membro_id,
            models.Pagamento.status_pagamento.in_(["pendente", "atrasado"])
        ).order_by(models.Pagamento.mes_referencia.asc()).all()

        pagamentos_payload = [
            {
                "pagamento_id": p.id,
                "mes": p.mes_referencia,
                "valor": float(p.valor_pago or 0),
                "status": p.status_pagamento,
                "data_previsto": str(p.data_pagamento) if p.data_pagamento else None
            }
            for p in pagamentos
        ]

        if not pagamentos_payload and mes_referencia:
            pagamentos_payload.append({
                "pagamento_id": None,
                "mes": mes_referencia,
                "valor": float(valor if valor is not None else (membro.valor_mensalidade or 0)),
                "status": "nao_gerado",
                "data_previsto": None,
                "membro_id": membro.id,
            })

        return {
            "membro_id": membro.id,
            "nome": membro.nome_completo,
            "email": membro.email,
            "pagamentos": pagamentos_payload
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao listar pagamentos: {str(e)}")


# ⭐ ENDPOINTS COM PARÂMETRO {conc_id}
@app.get("/api/conciliacao/{conc_id}/sugestoes-mensalidade")
def sugerir_membros_para_mensalidade(
    conc_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    conciliacao = _get_conciliacao_or_404(db, conc_id)
    if conciliacao.tipo != "credito":
        return {"termo_busca": "", "membros": []}
    return _sugerir_membros_para_conciliacao_credito(db, conciliacao)


@app.get("/api/conciliacao/{conc_id}/sugestoes")
def sugerir_matching_pagamentos(
    conc_id: str,
    tolerance: float = 0.01,  # tolerância de R$ 0,01
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Sugere pagamentos para matching baseado no valor do extrato"""
    c = db.query(models.Conciliacao).filter(models.Conciliacao.id == conc_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")

    # Se é débito (saída), procura por pagamentos devidos (pendentes/atrasados)
    if c.tipo == "debito":
        valor_min = float(c.valor_extrato) - tolerance
        valor_max = float(c.valor_extrato) + tolerance

        pagamentos = db.query(models.Pagamento).join(
            models.Membro, models.Pagamento.membro_id == models.Membro.id
        ).filter(
            models.Pagamento.status_pagamento.in_(["pendente", "atrasado"]),
            models.Pagamento.valor_pago >= valor_min,
            models.Pagamento.valor_pago <= valor_max
        ).all()

        sugestoes = []
        for p in pagamentos:
            membro = db.query(models.Membro).filter(models.Membro.id == p.membro_id).first()
            sugestoes.append({
                "pagamento_id": p.id,
                "membro_nome": membro.nome_completo if membro else "?",
                "mes": p.mes_referencia,
                "valor": float(p.valor_pago),
                "diferenca": abs(float(c.valor_extrato) - float(p.valor_pago))
            })

        return {"total": len(sugestoes), "sugestoes": sugestoes}
    
    return {"total": 0, "sugestoes": []}


@app.post("/api/conciliacao/{conc_id}/reconciliar")
def reconciliar_pagamento(
    conc_id: str,
    req: schemas.ConciliacaoImportRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Vincula extrato bancário a um pagamento e marca como conciliado"""
    c = db.query(models.Conciliacao).filter(models.Conciliacao.id == conc_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")

    if req.pagamento_id:
        pag = db.query(models.Pagamento).filter(models.Pagamento.id == req.pagamento_id).first()
        if not pag:
            raise HTTPException(status_code=404, detail="Pagamento não encontrado")

        # Vincula e marca como conciliado
        c.pagamento_id = req.pagamento_id
        c.conciliado = True
        c.updated_at = datetime.utcnow()

        # Atualiza pagamento
        pag.status_pagamento = "pago"
        pag.data_pagamento = c.data_extrato
        pag.forma_pagamento = "transferencia"
        pag.updated_at = datetime.utcnow()
    elif req.membro_id:
        membro = db.query(models.Membro).filter(models.Membro.id == req.membro_id).first()
        if not membro:
            raise HTTPException(status_code=404, detail="Membro não encontrado")
        _baixar_pagamento_mensalidade_por_conciliacao(
            db=db,
            conciliacao=c,
            membro=membro,
            user_id=current_user.id,
            observacao_origem="Baixa manual via conciliação OFX",
        )
    else:
        raise HTTPException(status_code=400, detail="pagamento_id ou membro_id obrigatório")

    db.commit()
    return {"ok": True, "detail": "Pagamento reconciliado com sucesso"}


@app.post("/api/conciliacao/{conc_id}/lancar-despesa", response_model=schemas.ConciliacaoResponse)
def lancar_conciliacao_em_despesa(
    conc_id: str,
    req: schemas.ConciliacaoLancarDespesaRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    conciliacao = _get_conciliacao_or_404(db, conc_id)
    if conciliacao.tipo != "debito":
        raise HTTPException(status_code=400, detail="Apenas lançamentos de débito podem ser enviados para despesas")
    if conciliacao.despesa_id:
        raise HTTPException(status_code=400, detail="Este lançamento já foi enviado para despesas")

    conta = _get_conta_or_400(db, req.conta_id, "saida")
    categoria = (req.categoria or conta.nome or "Outros").strip()
    mes_ref = conciliacao.mes_referencia or (
        conciliacao.data_extrato.strftime("%Y-%m") if conciliacao.data_extrato else date.today().strftime("%Y-%m")
    )
    observacoes = _montar_observacao_conciliacao_lancada(
        conciliacao,
        conciliacao.observacoes,
        req.observacoes,
        "despesa",
    )

    despesa = models.Despesa(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        descricao=(conciliacao.descricao_extrato or "Despesa importada do extrato").strip(),
        categoria=categoria,
        conta_id=conta.id,
        conta_codigo=conta.codigo,
        conta_nome=conta.nome,
        valor=float(conciliacao.valor_extrato or 0),
        data_despesa=conciliacao.data_extrato or date.today(),
        mes_referencia=mes_ref,
        forma_pagamento=(req.forma_pagamento or "extrato_bancario").strip() or "extrato_bancario",
        fornecedor=(req.fornecedor or "").strip() or None,
        nota_fiscal=(req.nota_fiscal or "").strip() or None,
        observacoes=observacoes,
        created_at=datetime.utcnow(),
    )
    db.add(despesa)
    db.flush()
    _sync_transacao_despesa(db, despesa, current_user.id)

    conciliacao.despesa_id = despesa.id
    conciliacao.conciliado = True
    conciliacao.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(conciliacao)
    return conciliacao


@app.post("/api/conciliacao/{conc_id}/lancar-receita", response_model=schemas.ConciliacaoResponse)
def lancar_conciliacao_em_receita(
    conc_id: str,
    req: schemas.ConciliacaoLancarReceitaRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    conciliacao = _get_conciliacao_or_404(db, conc_id)
    if conciliacao.tipo != "credito":
        raise HTTPException(status_code=400, detail="Apenas lançamentos de crédito podem ser enviados para receitas")
    if conciliacao.outra_renda_id:
        raise HTTPException(status_code=400, detail="Este lançamento já foi enviado para receitas")
    if conciliacao.pagamento_id:
        raise HTTPException(status_code=400, detail="Este crédito já está vinculado a uma mensalidade")

    conta = _get_conta_or_400(db, req.conta_id, "entrada")
    categoria = (req.categoria or conta.nome or "Outros").strip()
    mes_ref = conciliacao.mes_referencia or (
        conciliacao.data_extrato.strftime("%Y-%m") if conciliacao.data_extrato else date.today().strftime("%Y-%m")
    )
    observacoes = _montar_observacao_conciliacao_lancada(
        conciliacao,
        conciliacao.observacoes,
        req.observacoes,
        "receita",
    )

    renda = models.OutraRenda(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        descricao=(conciliacao.descricao_extrato or "Receita importada do extrato").strip(),
        categoria=categoria,
        conta_id=conta.id,
        conta_codigo=conta.codigo,
        conta_nome=conta.nome,
        valor=float(conciliacao.valor_extrato or 0),
        data_recebimento=conciliacao.data_extrato or date.today(),
        mes_referencia=mes_ref,
        fonte=(req.fonte or conciliacao.banco or "Extrato bancario").strip(),
        observacoes=observacoes,
        created_at=datetime.utcnow(),
    )
    db.add(renda)
    db.flush()
    _sync_transacao_outra_renda(db, renda, current_user.id)

    conciliacao.outra_renda_id = renda.id
    conciliacao.conciliado = True
    conciliacao.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(conciliacao)
    return conciliacao


@app.post("/api/conciliacao/processar-ofx/{mes_referencia}")
def processar_conciliacao_ofx_mes(
    mes_referencia: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    conciliacoes = db.query(models.Conciliacao).filter(
        models.Conciliacao.mes_referencia == mes_referencia
    ).order_by(models.Conciliacao.data_extrato.asc(), models.Conciliacao.created_at.asc()).all()

    total_creditos_baixados = 0
    total_debitos_lancados = 0
    creditos_sem_match = 0
    debitos_sem_conta = 0

    for conciliacao in conciliacoes:
        if conciliacao.tipo == "credito" and not conciliacao.pagamento_id and _descricao_credito_parece_mensalidade_ofx(conciliacao.descricao_extrato):
            membro = _escolher_membro_para_credito_ofx(db, conciliacao)
            if membro:
                _baixar_pagamento_mensalidade_por_conciliacao(
                    db=db,
                    conciliacao=conciliacao,
                    membro=membro,
                    user_id=current_user.id,
                    observacao_origem=f"Baixa automática via processamento OFX ({mes_referencia})",
                )
                total_creditos_baixados += 1
            else:
                creditos_sem_match += 1

        if conciliacao.tipo == "debito" and not conciliacao.despesa_id:
            conta = _inferir_conta_despesa_ofx(db, conciliacao.descricao_extrato)
            if conta:
                _lancar_despesa_por_conciliacao(
                    db=db,
                    conciliacao=conciliacao,
                    conta=conta,
                    current_user=current_user,
                    fornecedor=conciliacao.descricao_extrato,
                    forma_pagamento="extrato_bancario",
                    observacoes_extra=f"Lançamento automático via processamento OFX ({mes_referencia})",
                )
                total_debitos_lancados += 1
            else:
                debitos_sem_conta += 1

    db.commit()
    return {
        "ok": True,
        "mes_referencia": mes_referencia,
        "total_creditos_baixados": total_creditos_baixados,
        "total_debitos_lancados": total_debitos_lancados,
        "creditos_sem_match": creditos_sem_match,
        "debitos_sem_conta": debitos_sem_conta,
    }


@app.put("/api/conciliacao/{conc_id}", response_model=schemas.ConciliacaoResponse)
def update_conciliacao(
    conc_id: str,
    req: schemas.ConciliacaoUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Atualiza lançamento: vincula pagamento ou altera status"""
    c = db.query(models.Conciliacao).filter(models.Conciliacao.id == conc_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")

    nova_data_extrato = req.data_extrato or c.data_extrato

    # Se vincular um pagamento e marcar como conciliado, atualiza o pagamento
    if req.pagamento_id and req.conciliado:
        pag = db.query(models.Pagamento).filter(models.Pagamento.id == req.pagamento_id).first()
        if pag:
            pag.status_pagamento = "pago"
            pag.data_pagamento = nova_data_extrato
            pag.forma_pagamento = "transferencia"  # default para extrato bancário
            pag.updated_at = datetime.utcnow()

    for k, v in req.dict(exclude_none=True).items():
        setattr(c, k, v)
    c.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(c)
    return c


@app.delete("/api/conciliacao/{conc_id}")
def delete_conciliacao(
    conc_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Remove lançamento bancário"""
    c = db.query(models.Conciliacao).filter(models.Conciliacao.id == conc_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    db.delete(c)
    db.commit()
    return {"ok": True}


def _parse_valor_extrato(valor_raw: str) -> float:
    valor_limpo = (valor_raw or "").replace("R$", "").replace(" ", "").strip()
    if not valor_limpo:
        return 0.0

    if "," in valor_limpo and "." in valor_limpo:
        valor_limpo = valor_limpo.replace(".", "").replace(",", ".")
    elif "," in valor_limpo:
        valor_limpo = valor_limpo.replace(",", ".")

    return float(valor_limpo)


def _parse_data_ofx(data_raw: str) -> date:
    digits = "".join(ch for ch in (data_raw or "") if ch.isdigit())
    if len(digits) < 8:
        raise ValueError("DATA OFX invalida")
    return datetime.strptime(digits[:8], "%Y%m%d").date()


def _parse_data_extrato(data_raw: str) -> date:
    valor = (data_raw or "").strip()
    if not valor:
        raise ValueError("DATA invalida")

    # Remove timezone/textos adicionais comuns em exportacoes.
    valor = valor.replace("T", " ").split("[")[0].strip()

    formatos = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(valor, fmt).date()
        except Exception:
            continue

    # Fallback: pega apenas a parte de data antes do espaco.
    primeira_parte = valor.split(" ")[0]
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
        try:
            return datetime.strptime(primeira_parte, fmt).date()
        except Exception:
            continue

    raise ValueError("DATA invalida")


def _extrair_data_valida_yyyymmdd(valor_raw: str) -> Optional[date]:
    candidatos = re.findall(r"20\d{6}", valor_raw or "")
    for candidato in candidatos:
        try:
            return datetime.strptime(candidato, "%Y%m%d").date()
        except ValueError:
            continue
    return None


def _extrair_data_header_dabb(texto: str) -> Optional[date]:
    for raw_line in (texto or "").splitlines():
        line = raw_line.rstrip("\r\n").lstrip("\ufeff ").rstrip()
        if not line.startswith("A"):
            continue

        # Tenta primeiro datas YYYYMMDD do cabeçalho.
        data_header = _extrair_data_valida_yyyymmdd(line)
        if data_header:
            return data_header

        # Fallback para DDMMAAAA caso o banco use outro layout.
        candidatos = re.findall(r"\d{8}", line)
        for candidato in candidatos:
            try:
                return datetime.strptime(candidato, "%d%m%Y").date()
            except ValueError:
                continue
    return None


def _normalizar_header_csv(valor: str) -> str:
    norm = unicodedata.normalize("NFKD", (valor or "")).encode("ascii", "ignore").decode().lower().strip()
    norm = re.sub(r"\s+", " ", norm)
    return norm


def _escolher_delimitador_csv(header_line: str) -> str:
    if header_line.count(";") > header_line.count(","):
        return ";"
    if header_line.count("\t") > 0:
        return "\t"
    return ","


def _encontrar_inicio_csv(lines: List[str]) -> int:
    candidatos = [
        "data", "descricao", "descricao lancamento", "historico", "detalhes", "lancamento", "valor", "tipo"
    ]
    for idx, line in enumerate(lines[:20]):
        header_norm = _normalizar_header_csv(line)
        hits = sum(1 for c in candidatos if c in header_norm)
        if hits >= 2:
            return idx
    return 0


def _valor_por_alias(row: dict, aliases: List[str]) -> str:
    for key, value in row.items():
        key_norm = _normalizar_header_csv(key)
        if key_norm in aliases:
            return (value or "").strip()
    return ""


def _normalizar_texto_chave(valor: Optional[str]) -> str:
    txt = unicodedata.normalize("NFKD", (valor or "")).encode("ascii", "ignore").decode().lower().strip()
    return re.sub(r"\s+", " ", txt)


def _buscar_duplicado_conciliacao(
    db: Session,
    data_extrato: date,
    valor_extrato: float,
    banco: str,
    tipo: str,
    numero_documento: Optional[str],
    descricao_extrato: Optional[str],
):
    valor_abs = abs(float(valor_extrato or 0))
    banco_norm = (banco or "").strip()
    tipo_norm = (tipo or "").strip().lower()
    numero_doc_norm = (numero_documento or "").strip()
    descricao_norm = _normalizar_texto_chave(descricao_extrato)

    q = db.query(models.Conciliacao).filter(
        models.Conciliacao.data_extrato == data_extrato,
        models.Conciliacao.valor_extrato == valor_abs,
        models.Conciliacao.banco == banco_norm,
        models.Conciliacao.tipo == tipo_norm,
    )

    # Para DABB, o mesmo recebimento pode aparecer em REM e RET com numero_documento
    # diferente. Nesse caso a chave prática é descrição + data + valor + tipo + banco.
    if descricao_norm.startswith("mensalidade dabb "):
        candidatos_dabb = q.all()
        for cand in candidatos_dabb:
            if _normalizar_texto_chave(cand.descricao_extrato) == descricao_norm:
                return cand

    if numero_doc_norm:
        return q.filter(models.Conciliacao.numero_documento == numero_doc_norm).first()

    candidatos = q.filter(
        or_(
            models.Conciliacao.numero_documento.is_(None),
            models.Conciliacao.numero_documento == "",
        )
    ).all()
    for cand in candidatos:
        if _normalizar_texto_chave(cand.descricao_extrato) == descricao_norm:
            return cand
    return None


def _descricao_indica_linha_saldo(descricao: Optional[str]) -> bool:
    txt = _normalizar_texto_chave(descricao)
    if not txt:
        return False

    if txt in {"saldo", "saldo do dia", "saldo anterior"}:
        return True

    if txt.replace(" ", "") == "saldo":
        return True

    return txt.startswith("saldo ") or " saldo " in f" {txt} "


def _extrair_tag_ofx(bloco: str, tag: str) -> str:
    match = re.search(rf"<{tag}>([^<\r\n]+)", bloco, re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _decode_uploaded_text(contents: bytes) -> str:
    # Alguns bancos exportam OFX em UTF-16 ou UTF-8 com BOM.
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            return contents.decode(encoding)
        except Exception:
            continue
    return contents.decode("utf-8", errors="ignore")


def _normalizar_codigo_dabb(valor: Optional[str]) -> str:
    return re.sub(r"\D", "", (valor or "").strip())


def _variantes_codigo_dabb(valor: Optional[str]) -> set[str]:
    codigo = _normalizar_codigo_dabb(valor)
    if not codigo:
        return set()
    sem_zeros = codigo.lstrip("0")
    variantes = {codigo}
    if sem_zeros:
        variantes.add(sem_zeros)
    return variantes


def _extrair_codigo_dabb_das_observacoes(observacoes: Optional[str]) -> str:
    texto = observacoes or ""
    match = re.search(r"codigo_dabb=(\d+)", texto)
    return _normalizar_codigo_dabb(match.group(1) if match else "")


def _indexar_membros_por_codigo_dabb(db: Session) -> dict[str, list[models.Membro]]:
    indice = defaultdict(list)
    membros = db.query(models.Membro).filter(
        models.Membro.status == "ativo",
        models.Membro.codigo_dabb.isnot(None),
        models.Membro.codigo_dabb != ""
    ).all()
    for membro in membros:
        for variante in _variantes_codigo_dabb(membro.codigo_dabb):
            indice[variante].append(membro)
    return indice


def _parse_valor_dabb(valor_raw: str) -> float:
    digits = re.sub(r"\D", "", (valor_raw or "").strip())
    if not digits:
        raise ValueError("VALOR DABB invalido")
    # Layout bancario costuma enviar 3 casas implicitas.
    return round(int(digits) / 1000, 2)


def _extrair_codigo_barras_dabb(payload: str, data_tx: date) -> Optional[str]:
    payload = re.sub(r"\D", "", (payload or "").strip())
    if not payload:
        return None

    data_token = data_tx.strftime("%Y%m%d")
    data_idx = payload.find(data_token)
    if data_idx <= 0:
        return None

    codigo_barras = payload[:data_idx]
    return codigo_barras or None


def _iterar_transacoes_dabb(texto: str):
    for raw_line in texto.splitlines():
        line = raw_line.rstrip("\r\n")
        line_normalized = line.lstrip("\ufeff ").rstrip()
        if not line_normalized or line_normalized[:1] not in {"E", "F"}:
            continue

        match = re.match(r"^(?P<tipo_registro>[EF])(?P<codigo>\d{8,20})\s+(?P<payload>\d+)\s*(?P<status>\d)?$", line_normalized)
        if not match:
            continue

        codigo_dabb = _normalizar_codigo_dabb(match.group("codigo"))
        payload = match.group("payload") or ""
        data_tx = _extrair_data_dabb_payload(payload)
        if not codigo_dabb or not data_tx:
            continue

        codigo_barras = _extrair_codigo_barras_dabb(payload, data_tx)
        data_token = data_tx.strftime("%Y%m%d")
        data_idx = payload.find(data_token)
        if data_idx < 0:
            continue

        depois_data = payload[data_idx + len(data_token):]
        # Nos arquivos DABB de mensalidade, o valor vem logo após a data
        # com 16 dígitos implícitos e 3 casas decimais: 0000000000067000 -> 67,00.
        if len(depois_data) < 16:
            continue

        valor_tx = _parse_valor_dabb(depois_data[:16])
        numero_documento = depois_data[16:] or None

        yield {
            "data": data_tx,
            "valor": valor_tx,
            "tipo": "credito",
            "descricao": f"Mensalidade DABB {codigo_dabb}",
            "numero_documento": numero_documento,
            "codigo_dabb": codigo_dabb,
            "codigo_barras": codigo_barras,
            "linha_original": line_normalized,
            "tipo_registro": match.group("tipo_registro"),
        }


def _extrair_data_dabb_payload(payload: str, data_fallback: Optional[date] = None) -> Optional[date]:
    # No layout observado, a data costuma iniciar na posição 18 do payload.
    if len(payload) >= 26:
        candidato_fixo = payload[18:26]
        for formato in ("%Y%m%d", "%d%m%Y"):
            try:
                return datetime.strptime(candidato_fixo, formato).date()
            except ValueError:
                continue

    candidatos = []

    candidatos.extend(re.findall(r"20\d{6}", payload or ""))

    # Alguns arquivos podem trazer DDMMAAAA.
    candidatos.extend(re.findall(r"\d{8}", payload or ""))

    vistos = set()
    for candidato in candidatos:
        if candidato in vistos:
            continue
        vistos.add(candidato)
        for formato in ("%Y%m%d", "%d%m%Y"):
            try:
                return datetime.strptime(candidato, formato).date()
            except ValueError:
                continue

    return data_fallback


def _parse_linha_dabb(line: str, data_fallback: Optional[date] = None) -> dict:
    line_normalized = (line or "").rstrip("\r\n").lstrip("\ufeff ").rstrip()
    if not line_normalized or line_normalized[:1] not in {"E", "F"}:
        raise ValueError("registro_nao_detalhe")

    match = re.match(r"^(?P<tipo_registro>[EF])(?P<codigo>\d{8,20})\s+(?P<payload>\d+)\s*(?P<status>\d)?$", line_normalized)
    if not match:
        raise ValueError("layout_invalido")

    codigo_dabb = _normalizar_codigo_dabb(match.group("codigo"))
    if not codigo_dabb:
        raise ValueError("codigo_dabb_invalido")

    payload = match.group("payload") or ""
    data_tx = _extrair_data_dabb_payload(payload, data_fallback=data_fallback)
    if not data_tx:
        raise ValueError("data_invalida")

    codigo_barras = _extrair_codigo_barras_dabb(payload, data_tx)
    data_token = data_tx.strftime("%Y%m%d")
    data_idx = payload.find(data_token)
    if data_idx < 0:
        raise ValueError("data_nao_localizada")

    depois_data = payload[data_idx + len(data_token):]
    if len(depois_data) < 16:
        raise ValueError("valor_invalido")

    valor_tx = _parse_valor_dabb(depois_data[:16])
    numero_documento = depois_data[16:] or None

    return {
        "data": data_tx,
        "valor": valor_tx,
        "tipo": "credito",
        "descricao": f"Mensalidade DABB {codigo_dabb}",
        "numero_documento": numero_documento,
        "codigo_dabb": codigo_dabb,
        "codigo_barras": codigo_barras,
        "linha_original": line_normalized,
        "tipo_registro": match.group("tipo_registro"),
    }


def _iterar_transacoes_pdf_bb(texto: str):
    blocos = [
        bloco for bloco in re.split(r"(?=Nome[. ]*:)", texto or "")
        if bloco.strip().startswith("Nome")
    ]

    for bloco in blocos:
        nome_match = re.search(r"Nome[. ]*:\s*(.*)", bloco, flags=re.IGNORECASE)
        codigo_match = re.search(
            r"Identifica(?:ç|c)ão P/d[ée]bito:\s*(\d+)",
            bloco,
            flags=re.IGNORECASE,
        )
        banco_match = re.search(r"Banco[. ]*:\s*(.*)", bloco, flags=re.IGNORECASE)
        data_valor_match = re.search(
            r"Data[. ]*:\s*(\d{2}/\d{2}/\d{4})\s+Valor:\s*([\d.,]+)",
            bloco,
            flags=re.IGNORECASE,
        )

        codigo_dabb = _normalizar_codigo_dabb(codigo_match.group(1) if codigo_match else None)
        data_raw = (data_valor_match.group(1) if data_valor_match else "").strip()
        valor_raw = (data_valor_match.group(2) if data_valor_match else "").strip()

        try:
            data_tx = datetime.strptime(data_raw, "%d/%m/%Y").date()
            valor_tx = _parse_valor_extrato(valor_raw)
        except Exception:
            data_tx = None
            valor_tx = None

        yield {
            "data": data_tx,
            "valor": abs(float(valor_tx or 0)) if valor_tx is not None else None,
            "tipo": "credito",
            "descricao": f"Mensalidade DABB {codigo_dabb}" if codigo_dabb else "Mensalidade DABB",
            "numero_documento": None,
            "codigo_dabb": codigo_dabb,
            "nome": (nome_match.group(1) if nome_match else "").strip(),
            "banco_linha": (banco_match.group(1) if banco_match else "").strip(),
            "bloco_original": bloco,
            "tipo_registro": "PDF",
        }


def _baixar_pagamento_mensalidade_por_conciliacao(
    db: Session,
    conciliacao: models.Conciliacao,
    membro: models.Membro,
    user_id: str,
    observacao_origem: str,
):
    mes_ref = conciliacao.mes_referencia or (
        conciliacao.data_extrato.strftime("%Y-%m") if conciliacao.data_extrato else None
    )
    if not mes_ref:
        raise ValueError("Mes de referencia nao identificado para a conciliacao")

    pagamento = db.query(models.Pagamento).filter(
        models.Pagamento.membro_id == membro.id,
        models.Pagamento.mes_referencia == mes_ref
    ).first()

    observacao = observacao_origem.strip()
    if pagamento:
        pagamento.valor_pago = float(conciliacao.valor_extrato or 0)
        pagamento.status_pagamento = "pago"
        pagamento.data_pagamento = conciliacao.data_extrato
        pagamento.forma_pagamento = "debito_automatico"
        pagamento.observacoes = (
            (pagamento.observacoes + "\n") if pagamento.observacoes else ""
        ) + observacao
        pagamento.updated_at = datetime.utcnow()
    else:
        pagamento = models.Pagamento(
            id=str(uuid.uuid4()),
            membro_id=membro.id,
            valor_pago=float(conciliacao.valor_extrato or 0),
            mes_referencia=mes_ref,
            data_pagamento=conciliacao.data_extrato,
            status_pagamento="pago",
            forma_pagamento="debito_automatico",
            observacoes=observacao,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        db.add(pagamento)
        db.flush()

    conciliacao.pagamento_id = pagamento.id
    conciliacao.conciliado = True
    conciliacao.updated_at = datetime.utcnow()

    _register_transaction(db, pagamento, user_id)
    return pagamento


def _descricao_credito_parece_mensalidade_ofx(descricao: Optional[str]) -> bool:
    texto = _normalizar_texto_chave(descricao)
    if not texto:
        return False

    bloqueios = [
        "bb rf",
        "empresa agil",
        "resgate",
        "aplicacao",
        "invest",
        "cdb",
        "rendimento",
        "estorno",
        "tarifa",
        "juros",
    ]
    if any(item in texto for item in bloqueios):
        return False

    return len(_tokens_nome_busca_mensalidade(descricao)) >= 2


def _escolher_membro_para_credito_ofx(db: Session, conciliacao: models.Conciliacao) -> Optional[models.Membro]:
    sugestoes = _sugerir_membros_para_conciliacao_credito(db, conciliacao).get("membros", [])
    if not sugestoes:
        return None

    melhor = sugestoes[0]
    segundo = sugestoes[1] if len(sugestoes) > 1 else None
    score_gap = (melhor.get("score") or 0) - ((segundo or {}).get("score") or 0)

    if melhor.get("hits_nome", 0) < 2:
        return None
    if melhor.get("menor_diferenca", 9999) > 0.5:
        return None
    if segundo and score_gap < 80:
        return None

    return db.query(models.Membro).filter(models.Membro.id == melhor["membro_id"]).first()


def _conta_por_codigo(db: Session, codigo: str, tipo: str) -> Optional[models.PlanoConta]:
    codigo_norm = _normalizar_codigo_conta_seed(codigo)
    tipo_norm = _validar_tipo_conta(tipo)
    return db.query(models.PlanoConta).filter(
        models.PlanoConta.codigo == codigo_norm,
        models.PlanoConta.tipo == tipo_norm
    ).first()


def _inferir_conta_despesa_ofx(db: Session, descricao: Optional[str]) -> Optional[models.PlanoConta]:
    texto = _normalizar_texto_chave(descricao)
    mapa = [
        (["cpfl", "energia"], "2.21"),
        (["dae", "agua"], "2.22"),
        (["vivo", "telefone", "internet", "claro", "tim", "oi"], "2.24"),
        (["cobranca referente", "tarifa", "estorno solucao imediata", "solucao imediata", "iof", "juros", "taxa bancaria"], "2.23"),
    ]
    for termos, codigo in mapa:
        if any(termo in texto for termo in termos):
            conta = _conta_por_codigo(db, codigo, "saida")
            if conta:
                return conta
    return None


def _lancar_despesa_por_conciliacao(
    db: Session,
    conciliacao: models.Conciliacao,
    conta: models.PlanoConta,
    current_user: models.User,
    fornecedor: Optional[str] = None,
    forma_pagamento: Optional[str] = "extrato_bancario",
    observacoes_extra: Optional[str] = None,
):
    if conciliacao.despesa_id:
        existente = db.query(models.Despesa).filter(models.Despesa.id == conciliacao.despesa_id).first()
        if existente:
            return existente

    categoria = (conta.nome or "Outros").strip()
    mes_ref = conciliacao.mes_referencia or (
        conciliacao.data_extrato.strftime("%Y-%m") if conciliacao.data_extrato else date.today().strftime("%Y-%m")
    )
    observacoes = _montar_observacao_conciliacao_lancada(
        conciliacao,
        conciliacao.observacoes,
        observacoes_extra,
        "despesa",
    )

    despesa = models.Despesa(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        descricao=(conciliacao.descricao_extrato or "Despesa importada do extrato").strip(),
        categoria=categoria,
        conta_id=conta.id,
        conta_codigo=conta.codigo,
        conta_nome=conta.nome,
        valor=float(conciliacao.valor_extrato or 0),
        data_despesa=conciliacao.data_extrato or date.today(),
        mes_referencia=mes_ref,
        forma_pagamento=(forma_pagamento or "extrato_bancario").strip() or "extrato_bancario",
        fornecedor=(fornecedor or "").strip() or None,
        observacoes=observacoes,
        created_at=datetime.utcnow(),
    )
    db.add(despesa)
    db.flush()
    _sync_transacao_despesa(db, despesa, current_user.id)

    conciliacao.despesa_id = despesa.id
    conciliacao.conciliado = True
    conciliacao.updated_at = datetime.utcnow()
    return despesa


def _iterar_transacoes_ofx(texto: str):
    # Suporta OFX com e sem fechamento explicito de </STMTTRN>.
    padrao = r"<STMTTRN>(.*?)(?=(</STMTTRN>|<STMTTRN>|</BANKTRANLIST>|$))"
    for match in re.finditer(padrao, texto, flags=re.IGNORECASE | re.DOTALL):
        bloco = match.group(1)
        data_str = _extrair_tag_ofx(bloco, "DTPOSTED")
        valor_str = _extrair_tag_ofx(bloco, "TRNAMT")
        if not data_str or not valor_str:
            continue

        data_tx = _parse_data_ofx(data_str)
        valor_tx = _parse_valor_extrato(valor_str)
        trntype = _extrair_tag_ofx(bloco, "TRNTYPE").lower()
        fitid = _extrair_tag_ofx(bloco, "FITID")
        memo = _extrair_tag_ofx(bloco, "MEMO")
        name = _extrair_tag_ofx(bloco, "NAME")
        descricao = memo or name or trntype or "Lancamento bancario"

        if trntype in {"debit", "payment", "check", "atm", "pos", "fee"}:
            tipo = "debito"
        elif trntype in {"credit", "dep", "directdep", "int", "div"}:
            tipo = "credito"
        else:
            tipo = "credito" if valor_tx >= 0 else "debito"

        yield {
            "data": data_tx,
            "valor": abs(valor_tx),
            "tipo": tipo,
            "descricao": descricao,
            "numero_documento": fitid or None,
        }

@app.post("/api/conciliacao/importar/csv")
@app.post("/api/conciliacao/importar/extrato")
async def importar_extrato_arquivo(
    file: UploadFile = File(...),
    banco: str = "Importado",
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """
    Importa extrato bancário em formato CSV, OFX, REM ou RET.
    CSV suportado:
    1. Simples: data, descricao, tipo (credito|debito), valor
    2. Banco Real: Data, Lançamento, Detalhes, Nº documento, Valor, Tipo Lançamento
    REM/RET:
    - Registros de mensalidade com codigo DABB na linha de detalhe "E"
    - Faz baixa automática na mensalidade usando Membro.codigo_dabb
    """
    try:
        import csv

        contents = await file.read()
        text = _decode_uploaded_text(contents)

        text = text.strip()
        if not text:
            raise HTTPException(status_code=400, detail="Arquivo vazio")

        ext = Path((file.filename or "").strip().lower()).suffix
        is_ofx_content = "<OFX" in text.upper() or "<STMTTRN>" in text.upper()

        if ext not in {".csv", ".ofx", ".ret", ".rem"} and not is_ofx_content:
            raise HTTPException(status_code=400, detail="Arquivo deve ser CSV, OFX, RET ou REM")

        importados = []
        linhas_lidas = 0
        linhas_duplicadas = 0
        linhas_invalidas = 0
        meses_lidos = set()
        total_baixas_automaticas = 0
        total_despesas_automaticas = 0
        total_sem_membro = 0
        total_codigos_ambiguos = 0
        codigos_sem_membro = {}
        codigos_ambiguos = {}
        diagnostico_dabb = {
            "linhas_detalhe_encontradas": 0,
            "linhas_detalhe_validas": 0,
            "motivos_invalidos": {},
            "exemplos_invalidos": [],
        }

        if ext == ".ofx" or is_ofx_content:
            for tx in _iterar_transacoes_ofx(text):
                linhas_lidas += 1
                try:
                    data = tx["data"]
                    valor = tx["valor"]
                    tipo = tx["tipo"]
                    descricao = tx["descricao"]
                    numero_doc = tx["numero_documento"]
                    mes_ref = data.strftime("%Y-%m")
                    meses_lidos.add(mes_ref)

                    if _descricao_indica_linha_saldo(descricao):
                        linhas_invalidas += 1
                        continue

                    existe = _buscar_duplicado_conciliacao(
                        db=db,
                        data_extrato=data,
                        valor_extrato=valor,
                        banco=banco,
                        tipo=tipo,
                        numero_documento=numero_doc,
                        descricao_extrato=descricao,
                    )

                    if existe:
                        linhas_duplicadas += 1
                        continue

                    c = models.Conciliacao(
                        id=str(uuid.uuid4()),
                        user_id=current_user.id,
                        data_extrato=data,
                        descricao_extrato=descricao,
                        valor_extrato=valor,
                        tipo=tipo,
                        mes_referencia=mes_ref,
                        banco=banco,
                        numero_documento=numero_doc,
                        conciliado=False,
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    db.add(c)
                    db.flush()

                    if tipo == "credito" and _descricao_credito_parece_mensalidade_ofx(descricao):
                        membro = _escolher_membro_para_credito_ofx(db, c)
                        if membro:
                            _baixar_pagamento_mensalidade_por_conciliacao(
                                db=db,
                                conciliacao=c,
                                membro=membro,
                                user_id=current_user.id,
                                observacao_origem=f"Baixa automática via arquivo OFX ({mes_ref})",
                            )
                            total_baixas_automaticas += 1
                        else:
                            total_sem_membro += 1

                    if tipo == "debito":
                        conta_despesa = _inferir_conta_despesa_ofx(db, descricao)
                        if conta_despesa:
                            _lancar_despesa_por_conciliacao(
                                db=db,
                                conciliacao=c,
                                conta=conta_despesa,
                                current_user=current_user,
                                fornecedor=descricao,
                                forma_pagamento="extrato_bancario",
                                observacoes_extra=f"Lançamento automático via arquivo OFX ({mes_ref})",
                            )
                            total_despesas_automaticas += 1

                    importados.append({
                        "data": data.strftime("%Y-%m-%d"),
                        "descricao": descricao,
                        "valor": valor,
                        "tipo": tipo,
                        "numero_doc": numero_doc,
                        "conciliado": c.conciliado,
                    })
                except Exception:
                    linhas_invalidas += 1
                    continue
        elif ext in {".ret", ".rem"}:
            membros_por_codigo_dabb = _indexar_membros_por_codigo_dabb(db)
            data_header_dabb = _extrair_data_header_dabb(text)
            encontrou_linha_detalhe = False

            for numero_linha, raw_line in enumerate(text.splitlines(), start=1):
                line = raw_line.rstrip("\r\n")
                line_normalized = line.lstrip("\ufeff ").rstrip()
                if not line:
                    continue
                if line_normalized[:1] not in {"E", "F"}:
                    continue

                encontrou_linha_detalhe = True
                diagnostico_dabb["linhas_detalhe_encontradas"] += 1

                try:
                    tx = _parse_linha_dabb(line_normalized, data_fallback=data_header_dabb)
                except ValueError as exc:
                    motivo = str(exc)
                    linhas_invalidas += 1
                    diagnostico_dabb["motivos_invalidos"][motivo] = diagnostico_dabb["motivos_invalidos"].get(motivo, 0) + 1
                    if len(diagnostico_dabb["exemplos_invalidos"]) < 5:
                        diagnostico_dabb["exemplos_invalidos"].append({
                            "linha": numero_linha,
                            "motivo": motivo,
                            "conteudo": line_normalized[:120],
                        })
                    continue

                diagnostico_dabb["linhas_detalhe_validas"] += 1
                linhas_lidas += 1
                try:
                    data = tx["data"]
                    valor = tx["valor"]
                    tipo = tx["tipo"]
                    descricao = tx["descricao"]
                    numero_doc = tx["numero_documento"]
                    codigo_dabb = tx["codigo_dabb"]
                    codigo_barras = tx.get("codigo_barras")
                    linha_original = tx["linha_original"]
                    mes_ref = data.strftime("%Y-%m")
                    meses_lidos.add(mes_ref)

                    existe = _buscar_duplicado_conciliacao(
                        db=db,
                        data_extrato=data,
                        valor_extrato=valor,
                        banco=banco,
                        tipo=tipo,
                        numero_documento=numero_doc,
                        descricao_extrato=descricao,
                    )

                    if existe:
                        linhas_duplicadas += 1
                        continue

                    observacoes_partes = [
                        f"Arquivo DABB {ext.upper()}",
                        f"codigo_dabb={codigo_dabb}",
                    ]
                    if codigo_barras:
                        observacoes_partes.append(f"codigo_barras={codigo_barras}")
                    observacoes_partes.append(f"linha={linha_original}")

                    c = models.Conciliacao(
                        id=str(uuid.uuid4()),
                        user_id=current_user.id,
                        data_extrato=data,
                        descricao_extrato=descricao,
                        valor_extrato=valor,
                        tipo=tipo,
                        mes_referencia=mes_ref,
                        banco=banco,
                        numero_documento=numero_doc,
                        conciliado=False,
                        observacoes=" | ".join(observacoes_partes),
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    db.add(c)
                    db.flush()

                    membros_match = {}
                    for variante in _variantes_codigo_dabb(codigo_dabb):
                        for membro in membros_por_codigo_dabb.get(variante, []):
                            membros_match[membro.id] = membro

                    if len(membros_match) == 1:
                        membro = next(iter(membros_match.values()))
                        _baixar_pagamento_mensalidade_por_conciliacao(
                            db=db,
                            conciliacao=c,
                            membro=membro,
                            user_id=current_user.id,
                            observacao_origem=(
                                f"Baixa automática via arquivo bancário DABB ({mes_ref}) - "
                                f"codigo_dabb {codigo_dabb}"
                            ),
                        )
                        total_baixas_automaticas += 1
                    elif len(membros_match) > 1:
                        total_codigos_ambiguos += 1
                        item = codigos_ambiguos.setdefault(codigo_dabb, {
                            "codigo_dabb": codigo_dabb,
                            "quantidade": 0,
                            "valores": set(),
                            "meses": set(),
                            "registros": set(),
                        })
                        item["quantidade"] += 1
                        item["valores"].add(round(float(valor or 0), 2))
                        item["meses"].add(mes_ref)
                        item["registros"].add(tx.get("tipo_registro") or "?")
                        c.observacoes = (
                            (c.observacoes + "\n") if c.observacoes else ""
                        ) + "Codigo DABB encontrado em mais de um membro ativo; baixa nao realizada automaticamente."
                    else:
                        total_sem_membro += 1
                        item = codigos_sem_membro.setdefault(codigo_dabb, {
                            "codigo_dabb": codigo_dabb,
                            "quantidade": 0,
                            "valores": set(),
                            "meses": set(),
                            "registros": set(),
                        })
                        item["quantidade"] += 1
                        item["valores"].add(round(float(valor or 0), 2))
                        item["meses"].add(mes_ref)
                        item["registros"].add(tx.get("tipo_registro") or "?")
                        c.observacoes = (
                            (c.observacoes + "\n") if c.observacoes else ""
                        ) + "Nenhum membro ativo encontrado para o codigo DABB; baixa nao realizada automaticamente."

                    importados.append({
                        "data": data.strftime("%Y-%m-%d"),
                        "descricao": descricao,
                        "valor": valor,
                        "tipo": tipo,
                        "numero_doc": numero_doc,
                        "codigo_dabb": codigo_dabb,
                        "codigo_barras": codigo_barras,
                        "conciliado": c.conciliado,
                    })
                except Exception:
                    linhas_invalidas += 1
                    diagnostico_dabb["motivos_invalidos"]["erro_processamento"] = diagnostico_dabb["motivos_invalidos"].get("erro_processamento", 0) + 1
                    continue

            if not encontrou_linha_detalhe:
                return {
                    "ok": False,
                    "total_importados": 0,
                    "linhas_lidas": 0,
                    "linhas_duplicadas": 0,
                    "linhas_invalidas": 0,
                    "total_baixas_automaticas": 0,
                    "total_despesas_automaticas": 0,
                    "total_sem_membro": 0,
                    "total_codigos_ambiguos": 0,
                    "diagnostico_dabb": {
                        **diagnostico_dabb,
                        "motivos_invalidos": {"nenhuma_linha_ef_encontrada": 1},
                        "exemplos_invalidos": [],
                    },
                    "meses_importados": [],
                    "meses_lidos": [],
                    "registros": []
                }
        else:
            all_lines = [ln for ln in text.splitlines() if ln.strip()]
            if not all_lines:
                raise HTTPException(status_code=400, detail="Arquivo CSV vazio")

            start_idx = _encontrar_inicio_csv(all_lines)
            lines = all_lines[start_idx:]
            if not lines:
                raise HTTPException(status_code=400, detail="Cabecalho CSV nao encontrado")

            delimiter = _escolher_delimitador_csv(lines[0])
            reader = csv.DictReader(lines, delimiter=delimiter)

            for row in reader:
                linhas_lidas += 1
                try:
                    data_str = _valor_por_alias(row, ["data", "data lancamento", "data movimento", "dt lancamento"])
                    valor_str = _valor_por_alias(row, ["valor", "valor r$", "valor (r$)", "valor lancamento", "vlr"])
                    descricao = _valor_por_alias(
                        row,
                        ["descricao", "descricao lancamento", "historico", "detalhes", "lancamento", "memo", "name"]
                    )
                    tipo_raw = _valor_por_alias(row, ["tipo", "tipo lancamento", "natureza"])
                    numero_doc = _valor_por_alias(row, ["n documento", "no documento", "numero documento", "documento", "fitid"])

                    if not data_str or not valor_str:
                        linhas_invalidas += 1
                        continue

                    data = _parse_data_extrato(data_str)
                    valor = _parse_valor_extrato(valor_str)

                    tipo_norm = _normalizar_header_csv(tipo_raw)
                    if tipo_norm in ["entrada", "credito", "credit", "c"]:
                        tipo = "credito"
                    elif tipo_norm in ["saida", "debito", "debit", "d"]:
                        tipo = "debito"
                    else:
                        tipo = "credito" if valor >= 0 else "debito"

                    descricao_final = descricao or tipo_raw or "Lancamento bancario"
                    valor_abs = abs(valor)
                    mes_ref = data.strftime("%Y-%m")
                    meses_lidos.add(mes_ref)

                    if _descricao_indica_linha_saldo(descricao_final):
                        linhas_invalidas += 1
                        continue

                    existe = _buscar_duplicado_conciliacao(
                        db=db,
                        data_extrato=data,
                        valor_extrato=valor_abs,
                        banco=banco,
                        tipo=tipo,
                        numero_documento=numero_doc,
                        descricao_extrato=descricao_final,
                    )

                    if existe:
                        linhas_duplicadas += 1
                        continue

                    c = models.Conciliacao(
                        id=str(uuid.uuid4()),
                        user_id=current_user.id,
                        data_extrato=data,
                        descricao_extrato=descricao_final,
                        valor_extrato=valor_abs,
                        tipo=tipo,
                        mes_referencia=mes_ref,
                        banco=banco,
                        numero_documento=numero_doc if numero_doc else None,
                        conciliado=False,
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    db.add(c)
                    importados.append({
                        "data": data.strftime("%Y-%m-%d"),
                        "descricao": descricao_final,
                        "valor": valor_abs,
                        "tipo": tipo,
                        "numero_doc": numero_doc
                    })

                except Exception:
                    linhas_invalidas += 1
                    continue

        db.commit()
        meses_importados = sorted({item.get("data", "")[:7] for item in importados if item.get("data")})
        codigos_sem_membro_lista = [
            {
                "codigo_dabb": item["codigo_dabb"],
                "quantidade": item["quantidade"],
                "valores": sorted(item["valores"]),
                "meses": sorted(item["meses"]),
                "registros": sorted(item["registros"]),
            }
            for item in sorted(codigos_sem_membro.values(), key=lambda x: (x["codigo_dabb"]))
        ]
        codigos_ambiguos_lista = [
            {
                "codigo_dabb": item["codigo_dabb"],
                "quantidade": item["quantidade"],
                "valores": sorted(item["valores"]),
                "meses": sorted(item["meses"]),
                "registros": sorted(item["registros"]),
            }
            for item in sorted(codigos_ambiguos.values(), key=lambda x: (x["codigo_dabb"]))
        ]
        return {
            "ok": True,
            "total_importados": len(importados),
            "linhas_lidas": linhas_lidas,
            "linhas_duplicadas": linhas_duplicadas,
            "linhas_invalidas": linhas_invalidas,
            "total_baixas_automaticas": total_baixas_automaticas,
            "total_despesas_automaticas": total_despesas_automaticas,
            "total_sem_membro": total_sem_membro,
            "total_codigos_ambiguos": total_codigos_ambiguos,
            "codigos_sem_membro": codigos_sem_membro_lista,
            "codigos_ambiguos": codigos_ambiguos_lista,
            "diagnostico_dabb": diagnostico_dabb if ext in {".ret", ".rem"} else None,
            "meses_importados": meses_importados,
            "meses_lidos": sorted(meses_lidos),
            "registros": importados
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar: {str(e)}")


# ════════════════════════════════════════════════════════════════════════════════
# FINANCEIRO / BALANCETE
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/financeiro/balancete")
def balancete(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")

    pags = _pagamentos_pagos_membros_ativos_no_mes(db, mes_ref)
    desp = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_ref).all()
    rendas = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_ref).all()

    total_mensalidades = sum(float(p.valor_pago) for p in pags if p.valor_pago)
    total_outras = sum(r.valor for r in rendas if r.valor)
    total_entradas = total_mensalidades + total_outras
    total_saidas = sum(d.valor for d in desp if d.valor)
    saldo = total_entradas - total_saidas
    saldo_manual_registrado = db.query(models.SaldoMensal).filter(models.SaldoMensal.mes_referencia == mes_ref).first()
    saldo_anterior = _saldo_anterior_mes(db, mes_ref)
    saldo_final = round(saldo_anterior + saldo, 2)

    # Por categoria de despesa
    desp_por_cat = defaultdict(float)
    for d in desp:
        if d.valor:
            desp_por_cat[d.categoria or "Outros"] += d.valor

    # Por categoria de renda
    renda_por_cat = defaultdict(float)
    for r in rendas:
        if r.valor:
            renda_por_cat[r.categoria or "Outros"] += r.valor

    entradas_por_conta = defaultdict(float)
    entradas_meta = {}
    entradas_por_conta["1.1"] += total_mensalidades
    entradas_meta["1.1"] = {"codigo": "1.1", "nome": "Mensalidades"}

    for r in rendas:
        if not r.valor:
            continue
        codigo = (r.conta_codigo or "1.5").strip()
        nome = (r.conta_nome or r.categoria or "Outras arrecadações").strip()
        entradas_por_conta[codigo] += float(r.valor)
        entradas_meta[codigo] = {"codigo": codigo, "nome": nome}

    saidas_por_conta = defaultdict(float)
    saidas_meta = {}
    for d in desp:
        if not d.valor:
            continue
        codigo = (d.conta_codigo or "2.99").strip()
        nome = (d.conta_nome or d.categoria or "Outros").strip()
        saidas_por_conta[codigo] += float(d.valor)
        saidas_meta[codigo] = {"codigo": codigo, "nome": nome}

    contas_entrada = db.query(models.PlanoConta).filter(models.PlanoConta.tipo == "entrada").all()
    contas_saida = db.query(models.PlanoConta).filter(models.PlanoConta.tipo == "saida").all()

    for c in contas_entrada:
        codigo = (c.codigo or "").strip()
        if not codigo:
            continue
        entradas_meta[codigo] = {"codigo": codigo, "nome": c.nome or ""}

    for c in contas_saida:
        codigo = (c.codigo or "").strip()
        if not codigo:
            continue
        saidas_meta[codigo] = {"codigo": codigo, "nome": c.nome or ""}

    def _ordena_codigo_conta(item):
        codigo = item[0]
        partes = []
        for p in codigo.split('.'):
            try:
                partes.append(int(p))
            except Exception:
                partes.append(9999)
        return tuple(partes)

    entradas_ordenadas = [
        {
            "codigo": c,
            "nome": entradas_meta.get(c, {}).get("nome") or "",
            "valor": round(v, 2),
        }
        for c, v in sorted(entradas_por_conta.items(), key=_ordena_codigo_conta)
        if abs(float(v or 0)) > 0.000001
    ]

    saidas_ordenadas = [
        {
            "codigo": c,
            "nome": saidas_meta.get(c, {}).get("nome") or "",
            "valor": round(v, 2),
        }
        for c, v in sorted(saidas_por_conta.items(), key=_ordena_codigo_conta)
        if abs(float(v or 0)) > 0.000001
    ]

    return {
        "mes_referencia": mes_ref,
        "mes_referencia_anterior": _mes_anterior(mes_ref),
        "saldo_anterior": saldo_anterior,
        "origem_saldo_anterior": "manual" if saldo_manual_registrado else "calculado",
        "total_mensalidades": total_mensalidades,
        "total_outras_rendas": total_outras,
        "total_entradas": total_entradas,
        "total_despesas": total_saidas,
        "saldo": saldo,
        "saldo_final": saldo_final,
        "qtd_pagantes": len(pags),
        "despesas_por_categoria": dict(desp_por_cat),
        "rendas_por_categoria": dict(renda_por_cat),
        "entradas_por_conta": entradas_ordenadas,
        "saidas_por_conta": saidas_ordenadas,
        "pagamentos": [{
            "membro_id": p.membro_id, "valor_pago": float(p.valor_pago) if p.valor_pago else 0,
            "data_pagamento": str(p.data_pagamento) if p.data_pagamento else None,
            "forma_pagamento": p.forma_pagamento
        } for p in pags],
        "despesas": [{
            "id": d.id, "descricao": d.descricao, "categoria": d.categoria,
            "valor": d.valor, "data_despesa": str(d.data_despesa) if d.data_despesa else None,
            "conta_id": d.conta_id, "conta_codigo": d.conta_codigo, "conta_nome": d.conta_nome
        } for d in desp],
        "rendas": [{
            "id": r.id, "descricao": r.descricao, "categoria": r.categoria,
            "valor": r.valor, "data_recebimento": str(r.data_recebimento) if r.data_recebimento else None,
            "conta_id": r.conta_id, "conta_codigo": r.conta_codigo, "conta_nome": r.conta_nome
        } for r in rendas],
    }


# ════════════════════════════════════════════════════════════════════════════════
# RELATÓRIOS / EXPORTAÇÃO EXCEL
# ════════════════════════════════════════════════════════════════════════════════
def _excel_header_style(cell):
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _excel_thin_border() -> Border:
    return Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )


def _excel_apply_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    border = _excel_thin_border()
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border


def _excel_apply_zebra(ws, start_row: int, end_row: int, start_col: int, end_col: int, fill_color: str = "F8FAFC"):
    if end_row < start_row:
        return
    zebra_fill = PatternFill("solid", fgColor=fill_color)
    for index, row in enumerate(range(start_row, end_row + 1)):
        if index % 2 == 0:
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).fill = zebra_fill


def _excel_autofit_columns(ws, min_width: int = 10, max_width: int = 48):
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        max_length = 0
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max(min_width, min(max_length + 2, max_width))


def _excel_set_date_cell(ws, row: int, column: int, value, alignment: str = "center"):
    cell = ws.cell(row=row, column=column, value=value if value else None)
    if value:
        cell.number_format = 'DD/MM/YYYY'
    cell.alignment = Alignment(horizontal=alignment)
    return cell


@app.get("/api/relatorios/membros")
def exportar_membros(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Membro)
    if status:
        q = q.filter(models.Membro.status == status)
    membros = q.order_by(models.Membro.nome_completo).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Membros"

    ws.merge_cells("A1:Q1")
    ws["A1"] = "RELATÓRIO DE MEMBROS"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Matrícula", "Nome", "CPF", "Código DABB", "Email", "Telefone", "Celular",
               "Endereço", "Bairro", "Cidade", "Estado", "CEP",
               "Data Nascimento", "Data Associação", "Status", "Benefício", "Valor Mensalidade"]

    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        _excel_header_style(cell)

    first_data_row = header_row + 1
    for row, m in enumerate(membros, first_data_row):
        ws.cell(row=row, column=1, value=m.matricula)
        ws.cell(row=row, column=2, value=m.nome_completo)
        ws.cell(row=row, column=3, value=m.cpf)
        ws.cell(row=row, column=4, value=m.codigo_dabb)
        ws.cell(row=row, column=5, value=m.email)
        ws.cell(row=row, column=6, value=m.telefone)
        ws.cell(row=row, column=7, value=m.celular)
        ws.cell(row=row, column=8, value=m.endereco)
        ws.cell(row=row, column=9, value=m.bairro)
        ws.cell(row=row, column=10, value=m.cidade)
        ws.cell(row=row, column=11, value=m.estado)
        ws.cell(row=row, column=12, value=m.cep)
        _excel_set_date_cell(ws, row, 13, m.data_nascimento)
        _excel_set_date_cell(ws, row, 14, m.data_filiacao)
        ws.cell(row=row, column=15, value=m.status)
        ws.cell(row=row, column=16, value=m.beneficio)
        ws.cell(row=row, column=17, value=float(m.valor_mensalidade) if m.valor_mensalidade else 0)
        ws.cell(row=row, column=17).number_format = 'R$ #,##0.00'

    last_row = max(first_data_row, first_data_row + len(membros) - 1)
    _excel_apply_zebra(ws, first_data_row, last_row, 1, 17)
    _excel_apply_borders(ws, header_row, last_row, 1, 17)
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:Q{last_row}"

    for row in range(first_data_row, last_row + 1):
        ws.cell(row=row, column=17).alignment = Alignment(horizontal="right")

    _excel_autofit_columns(ws, min_width=12, max_width=44)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=membros.xlsx"}
    )

@app.get("/api/relatorios/pagamentos")
def exportar_pagamentos(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")

    membros = db.query(models.Membro).filter(models.Membro.status == 'ativo').order_by(models.Membro.nome_completo).all()
    pagamentos = {p.membro_id: p for p in db.query(models.Pagamento).filter(models.Pagamento.mes_referencia == mes_ref).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mensalidades {mes_ref}"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"RECEBIMENTO DE MENSALIDADES - {mes_ref}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Matrícula", "Nome", "Valor Mensalidade", "Valor Pago", "Data Pagamento", "Status", "Forma Pagamento"]

    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        _excel_header_style(cell)

    red_fill = PatternFill("solid", fgColor="FFB3B3")
    green_fill = PatternFill("solid", fgColor="B3FFB3")

    first_data_row = header_row + 1
    total_mens = 0.0
    total_pago = 0.0
    status_by_row = {}

    for row, m in enumerate(membros, first_data_row):
        p = pagamentos.get(m.id)
        status = p.status_pagamento if p else "pendente"
        fill = green_fill if status == "pago" else red_fill
        
        values = [
            m.matricula, m.nome_completo,
            float(m.valor_mensalidade) if m.valor_mensalidade else 0,
            float(p.valor_pago) if p and p.valor_pago else 0,
            p.data_pagamento if p and p.data_pagamento else None,
            status,
            p.forma_pagamento if p else ""
        ]
        for col, val in enumerate(values, 1):
            if col == 5:
                cell = _excel_set_date_cell(ws, row, col, val)
            else:
                cell = ws.cell(row=row, column=col, value=val)
            if col in (3, 4):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col == 6:
                cell.alignment = Alignment(horizontal="center")

        status_by_row[row] = fill

        total_mens += float(values[2] or 0)
        total_pago += float(values[3] or 0)

    total_row = first_data_row + len(membros)
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=round(total_mens, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=round(total_pago, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=3).number_format = 'R$ #,##0.00'
    ws.cell(row=total_row, column=4).number_format = 'R$ #,##0.00'
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor="E2E8F0")
    ws.cell(row=total_row, column=3).fill = PatternFill("solid", fgColor="E2E8F0")
    ws.cell(row=total_row, column=4).fill = PatternFill("solid", fgColor="E2E8F0")

    _excel_apply_zebra(ws, first_data_row, total_row - 1, 1, 7)
    for row, status_fill in status_by_row.items():
        ws.cell(row=row, column=6).fill = status_fill

    _excel_apply_borders(ws, header_row, total_row, 1, 7)
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:G{total_row}"
    _excel_autofit_columns(ws, min_width=12, max_width=42)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=recebimento_mensalidades_{mes_ref}.xlsx"}
    )

@app.get("/api/relatorios/aniversariantes")
def exportar_aniversariantes(
    mes: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_filtro = mes or today.month
    membros = db.query(models.Membro).filter(
        models.Membro.data_nascimento.isnot(None),
        sql_func.strftime('%m', models.Membro.data_nascimento) == str(mes_filtro).zfill(2)
    ).order_by(sql_func.strftime('%d', models.Membro.data_nascimento)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Aniversariantes Mês {mes_filtro}"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"RELATÓRIO DE ANIVERSARIANTES - {mes_filtro:02d}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Nome", "Data Nascimento", "Dia", "Idade", "Email", "Celular", "Telefone"]

    header_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        _excel_header_style(cell)

    first_data_row = header_row + 1
    for row, m in enumerate(membros, first_data_row):
        idade = today.year - m.data_nascimento.year
        ws.cell(row=row, column=1, value=m.nome_completo)
        _excel_set_date_cell(ws, row, 2, m.data_nascimento)
        ws.cell(row=row, column=3, value=m.data_nascimento.day)
        ws.cell(row=row, column=4, value=idade)
        ws.cell(row=row, column=5, value=m.email)
        ws.cell(row=row, column=6, value=m.celular)
        ws.cell(row=row, column=7, value=m.telefone)

    last_row = max(first_data_row, first_data_row + len(membros) - 1)
    _excel_apply_zebra(ws, first_data_row, last_row, 1, 7)
    _excel_apply_borders(ws, header_row, last_row, 1, 7)
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:G{last_row}"
    _excel_autofit_columns(ws, min_width=12, max_width=40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=aniversariantes_mes_{mes_filtro}.xlsx"}
    )

@app.get("/api/relatorios/balancete")
def exportar_balancete(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")
    pags = _pagamentos_pagos_membros_ativos_no_mes(db, mes_ref)
    desp = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_ref).all()
    rendas = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_ref).all()
    saldo_anterior = _saldo_anterior_mes(db, mes_ref)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Balancete"

    title_fill = PatternFill("solid", fgColor="1E3A5F")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="FCE4D6")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    bold_font = Font(bold=True)

    total_mensalidades = sum(float(p.valor_pago) for p in pags if p.valor_pago)
    entradas_por_conta = defaultdict(float)
    entradas_meta = {"1.1": "Mensalidades"}
    entradas_por_conta["1.1"] = total_mensalidades

    for r in rendas:
        if not r.valor:
            continue
        codigo = (r.conta_codigo or "1.5").strip()
        nome = (r.conta_nome or r.categoria or "Outras arrecadações").strip()
        entradas_meta[codigo] = nome
        entradas_por_conta[codigo] += float(r.valor)

    saidas_por_conta = defaultdict(float)
    saidas_meta = {}
    for d in desp:
        if not d.valor:
            continue
        codigo = (d.conta_codigo or "2.99").strip()
        nome = (d.conta_nome or d.categoria or "Outros").strip()
        saidas_meta[codigo] = nome
        saidas_por_conta[codigo] += float(d.valor)

    contas_entrada = db.query(models.PlanoConta).filter(models.PlanoConta.tipo == "entrada").all()
    contas_saida = db.query(models.PlanoConta).filter(models.PlanoConta.tipo == "saida").all()
    for c in contas_entrada:
        codigo = (c.codigo or "").strip()
        if not codigo:
            continue
        entradas_meta[codigo] = c.nome or ""
    for c in contas_saida:
        codigo = (c.codigo or "").strip()
        if not codigo:
            continue
        saidas_meta[codigo] = c.nome or ""

    def _sort_codigo(codigo: str):
        out = []
        for p in codigo.split('.'):
            try:
                out.append(int(p))
            except Exception:
                out.append(9999)
        return tuple(out)

    row = 1
    ws1.merge_cells("A1:D1")
    ws1[f"A{row}"] = f"BALANCETE MENSAL - {mes_ref}"
    ws1[f"A{row}"].fill = title_fill
    ws1[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws1[f"A{row}"].font = title_font
    row += 2

    ws1[f"B{row}"] = "SALDO ANTERIOR"
    ws1[f"C{row}"] = "R$"
    ws1[f"D{row}"] = round(float(saldo_anterior or 0), 2)
    ws1[f"B{row}"].font = bold_font
    ws1[f"C{row}"].font = bold_font
    ws1[f"D{row}"].font = bold_font
    ws1[f"D{row}"].alignment = Alignment(horizontal="right")
    row += 2

    ws1[f"A{row}"] = "Contas"
    ws1[f"B{row}"] = "ENTRADAS"
    ws1[f"C{row}"] = "Total"
    total_entradas = round(sum(entradas_por_conta.values()), 2)
    ws1[f"D{row}"] = total_entradas
    for col in ("A", "B", "C", "D"):
        ws1[f"{col}{row}"].fill = header_fill
        ws1[f"{col}{row}"].font = bold_font
    row += 1

    entrada_start_row = row
    for codigo in sorted(entradas_por_conta.keys(), key=_sort_codigo):
        if abs(float(entradas_por_conta[codigo] or 0)) <= 0.000001:
            continue
        ws1[f"A{row}"] = codigo
        ws1[f"B{row}"] = entradas_meta.get(codigo, "")
        ws1[f"C{row}"] = "R$"
        ws1[f"D{row}"] = round(float(entradas_por_conta[codigo] or 0), 2)
        ws1[f"D{row}"].alignment = Alignment(horizontal="right")
        row += 1
    entrada_end_row = row - 1
    _excel_apply_zebra(ws1, entrada_start_row, entrada_end_row, 1, 4)

    row += 1
    ws1[f"A{row}"] = "Contas"
    ws1[f"B{row}"] = "SAÍDAS"
    ws1[f"C{row}"] = "Total"
    total_saidas = round(sum(saidas_por_conta.values()), 2)
    ws1[f"D{row}"] = total_saidas
    for col in ("A", "B", "C", "D"):
        ws1[f"{col}{row}"].fill = header_fill
        ws1[f"{col}{row}"].font = bold_font
    row += 1

    saida_start_row = row
    for codigo in sorted(saidas_por_conta.keys(), key=_sort_codigo):
        if abs(float(saidas_por_conta[codigo] or 0)) <= 0.000001:
            continue
        ws1[f"A{row}"] = codigo
        ws1[f"B{row}"] = saidas_meta.get(codigo, "")
        ws1[f"C{row}"] = "R$"
        ws1[f"D{row}"] = round(float(saidas_por_conta[codigo] or 0), 2)
        ws1[f"D{row}"].alignment = Alignment(horizontal="right")
        row += 1
    saida_end_row = row - 1
    _excel_apply_zebra(ws1, saida_start_row, saida_end_row, 1, 4)

    row += 1
    saldo_mes = round(total_entradas - total_saidas, 2)
    saldo_final = round(float(saldo_anterior or 0) + saldo_mes, 2)

    ws1[f"B{row}"] = "SALDO DO MÊS"
    ws1[f"C{row}"] = "R$"
    ws1[f"D{row}"] = saldo_mes
    for col in ("B", "C", "D"):
        ws1[f"{col}{row}"].fill = total_fill
        ws1[f"{col}{row}"].font = bold_font
    ws1[f"D{row}"].alignment = Alignment(horizontal="right")
    row += 1

    ws1[f"B{row}"] = "SALDO FINAL"
    ws1[f"C{row}"] = "R$"
    ws1[f"D{row}"] = saldo_final
    for col in ("B", "C", "D"):
        ws1[f"{col}{row}"].fill = total_fill
        ws1[f"{col}{row}"].font = bold_font
    ws1[f"D{row}"].alignment = Alignment(horizontal="right")

    for col_letter, width in {"A": 12, "B": 56, "C": 8, "D": 18}.items():
        ws1.column_dimensions[col_letter].width = width

    for r_idx in range(1, row + 1):
        ws1[f"D{r_idx}"].number_format = 'R$ #,##0.00'

    _excel_apply_borders(ws1, 3, row, 1, 4)
    ws1.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=balancete_{mes_ref}.xlsx"}
    )


@app.get("/api/relatorios/livro-diario")
def exportar_livro_diario(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")

    pags = _pagamentos_pagos_membros_ativos_no_mes(db, mes_ref)
    rendas = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_ref).all()
    despesas = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_ref).all()
    saldo_anterior = _saldo_anterior_mes(db, mes_ref)

    nomes_membros = {}
    membro_ids = [p.membro_id for p in pags if p.membro_id]
    if membro_ids:
        nomes_membros = {
            m.id: m.nome_completo
            for m in db.query(models.Membro).filter(models.Membro.id.in_(membro_ids)).all()
        }

    linhas = []

    for p in pags:
        valor = float(p.valor_pago or 0)
        if valor <= 0:
            continue
        nome_membro = nomes_membros.get(p.membro_id, p.membro_id or "Membro")
        linhas.append({
            "data": p.data_pagamento,
            "codigo": "1.1",
            "conta": "Mensalidades",
            "historico": f"Mensalidade de {nome_membro}",
            "tipo": "entrada",
            "valor": valor,
        })

    for r in rendas:
        valor = float(r.valor or 0)
        if valor <= 0:
            continue
        linhas.append({
            "data": r.data_recebimento,
            "codigo": (r.conta_codigo or "1.5"),
            "conta": (r.conta_nome or r.categoria or "Outras arrecadações"),
            "historico": r.descricao or "Outras receitas",
            "tipo": "entrada",
            "valor": valor,
        })

    for d in despesas:
        valor = float(d.valor or 0)
        if valor <= 0:
            continue
        linhas.append({
            "data": d.data_despesa,
            "codigo": (d.conta_codigo or "2.99"),
            "conta": (d.conta_nome or d.categoria or "Outros"),
            "historico": d.descricao or "Despesa",
            "tipo": "saida",
            "valor": valor,
        })

    linhas.sort(key=lambda x: (x["data"] or date.min, x["codigo"] or "", x["historico"] or ""))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livro Diário"

    title_fill = PatternFill("solid", fgColor="1E3A5F")
    total_fill = PatternFill("solid", fgColor="FCE4D6")
    bold_font = Font(bold=True)
    saida_font = Font(color="C00000")
    saida_bold_font = Font(bold=True, color="C00000")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"LIVRO DIÁRIO - {mes_ref}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"] = "Saldo Anterior"
    ws["B3"] = round(float(saldo_anterior or 0), 2)
    ws["A3"].font = bold_font
    ws["B3"].font = bold_font
    ws["B3"].number_format = 'R$ #,##0.00'

    headers = ["Data", "Conta", "Descrição da Conta", "Histórico", "Tipo", "Entradas (R$)", "Saídas (R$)", "Saldo Acumulado (R$)"]
    start_row = 5
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=idx, value=h)
        _excel_header_style(cell)

    saldo_acumulado = float(saldo_anterior or 0)
    row = start_row + 1
    total_entradas = 0.0
    total_saidas = 0.0

    for item in linhas:
        entrada = float(item["valor"] or 0) if item["tipo"] == "entrada" else 0.0
        saida = float(item["valor"] or 0) if item["tipo"] == "saida" else 0.0
        saldo_acumulado += entrada - saida
        total_entradas += entrada
        total_saidas += saida

        ws.cell(row=row, column=1, value=item["data"])
        ws.cell(row=row, column=2, value=item["codigo"])
        ws.cell(row=row, column=3, value=item["conta"])
        ws.cell(row=row, column=4, value=item["historico"])
        ws.cell(row=row, column=5, value="Entrada" if item["tipo"] == "entrada" else "Saída")
        ws.cell(row=row, column=6, value=round(entrada, 2))
        ws.cell(row=row, column=7, value=round(saida, 2))
        ws.cell(row=row, column=8, value=round(saldo_acumulado, 2))

        if item["tipo"] == "saida":
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = saida_font

        row += 1

    movimentos_end_row = row - 1
    _excel_apply_zebra(ws, start_row + 1, movimentos_end_row, 1, 8)

    ws.cell(row=row, column=4, value="TOTAL DO MÊS").font = bold_font
    ws.cell(row=row, column=6, value=round(total_entradas, 2)).font = bold_font
    ws.cell(row=row, column=7, value=round(total_saidas, 2)).font = saida_bold_font
    ws.cell(row=row, column=8, value=round(saldo_acumulado, 2)).font = bold_font
    for col in range(4, 9):
        ws.cell(row=row, column=col).fill = total_fill

    for col_letter, width in {"A": 14, "B": 10, "C": 34, "D": 44, "E": 12, "F": 16, "G": 16, "H": 20}.items():
        ws.column_dimensions[col_letter].width = width

    for r in range(start_row + 1, row + 1):
        ws.cell(row=r, column=1).number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=7).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=8).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="right")

    _excel_apply_borders(ws, start_row, row, 1, 8)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:H{row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=livro_diario_{mes_ref}.xlsx"}
    )


@app.get("/api/relatorios/conciliacao")
def exportar_conciliacao(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")
    saldo_anterior = _saldo_anterior_mes(db, mes_ref)

    conciliacoes = db.query(models.Conciliacao).filter(
        models.Conciliacao.mes_referencia == mes_ref
    ).order_by(models.Conciliacao.data_extrato.desc()).all()

    total_creditos = round(sum(float(c.valor_extrato or 0) for c in conciliacoes if c.tipo == "credito"), 2)
    total_debitos = round(sum(float(c.valor_extrato or 0) for c in conciliacoes if c.tipo == "debito"), 2)
    saldo_mes = round(total_creditos - total_debitos, 2)
    saldo_final = round(saldo_anterior + saldo_mes, 2)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Conciliação {mes_ref}"

    ws["A1"] = f"CONCILIAÇÃO BANCÁRIA - {mes_ref}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws.merge_cells("A1:J1")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "RESUMO"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Saldo Anterior"
    ws["B4"] = saldo_anterior
    ws["A5"] = "Total Créditos"
    ws["B5"] = total_creditos
    ws["A6"] = "Total Débitos"
    ws["B6"] = total_debitos
    ws["A7"] = "Saldo do Extrato"
    ws["A7"].font = Font(bold=True)
    ws["B7"] = saldo_mes
    ws["B7"].font = Font(bold=True)
    ws["A8"] = "Saldo Final"
    ws["A8"].font = Font(bold=True)
    ws["B8"] = saldo_final
    ws["B8"].font = Font(bold=True)

    for row in range(4, 9):
        ws.cell(row=row, column=2).number_format = 'R$ #,##0.00'
    _excel_apply_borders(ws, 3, 8, 1, 2)

    headers = [
        "Data", "Descrição", "Valor", "Tipo", "Conciliado",
        "Mês Referência", "Banco", "Documento", "Pagamento ID", "Observações"
    ]

    header_row = 10
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        _excel_header_style(cell)

    first_data_row = header_row + 1
    for row, c in enumerate(conciliacoes, first_data_row):
        _excel_set_date_cell(ws, row, 1, c.data_extrato)
        ws.cell(row=row, column=2, value=c.descricao_extrato)
        ws.cell(row=row, column=3, value=float(c.valor_extrato) if c.valor_extrato else 0)
        ws.cell(row=row, column=4, value=c.tipo)
        ws.cell(row=row, column=5, value="Sim" if c.conciliado else "Não")
        ws.cell(row=row, column=6, value=c.mes_referencia)
        ws.cell(row=row, column=7, value=c.banco)
        ws.cell(row=row, column=8, value=c.numero_documento)
        ws.cell(row=row, column=9, value=c.pagamento_id)
        ws.cell(row=row, column=10, value=c.observacoes)

        ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

    last_row = max(first_data_row, first_data_row + len(conciliacoes) - 1)
    _excel_apply_zebra(ws, first_data_row, last_row, 1, 10)
    _excel_apply_borders(ws, header_row, last_row, 1, 10)
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:J{last_row}"
    _excel_autofit_columns(ws, min_width=12, max_width=46)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=conciliacao_{mes_ref}.xlsx"}
    )


@app.get("/api/relatorios/aplicacoes-financeiras")
def exportar_aplicacoes_financeiras(
    mes_referencia: Optional[str] = None,
    instituicao: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")

    q = db.query(models.AplicacaoFinanceira).filter(
        models.AplicacaoFinanceira.mes_referencia == mes_ref
    )
    if instituicao:
        q = q.filter(models.AplicacaoFinanceira.instituicao.ilike(f"%{instituicao}%"))
    registros = q.order_by(models.AplicacaoFinanceira.instituicao.asc(), models.AplicacaoFinanceira.produto.asc()).all()

    totais = {
        "saldo_anterior": round(sum(float(r.saldo_anterior or 0) for r in registros), 2),
        "aplicacoes": round(sum(float(r.aplicacoes or 0) for r in registros), 2),
        "rendimento_bruto": round(sum(float(r.rendimento_bruto or 0) for r in registros), 2),
        "imposto_renda": round(sum(float(r.imposto_renda or 0) for r in registros), 2),
        "iof": round(sum(float(r.iof or 0) for r in registros), 2),
        "impostos": round(sum(float(r.impostos or 0) for r in registros), 2),
        "rendimento_liquido": round(sum(float(r.rendimento_liquido or 0) for r in registros), 2),
        "resgate": round(sum(float(r.resgate or 0) for r in registros), 2),
        "saldo_atual": round(sum(float(r.saldo_atual or 0) for r in registros), 2),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Aplicações {mes_ref}"

    title_fill = PatternFill("solid", fgColor="1e3a5f")
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    total_fill = PatternFill("solid", fgColor="E6F4EA")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    money_fmt = 'R$ #,##0.00'

    ws.merge_cells("A1:K1")
    ws["A1"] = f"RELATÓRIO DE APLICAÇÕES FINANCEIRAS - {mes_ref}"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "Mês Referência"
    ws["B3"] = mes_ref
    ws["A4"] = "Total de Registros"
    ws["B4"] = len(registros)

    ws["A6"] = "Resumo Financeiro"
    ws["A6"].font = bold_font
    ws["A7"] = "SALDO ANTERIOR"
    ws["B7"] = totais["saldo_anterior"]
    ws["A8"] = "APLICAÇÕES (+)"
    ws["B8"] = totais["aplicacoes"]
    ws["A9"] = "RENDIMENTO BRUTO (+)"
    ws["B9"] = totais["rendimento_bruto"]
    ws["A10"] = "IMPOSTO DE RENDA (-)"
    ws["B10"] = totais["imposto_renda"]
    ws["A11"] = "IOF (-)"
    ws["B11"] = totais["iof"]
    ws["A12"] = "IMPOSTOS TOTAIS"
    ws["B12"] = totais["impostos"]
    ws["A13"] = "RENDIMENTO LÍQUIDO"
    ws["B13"] = totais["rendimento_liquido"]
    ws["A14"] = "RESGATES (-)"
    ws["B14"] = totais["resgate"]
    ws["A15"] = "SALDO ATUAL"
    ws["B15"] = totais["saldo_atual"]

    for row in range(7, 16):
        ws.cell(row=row, column=2).number_format = money_fmt

    headers = [
        "Instituição",
        "Produto",
        "Origem do Registro",
        "Conta",
        "Arquivo Importado",
        "SALDO ANTERIOR",
        "APLICAÇÕES (+)",
        "RENDIMENTO BRUTO (+)",
        "IMPOSTO DE RENDA (-)",
        "IOF",
        "IMPOSTOS TOTAIS",
        "RENDIMENTO LÍQUIDO",
        "RESGATES (-)",
        "SALDO ATUAL",
        "Observações"
    ]

    header_row = 17
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    first_data_row = header_row + 1
    for row_idx, reg in enumerate(registros, first_data_row):
        ws.cell(row=row_idx, column=1, value=reg.instituicao or "")
        ws.cell(row=row_idx, column=2, value=reg.produto or "")
        ws.cell(row=row_idx, column=3, value=(reg.origem_registro or "manual").replace("_", " ").title())
        ws.cell(row=row_idx, column=4, value=reg.conta_origem or "")
        ws.cell(row=row_idx, column=5, value=reg.arquivo_origem or "")
        ws.cell(row=row_idx, column=6, value=float(reg.saldo_anterior or 0))
        ws.cell(row=row_idx, column=7, value=float(reg.aplicacoes or 0))
        ws.cell(row=row_idx, column=8, value=float(reg.rendimento_bruto or 0))
        ws.cell(row=row_idx, column=9, value=float(reg.imposto_renda or 0))
        ws.cell(row=row_idx, column=10, value=float(reg.iof or 0))
        ws.cell(row=row_idx, column=11, value=float(reg.impostos or 0))
        ws.cell(row=row_idx, column=12, value=float(reg.rendimento_liquido or 0))
        ws.cell(row=row_idx, column=13, value=float(reg.resgate or 0))
        ws.cell(row=row_idx, column=14, value=float(reg.saldo_atual or 0))
        ws.cell(row=row_idx, column=15, value=reg.observacoes or "")

        for col in range(1, 16):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in [6, 7, 8, 9, 10, 11, 12, 13, 14]:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")

    total_row = first_data_row + len(registros)
    _excel_apply_zebra(ws, first_data_row, total_row - 1, 1, 15)
    ws.cell(row=total_row, column=1, value="TOTAIS").font = bold_font
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    for col, key in zip(
        [6, 7, 8, 9, 10, 11, 12, 13, 14],
        ["saldo_anterior", "aplicacoes", "rendimento_bruto", "imposto_renda", "iof", "impostos", "rendimento_liquido", "resgate", "saldo_atual"]
    ):
        cell = ws.cell(row=total_row, column=col, value=totais[key])
        cell.font = bold_font
        cell.number_format = money_fmt
        cell.alignment = Alignment(horizontal="right", vertical="center")

    for col in range(1, 16):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.border = thin_border

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:O{max(total_row, header_row + 1)}"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 38

    for row in [3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
        ws.cell(row=row, column=1).font = bold_font

    ws["A6"].fill = PatternFill("solid", fgColor="E2E8F0")
    ws["B6"].fill = PatternFill("solid", fgColor="E2E8F0")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=aplicacoes_financeiras_{mes_ref}.xlsx"}
    )


@app.get("/api/relatorios/consolidado-financeiro")
def exportar_consolidado_financeiro(
    ano: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    ano_ref = ano or date.today().year
    meses = [f"{ano_ref}-{m:02d}" for m in range(1, 13)]
    previsoes_ano = db.query(models.PrevisaoOrcamentaria).filter(models.PrevisaoOrcamentaria.ano == ano_ref).all()
    previsao_map = {
        (p.conta_id, p.mes): float(p.valor_previsto or 0)
        for p in previsoes_ano
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Consolidado {ano_ref}"

    detalhes_por_conta = []
    linhas_consolidado = []

    ws.merge_cells("A1:G1")
    ws["A1"] = f"RELATÓRIO CONSOLIDADO FINANCEIRO - {ano_ref}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    header_row = 3
    headers = [
        "Classificação (Nome)",
        "Previsão",
        "Mês",
        "Entradas",
        "Saídas",
        "Aplicações",
        "Saldo Líquido",
        "Diferença (Realizado - Previsão)",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        _excel_header_style(cell)

    first_data_row = header_row + 1
    total_entradas = 0.0
    total_saidas = 0.0
    total_aplicacoes = 0.0
    total_previsao = 0.0

    for idx, mes_ref in enumerate(meses):
        row = first_data_row + idx

        mensalidades = _pagamentos_pagos_membros_ativos_no_mes(db, mes_ref)
        outras_rendas = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_ref).all()
        despesas = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_ref).all()
        aplicacoes = db.query(models.AplicacaoFinanceira).filter(models.AplicacaoFinanceira.mes_referencia == mes_ref).all()

        entradas_conta_mes = defaultdict(float)
        entradas_meta_mes = {"1.1": "Mensalidades"}
        entradas_conta_mes["1.1"] += round(sum(float(p.valor_pago or 0) for p in mensalidades), 2)

        for r in outras_rendas:
            if not r.valor:
                continue
            codigo = (r.conta_codigo or "1.5").strip()
            nome = (r.conta_nome or r.categoria or "Outras arrecadações").strip()
            entradas_conta_mes[codigo] += float(r.valor)
            entradas_meta_mes[codigo] = nome

        saidas_conta_mes = defaultdict(float)
        saidas_meta_mes = {}
        for d in despesas:
            if not d.valor:
                continue
            codigo = (d.conta_codigo or "2.99").strip()
            nome = (d.conta_nome or d.categoria or "Outros").strip()
            saidas_conta_mes[codigo] += float(d.valor)
            saidas_meta_mes[codigo] = nome

        for codigo, valor in entradas_conta_mes.items():
            if abs(float(valor or 0)) <= 0.000001:
                continue
            conta_entrada = db.query(models.PlanoConta).filter(models.PlanoConta.codigo == codigo, models.PlanoConta.tipo == "entrada").first()
            previsao_valor = previsao_map.get((conta_entrada.id, int(mes_ref[-2:])), 0.0) if conta_entrada else 0.0
            linhas_consolidado.append({
                "classificacao": f"{codigo} - {entradas_meta_mes.get(codigo, '')}".strip(" -"),
                "previsao": float(previsao_valor or 0),
                "mes": mes_ref,
                "entradas": round(float(valor), 2),
                "saidas": 0.0,
                "aplicacoes": 0.0,
            })
            detalhes_por_conta.append({
                "mes": mes_ref,
                "tipo": "entrada",
                "codigo": codigo,
                "descricao": entradas_meta_mes.get(codigo, ""),
                "valor": round(float(valor), 2),
            })

        for codigo, valor in saidas_conta_mes.items():
            if abs(float(valor or 0)) <= 0.000001:
                continue
            conta_saida = db.query(models.PlanoConta).filter(models.PlanoConta.codigo == codigo, models.PlanoConta.tipo == "saida").first()
            previsao_valor = previsao_map.get((conta_saida.id, int(mes_ref[-2:])), 0.0) if conta_saida else 0.0
            linhas_consolidado.append({
                "classificacao": f"{codigo} - {saidas_meta_mes.get(codigo, '')}".strip(" -"),
                "previsao": float(previsao_valor or 0),
                "mes": mes_ref,
                "entradas": 0.0,
                "saidas": round(float(valor), 2),
                "aplicacoes": 0.0,
            })
            detalhes_por_conta.append({
                "mes": mes_ref,
                "tipo": "saida",
                "codigo": codigo,
                "descricao": saidas_meta_mes.get(codigo, ""),
                "valor": round(float(valor), 2),
            })

        for a in aplicacoes:
            aplicacao_valor = round(float(a.aplicacoes or 0), 2)
            if abs(aplicacao_valor) <= 0.000001:
                continue
            nome_aplicacao = " / ".join([x for x in [a.instituicao, a.produto] if x]) or "Aplicação financeira"
            linhas_consolidado.append({
                "classificacao": nome_aplicacao,
                "previsao": 0,
                "mes": mes_ref,
                "entradas": 0.0,
                "saidas": 0.0,
                "aplicacoes": aplicacao_valor,
            })

        entradas_valor = round(
            sum(float(p.valor_pago or 0) for p in mensalidades) +
            sum(float(r.valor or 0) for r in outras_rendas),
            2
        )
        saidas_valor = round(sum(float(d.valor or 0) for d in despesas), 2)
        aplicacoes_valor = round(sum(float(a.aplicacoes or 0) for a in aplicacoes), 2)
        saldo_liquido = round(entradas_valor - saidas_valor + aplicacoes_valor, 2)

        total_entradas += entradas_valor
        total_saidas += saidas_valor
        total_aplicacoes += aplicacoes_valor

    if not linhas_consolidado:
        linhas_consolidado.append({
            "classificacao": "Sem movimentação",
            "previsao": 0,
            "mes": f"{ano_ref}-01",
            "entradas": 0.0,
            "saidas": 0.0,
            "aplicacoes": 0.0,
        })

    linhas_consolidado = sorted(
        linhas_consolidado,
        key=lambda item: (item["mes"], item["classificacao"].lower())
    )

    for idx, item in enumerate(linhas_consolidado):
        row = first_data_row + idx
        previsao_linha = float(item["previsao"] or 0)
        saldo_linha = round(
            float(item["entradas"] or 0) - float(item["saidas"] or 0) + float(item["aplicacoes"] or 0),
            2
        )
        diferenca_linha = round(saldo_linha - previsao_linha, 2)
        ws.cell(row=row, column=1, value=item["classificacao"])
        ws.cell(row=row, column=2, value=previsao_linha)
        ws.cell(row=row, column=3, value=item["mes"])
        ws.cell(row=row, column=4, value=float(item["entradas"] or 0))
        ws.cell(row=row, column=5, value=float(item["saidas"] or 0))
        ws.cell(row=row, column=6, value=float(item["aplicacoes"] or 0))
        ws.cell(row=row, column=7, value=saldo_linha)
        ws.cell(row=row, column=8, value=diferenca_linha)
        total_previsao += previsao_linha

    fill_diff_positive = PatternFill("solid", fgColor="E6F4EA")
    fill_diff_negative = PatternFill("solid", fgColor="FDE8E8")
    fill_diff_neutral = PatternFill("solid", fgColor="F8FAFC")
    font_diff_positive = Font(color="166534")
    font_diff_negative = Font(color="B91C1C")
    font_diff_neutral = Font(color="374151")

    total_row = first_data_row + len(linhas_consolidado)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=round(total_entradas, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=round(total_saidas, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=round(total_aplicacoes, 2)).font = Font(bold=True)
    ws.cell(
        row=total_row,
        column=7,
        value=round(total_entradas - total_saidas + total_aplicacoes, 2)
    ).font = Font(bold=True)
    ws.cell(
        row=total_row,
        column=8,
        value=round((total_entradas - total_saidas + total_aplicacoes) - total_previsao, 2)
    ).font = Font(bold=True)

    for row in range(first_data_row, total_row + 1):
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        for col in (2, 4, 5, 6, 7, 8):
            ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")

    for col in range(1, 9):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="E2E8F0")

    _excel_apply_zebra(ws, first_data_row, total_row - 1, 1, 8)
    _excel_apply_borders(ws, header_row, total_row, 1, 8)

    for row in range(first_data_row, total_row + 1):
        cell_diff = ws.cell(row=row, column=8)
        valor_diff = float(cell_diff.value or 0)
        if valor_diff > 0:
            cell_diff.fill = fill_diff_positive
            cell_diff.font = Font(bold=(row == total_row), color=font_diff_positive.color)
        elif valor_diff < 0:
            cell_diff.fill = fill_diff_negative
            cell_diff.font = Font(bold=(row == total_row), color=font_diff_negative.color)
        else:
            cell_diff.fill = fill_diff_neutral
            cell_diff.font = Font(bold=(row == total_row), color=font_diff_neutral.color)

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:H{total_row}"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 22

    ws_det = wb.create_sheet(title=f"Detalhes Contas {ano_ref}")
    ws_det.merge_cells("A1:E1")
    ws_det["A1"] = f"DETALHAMENTO POR CONTA - {ano_ref}"
    ws_det["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws_det["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws_det["A1"].alignment = Alignment(horizontal="center", vertical="center")

    det_header_row = 3
    det_headers = ["Mês", "Tipo", "Código da Conta", "Descrição da Conta", "Valor"]
    for col, header in enumerate(det_headers, 1):
        _excel_header_style(ws_det.cell(row=det_header_row, column=col, value=header))

    detalhes_ordenados = sorted(
        detalhes_por_conta,
        key=lambda item: (item["mes"], item["tipo"], _ordenar_codigo_conta(item["codigo"]))
    )

    det_first_data_row = det_header_row + 1
    for idx, item in enumerate(detalhes_ordenados):
        row = det_first_data_row + idx
        ws_det.cell(row=row, column=1, value=item["mes"])
        ws_det.cell(row=row, column=2, value="Entrada" if item["tipo"] == "entrada" else "Saída")
        ws_det.cell(row=row, column=3, value=item["codigo"])
        ws_det.cell(row=row, column=4, value=item["descricao"])
        ws_det.cell(row=row, column=5, value=item["valor"])
        ws_det.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        ws_det.cell(row=row, column=5).alignment = Alignment(horizontal="right")

    det_last_row = max(det_first_data_row, det_first_data_row + len(detalhes_ordenados) - 1)
    _excel_apply_zebra(ws_det, det_first_data_row, det_last_row, 1, 5)
    _excel_apply_borders(ws_det, det_header_row, det_last_row, 1, 5)
    ws_det.freeze_panes = f"A{det_first_data_row}"
    ws_det.auto_filter.ref = f"A{det_header_row}:E{det_last_row}"
    ws_det.column_dimensions["A"].width = 14
    ws_det.column_dimensions["B"].width = 12
    ws_det.column_dimensions["C"].width = 18
    ws_det.column_dimensions["D"].width = 48
    ws_det.column_dimensions["E"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=consolidado_financeiro_{ano_ref}.xlsx"}
    )

@app.get("/api/relatorios/festas/{festa_id}")
def exportar_festa(
    festa_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    festa = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not festa:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    
    parts = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.festa_id == festa_id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participantes"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"LISTA DE PARTICIPANTES - {festa.nome_festa}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = f"Data: {festa.data_festa} | Local: {festa.local_festa}"
    ws.merge_cells("A2:F2")
    
    headers = ["Nome Participante", "Tipo", "Membro Titular", "Dependente/Convidado", "Custo Convite", "Pago"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        _excel_header_style(cell)
    
    first_data_row = header_row + 1
    for row, p in enumerate(parts, first_data_row):
        membro_nome = ""
        if p.membro_id:
            m = db.query(models.Membro).filter(models.Membro.id == p.membro_id).first()
            if m:
                membro_nome = m.nome_completo
        ws.cell(row=row, column=1, value=p.nome_participante)
        ws.cell(row=row, column=2, value=p.tipo_participante)
        ws.cell(row=row, column=3, value=membro_nome)
        ws.cell(row=row, column=4, value=p.nome_dependente or "")
        ws.cell(row=row, column=5, value=float(p.custo_convite) if p.custo_convite else 0)
        ws.cell(row=row, column=6, value="Sim" if p.pago else "Não")
        ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")

    total_row = first_data_row + len(parts)
    _excel_apply_zebra(ws, first_data_row, total_row - 1, 1, 6)
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=round(sum(float(p.custo_convite or 0) for p in parts), 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = 'R$ #,##0.00'
    ws.cell(row=total_row, column=4).fill = PatternFill("solid", fgColor="E2E8F0")
    ws.cell(row=total_row, column=5).fill = PatternFill("solid", fgColor="E2E8F0")

    _excel_apply_borders(ws, header_row, total_row, 1, 6)
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:F{total_row}"
    _excel_autofit_columns(ws, min_width=12, max_width=44)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=festa_{festa_id}.xlsx"}
    )

# ════════════════════════════════════════════════════════════════════════════════
# ETIQUETAS
# ════════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════════
# ETIQUETAS - LISTAR MEMBROS (JSON PARA FRONTEND)
# ════════════════════════════════════════════════════════════════════════════════

@app.get("/api/etiquetas/membros")
def listar_membros(
    status: Optional[str] = "ativo",
    nome: Optional[str] = None,
    cidade: Optional[str] = None,
    sexo: Optional[str] = None,
    categoria: Optional[str] = None,
    sem_email: Optional[bool] = False,
    sem_whatsapp: Optional[bool] = False,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Membro)

    if status:
        q = q.filter(models.Membro.status == status)

    if nome:
        q = q.filter(models.Membro.nome_completo.ilike(f"%{nome}%"))

    if cidade:
        q = q.filter(models.Membro.cidade.ilike(f"%{cidade}%"))

    if sexo:
        sexo_normalizado = sexo.strip().lower()
        if sexo_normalizado in ("masculino", "m"):
            valores_sexo = ["masculino", "m"]
        elif sexo_normalizado in ("feminino", "f"):
            valores_sexo = ["feminino", "f"]
        else:
            valores_sexo = [sexo_normalizado]

        q = q.filter(sql_func.lower(models.Membro.sexo).in_(valores_sexo))

    if categoria:
        categoria_normalizada = categoria.strip()
        if categoria_normalizada.lower() == "outros":
            q = q.filter(
                or_(
                    models.Membro.cat == None,
                    models.Membro.cat == "",
                    sql_func.lower(models.Membro.cat).notin_(["clt", "1711", "1712"])
                )
            )
        else:
            q = q.filter(sql_func.lower(models.Membro.cat) == categoria_normalizada.lower())

    if sem_email:
        q = q.filter(
            (models.Membro.email == None) |
            (models.Membro.email == "")
        )

    if sem_whatsapp:
        q = q.filter(
            (models.Membro.celular == None) |
            (models.Membro.celular == "")
        )

    membros = q.order_by(models.Membro.nome_completo).all()

    return {
        "total": len(membros),
        "membros": [
            {
                "id": m.id,
                "nome_completo": m.nome_completo,
                "endereco": m.endereco,
                "bairro": m.bairro,
                "cidade": m.cidade,
                "estado": m.estado,
                "cep": m.cep,
                "email": m.email,
                "telefone": m.telefone,
                "celular": m.celular,
            }
            for m in membros
        ]
    }


@app.get("/api/etiquetas")
def gerar_etiquetas(
    status: Optional[str] = "ativo",
    ids: Optional[str] = None,
    categoria: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors

    q = db.query(models.Membro)
    if ids:
        id_list = ids.split(",")
        q = q.filter(models.Membro.id.in_(id_list))
    elif status:
        q = q.filter(models.Membro.status == status)
    if categoria:
        categoria_normalizada = categoria.strip()
        if categoria_normalizada.lower() == "outros":
            q = q.filter(
                or_(
                    models.Membro.cat == None,
                    models.Membro.cat == "",
                    sql_func.lower(models.Membro.cat).notin_(["clt", "1711", "1712"])
                )
            )
        else:
            q = q.filter(sql_func.lower(models.Membro.cat) == categoria_normalizada.lower())
    membros = q.order_by(models.Membro.nome_completo).all()

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # 3x8 labels per page
    label_w = width / 3
    label_h = height / 8
    cols, rows_per_page = 3, 8

    x, y = 0, height - label_h
    count = 0

    for m in membros:
        if count > 0 and count % (cols * rows_per_page) == 0:
            c.showPage()
            x, y = 0, height - label_h

        col = count % cols
        row = (count % (cols * rows_per_page)) // cols

        lx = col * label_w
        ly = height - (row + 1) * label_h

        # Draw label border
        c.setStrokeColor(colors.grey)
        c.rect(lx + 2, ly + 2, label_w - 4, label_h - 4)

        # Write content
        c.setFont("Helvetica-Bold", 9)
        c.drawString(lx + 8, ly + label_h - 20, (m.nome_completo or "")[:40])
        c.setFont("Helvetica", 8)
        endereco = f"{m.endereco or ''}, {m.numero or ''}"
        if m.complemento:
            endereco += f" - {m.complemento}"
        c.drawString(lx + 8, ly + label_h - 35, endereco[:45])
        c.drawString(lx + 8, ly + label_h - 48, f"{m.bairro or ''}")
        c.drawString(lx + 8, ly + label_h - 61, f"{m.cidade or ''} - {m.estado or ''}")
        c.drawString(lx + 8, ly + label_h - 74, f"CEP: {m.cep or ''}")
        if m.matricula:
            c.drawString(lx + 8, ly + label_h - 87, f"Matrícula: {m.matricula}")

        count += 1

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=etiquetas.pdf"}
    )

# ════════════════════════════════════════════════════════════════════════════════
# SERVE FRONTEND STATIC FILES
# ════════════════════════════════════════════════════════════════════════════════
from fastapi.staticfiles import StaticFiles

_static_dir = os.path.join(os.path.dirname(__file__), "../frontend/dist")
if os.path.exists(_static_dir):
    app.mount("/assets", StaticFiles(directory=f"{_static_dir}/assets"), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_frontend(full_path: str):
        index_file = os.path.join(_static_dir, "index.html")
        return FileResponse(index_file)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
