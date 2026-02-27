from fastapi import FastAPI, Depends, HTTPException, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, and_, or_, cast, String
from datetime import date, datetime, timedelta
from typing import Optional, List
import uuid
import io
import os
import calendar
from collections import defaultdict

import models
import schemas
from database import engine, get_db, Base
from auth import (
    get_current_user, get_optional_user, verify_password,
    get_password_hash, create_access_token
)

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI(title="Associação de Aposentados API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Seed initial admin user ───────────────────────────────────────────────────
def seed_admin():
    from database import SessionLocal
    db = SessionLocal()
    try:
        admin = db.query(models.User).filter(models.User.email == "admin@associacao.com").first()
        if not admin:
            admin = models.User(
                id=str(uuid.uuid4()),
                email="admin@associacao.com",
                nome_completo="Administrador",
                role="administrador",
                password=get_password_hash("admin123"),
                ativo=True,
                created_at=datetime.utcnow()
            )
            db.add(admin)
            db.commit()
    finally:
        db.close()

seed_admin()

# ════════════════════════════════════════════════════════════════════════════════
# AUTH
# ════════════════════════════════════════════════════════════════════════════════
@app.post("/api/auth/login", response_model=schemas.TokenResponse)
def login(req: schemas.LoginRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == req.email).first()
    if not user or not verify_password(req.password, user.password):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    token = create_access_token({"sub": user.id})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user.id, "email": user.email, "nome_completo": user.nome_completo, "role": user.role}
    }

@app.get("/api/auth/me")
def me(current_user: models.User = Depends(get_current_user)):
    return {"id": current_user.id, "email": current_user.email,
            "nome_completo": current_user.nome_completo, "role": current_user.role}


# ════════════════════════════════════════════════════════════════════════════════
# USERS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/users", response_model=List[schemas.UserResponse])
def list_users(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    return db.query(models.User).all()

@app.post("/api/users", response_model=schemas.UserResponse)
def create_user(req: schemas.UserCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role != "administrador":
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar usuários")
    existing = db.query(models.User).filter(models.User.email == req.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    user = models.User(
        id=str(uuid.uuid4()),
        email=req.email,
        nome_completo=req.nome_completo,
        role=req.role,
        password=get_password_hash(req.password),
        ativo=True,
        created_at=datetime.utcnow()
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

@app.put("/api/users/{user_id}", response_model=schemas.UserResponse)
def update_user(user_id: str, req: schemas.UserUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        if k == "password":
            setattr(user, k, get_password_hash(v))
        else:
            setattr(user, k, v)
    db.commit()
    db.refresh(user)
    return user

@app.delete("/api/users/{user_id}")
def delete_user(user_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    db.delete(user)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# MEMBROS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/membros", response_model=List[schemas.MembroResponse])
def list_membros(
    skip: int = 0, limit: int = 500,
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Membro)
    if search:
        q = q.filter(or_(
            models.Membro.nome_completo.ilike(f"%{search}%"),
            models.Membro.cpf.ilike(f"%{search}%"),
            models.Membro.matricula.ilike(f"%{search}%"),
            models.Membro.email.ilike(f"%{search}%"),
        ))
    if status:
        q = q.filter(models.Membro.status == status)
    return q.order_by(models.Membro.nome_completo).offset(skip).limit(limit).all()

@app.post("/api/membros", response_model=schemas.MembroResponse)
def create_membro(req: schemas.MembroCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    membro = models.Membro(id=str(uuid.uuid4()), **req.dict(), created_at=datetime.utcnow(), updated_at=datetime.utcnow())
    db.add(membro)
    db.commit()
    db.refresh(membro)
    return membro

@app.get("/api/membros/{membro_id}", response_model=schemas.MembroResponse)
def get_membro(membro_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    m = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    return m

@app.put("/api/membros/{membro_id}", response_model=schemas.MembroResponse)
def update_membro(membro_id: str, req: schemas.MembroUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    m = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        setattr(m, k, v)
    m.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(m)
    return m

@app.delete("/api/membros/{membro_id}")
def delete_membro(membro_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    m = db.query(models.Membro).filter(models.Membro.id == membro_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    db.delete(m)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# PAGAMENTOS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/pagamentos")
def list_pagamentos(
    mes_referencia: Optional[str] = None,
    membro_id: Optional[str] = None,
    status_pagamento: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Pagamento)
    if mes_referencia:
        q = q.filter(models.Pagamento.mes_referencia == mes_referencia)
    if membro_id:
        q = q.filter(models.Pagamento.membro_id == membro_id)
    if status_pagamento:
        q = q.filter(models.Pagamento.status_pagamento == status_pagamento)
    pagamentos = q.all()
    result = []
    for p in pagamentos:
        pd = {
            "id": p.id, "membro_id": p.membro_id, "valor_pago": float(p.valor_pago) if p.valor_pago else 0,
            "mes_referencia": p.mes_referencia, "data_pagamento": str(p.data_pagamento) if p.data_pagamento else None,
            "status_pagamento": p.status_pagamento, "forma_pagamento": p.forma_pagamento,
            "observacoes": p.observacoes, "created_at": str(p.created_at) if p.created_at else None,
            "membro_nome": None
        }
        if p.membro_id:
            m = db.query(models.Membro).filter(models.Membro.id == p.membro_id).first()
            if m:
                pd["membro_nome"] = m.nome_completo
        result.append(pd)
    return result

@app.get("/api/pagamentos/painel")
def painel_pagamentos(
    mes_referencia: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    membros = db.query(models.Membro).filter(models.Membro.status == 'ativo').all()
    pagamentos = {p.membro_id: p for p in db.query(models.Pagamento).filter(
        models.Pagamento.mes_referencia == mes_referencia).all()}
    result = []
    for m in membros:
        p = pagamentos.get(m.id)
        result.append({
            "membro_id": m.id,
            "nome": m.nome_completo,
            "matricula": m.matricula,
            "valor_mensalidade": float(m.valor_mensalidade) if m.valor_mensalidade else 0,
            "pagamento_id": p.id if p else None,
            "valor_pago": float(p.valor_pago) if p and p.valor_pago else 0,
            "data_pagamento": str(p.data_pagamento) if p and p.data_pagamento else None,
            "status": p.status_pagamento if p else "pendente",
            "forma_pagamento": p.forma_pagamento if p else None,
        })
    return result

@app.post("/api/pagamentos", response_model=schemas.PagamentoResponse)
def create_pagamento(req: schemas.PagamentoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    # Check if payment already exists
    existing = db.query(models.Pagamento).filter(
        models.Pagamento.membro_id == req.membro_id,
        models.Pagamento.mes_referencia == req.mes_referencia
    ).first()
    if existing:
        for k, v in req.dict(exclude_none=True).items():
            setattr(existing, k, v)
        existing.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(existing)
        # Register transaction
        _register_transaction(db, existing, current_user.id)
        return existing
    
    p = models.Pagamento(
        id=str(uuid.uuid4()),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        **req.dict()
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    _register_transaction(db, p, current_user.id)
    return p

def _register_transaction(db, pagamento, user_id):
    m = db.query(models.Membro).filter(models.Membro.id == pagamento.membro_id).first()
    nome = m.nome_completo if m else "Membro"
    # Check if transaction exists for this pagamento
    existing = db.query(models.Transacao).filter(
        models.Transacao.origem == "mensalidade",
        models.Transacao.membro_id == pagamento.membro_id,
        cast(models.Transacao.categoria, String).ilike(f"%{pagamento.mes_referencia}%")
    ).first()
    if not existing:
        t = models.Transacao(
            id=str(uuid.uuid4()),
            user_id=user_id,
            descricao=f"Mensalidade {nome} - {pagamento.mes_referencia}",
            valor=pagamento.valor_pago,
            tipo="entrada",
            categoria=f"Mensalidade {pagamento.mes_referencia}",
            data_transacao=pagamento.data_pagamento or date.today(),
            origem="mensalidade",
            membro_id=pagamento.membro_id,
            created_at=datetime.utcnow()
        )
        db.add(t)
        db.commit()

@app.put("/api/pagamentos/{pagamento_id}", response_model=schemas.PagamentoResponse)
def update_pagamento(pagamento_id: str, req: schemas.PagamentoUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.Pagamento).filter(models.Pagamento.id == pagamento_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        setattr(p, k, v)
    p.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(p)
    return p

@app.delete("/api/pagamentos/{pagamento_id}")
def delete_pagamento(pagamento_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.Pagamento).filter(models.Pagamento.id == pagamento_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    db.delete(p)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# DESPESAS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/despesas", response_model=List[schemas.DespesaResponse])
def list_despesas(
    mes_referencia: Optional[str] = None,
    categoria: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Despesa)
    if mes_referencia:
        q = q.filter(models.Despesa.mes_referencia == mes_referencia)
    if categoria:
        q = q.filter(models.Despesa.categoria == categoria)
    return q.order_by(models.Despesa.data_despesa.desc()).all()

@app.post("/api/despesas", response_model=schemas.DespesaResponse)
def create_despesa(req: schemas.DespesaCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    mes_ref = req.mes_referencia or req.data_despesa.strftime("%Y-%m")
    descricao = (req.descricao or "").strip()
    d = models.Despesa(id=str(uuid.uuid4()), user_id=current_user.id, mes_referencia=mes_ref,
                        created_at=datetime.utcnow(), **req.dict(exclude={"mes_referencia", "descricao"}), descricao=descricao)
    d.mes_referencia = mes_ref
    db.add(d)
    # Register transaction
    t = models.Transacao(
        id=str(uuid.uuid4()), user_id=current_user.id,
        descricao=descricao, valor=req.valor, tipo="saida",
        categoria=req.categoria, data_transacao=req.data_despesa,
        origem="despesa", created_at=datetime.utcnow()
    )
    db.add(t)
    db.commit()
    db.refresh(d)
    return d

@app.put("/api/despesas/{despesa_id}", response_model=schemas.DespesaResponse)
def update_despesa(despesa_id: str, req: schemas.DespesaUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    d = db.query(models.Despesa).filter(models.Despesa.id == despesa_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    for k, v in req.dict(exclude_none=True).items():
        if k == "descricao" and isinstance(v, str):
            v = v.strip()
        setattr(d, k, v)
    db.commit()
    db.refresh(d)
    return d

@app.delete("/api/despesas/{despesa_id}")
def delete_despesa(despesa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    d = db.query(models.Despesa).filter(models.Despesa.id == despesa_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    db.delete(d)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# OUTRAS RENDAS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/outras-rendas", response_model=List[schemas.OutraRendaResponse])
def list_outras_rendas(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.OutraRenda)
    if mes_referencia:
        q = q.filter(models.OutraRenda.mes_referencia == mes_referencia)
    return q.order_by(models.OutraRenda.data_recebimento.desc()).all()

@app.post("/api/outras-rendas", response_model=schemas.OutraRendaResponse)
def create_outra_renda(req: schemas.OutraRendaCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    mes_ref = req.mes_referencia or req.data_recebimento.strftime("%Y-%m")
    r = models.OutraRenda(id=str(uuid.uuid4()), user_id=current_user.id, mes_referencia=mes_ref,
                           created_at=datetime.utcnow(), **req.dict(exclude={"mes_referencia"}))
    r.mes_referencia = mes_ref
    db.add(r)
    t = models.Transacao(
        id=str(uuid.uuid4()), user_id=current_user.id,
        descricao=req.descricao, valor=req.valor, tipo="entrada",
        categoria=req.categoria, data_transacao=req.data_recebimento,
        origem=req.categoria, created_at=datetime.utcnow()
    )
    db.add(t)
    db.commit()
    db.refresh(r)
    return r

@app.put("/api/outras-rendas/{renda_id}", response_model=schemas.OutraRendaResponse)
def update_outra_renda(renda_id: str, req: schemas.OutraRendaUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    r = db.query(models.OutraRenda).filter(models.OutraRenda.id == renda_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Renda não encontrada")
    for k, v in req.dict(exclude_none=True).items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    return r

@app.delete("/api/outras-rendas/{renda_id}")
def delete_outra_renda(renda_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    r = db.query(models.OutraRenda).filter(models.OutraRenda.id == renda_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Renda não encontrada")
    db.delete(r)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# TRANSAÇÕES / FLUXO DE CAIXA
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/transacoes")
def list_transacoes(
    mes_referencia: Optional[str] = None,
    tipo: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Transacao)
    if mes_referencia:
        ano, mes = mes_referencia.split("-")
        inicio = date(int(ano), int(mes), 1)
        ultimo_dia = calendar.monthrange(int(ano), int(mes))[1]
        fim = date(int(ano), int(mes), ultimo_dia)
        q = q.filter(models.Transacao.data_transacao.between(inicio, fim))
    if tipo:
        q = q.filter(models.Transacao.tipo == tipo)
    return q.order_by(models.Transacao.data_transacao.desc()).all()

@app.get("/api/fluxo-caixa")
def fluxo_caixa(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")
    ano, mes = mes_ref.split("-")
    inicio = date(int(ano), int(mes), 1)
    ultimo_dia = calendar.monthrange(int(ano), int(mes))[1]
    fim = date(int(ano), int(mes), ultimo_dia)

    transacoes = db.query(models.Transacao).filter(
        models.Transacao.data_transacao.between(inicio, fim)
    ).order_by(models.Transacao.data_transacao).all()

    total_entradas = sum(float(t.valor) for t in transacoes if t.tipo == "entrada" and t.valor)
    total_saidas = sum(float(t.valor) for t in transacoes if t.tipo == "saida" and t.valor)
    saldo = total_entradas - total_saidas

    # Monthly evolution (last 12 months)
    evolucao = []
    for i in range(11, -1, -1):
        ref_date = today - timedelta(days=i * 30)
        mes_iter = ref_date.strftime("%Y-%m")
        ano_i, mes_i = mes_iter.split("-")
        d_inicio = date(int(ano_i), int(mes_i), 1)
        d_fim = date(int(ano_i), int(mes_i), calendar.monthrange(int(ano_i), int(mes_i))[1])
        ts = db.query(models.Transacao).filter(models.Transacao.data_transacao.between(d_inicio, d_fim)).all()
        ent = sum(float(t.valor) for t in ts if t.tipo == "entrada" and t.valor)
        sai = sum(float(t.valor) for t in ts if t.tipo == "saida" and t.valor)
        evolucao.append({"mes": mes_iter, "entradas": ent, "saidas": sai, "saldo": ent - sai})

    return {
        "mes_referencia": mes_ref,
        "total_entradas": total_entradas,
        "total_saidas": total_saidas,
        "saldo": saldo,
        "transacoes": [{
            "id": t.id, "descricao": t.descricao, "valor": float(t.valor) if t.valor else 0,
            "tipo": t.tipo, "categoria": t.categoria,
            "data_transacao": str(t.data_transacao) if t.data_transacao else None,
            "origem": t.origem
        } for t in transacoes],
        "evolucao_mensal": evolucao
    }

@app.post("/api/transacoes", response_model=schemas.TransacaoResponse)
def create_transacao(req: schemas.TransacaoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    t = models.Transacao(id=str(uuid.uuid4()), user_id=current_user.id, created_at=datetime.utcnow(), **req.dict())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


# ════════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    today = date.today()
    mes_atual = today.strftime("%Y-%m")
    ano, mes = mes_atual.split("-")

    total_membros = db.query(models.Membro).filter(models.Membro.status == 'ativo').count()
    total_inativos = db.query(models.Membro).filter(models.Membro.status == 'inativo').count()

    # Pagamentos do mês
    pags_mes = db.query(models.Pagamento).filter(
        models.Pagamento.mes_referencia == mes_atual,
        models.Pagamento.status_pagamento == 'pago'
    ).all()
    total_arrecadado = sum(float(p.valor_pago) for p in pags_mes if p.valor_pago)
    total_pagantes = len(pags_mes)
    inadimplentes = total_membros - total_pagantes

    # Despesas do mês
    despesas_mes = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_atual).all()
    total_despesas = sum(d.valor for d in despesas_mes if d.valor)

    # Outras rendas do mês
    rendas_mes = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_atual).all()
    total_outras_rendas = sum(r.valor for r in rendas_mes if r.valor)

    saldo_mes = total_arrecadado + total_outras_rendas - total_despesas

    # Aniversariantes do mês
    aniv_mes = db.query(models.Membro).filter(
        func.strftime('%m', models.Membro.data_nascimento) == str(int(mes)).zfill(2)
    ).count()

    # Aniversariantes do dia
    aniv_hoje = db.query(models.Membro).filter(
        func.strftime('%m-%d', models.Membro.data_nascimento) == today.strftime('%m-%d')
    ).all()

    # Evolução últimos 6 meses
    evolucao = []
    for i in range(5, -1, -1):
        ref_date = today.replace(day=1) - timedelta(days=i * 28)
        mes_iter = ref_date.strftime("%Y-%m")
        pags_i = db.query(models.Pagamento).filter(
            models.Pagamento.mes_referencia == mes_iter,
            models.Pagamento.status_pagamento == 'pago'
        ).all()
        desp_i = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_iter).all()
        rend_i = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_iter).all()
        ent_i = sum(float(p.valor_pago) for p in pags_i if p.valor_pago) + sum(r.valor for r in rend_i if r.valor)
        sai_i = sum(d.valor for d in desp_i if d.valor)
        evolucao.append({"mes": mes_iter, "entradas": ent_i, "saidas": sai_i, "saldo": ent_i - sai_i})

    # Por status
    status_membros = db.query(models.Membro.status, func.count(models.Membro.id)).group_by(models.Membro.status).all()

    return {
        "total_membros": total_membros,
        "total_inativos": total_inativos,
        "total_pagantes": total_pagantes,
        "inadimplentes": inadimplentes,
        "total_arrecadado": total_arrecadado,
        "total_despesas": total_despesas,
        "total_outras_rendas": total_outras_rendas,
        "saldo_mes": saldo_mes,
        "aniversariantes_mes": aniv_mes,
        "aniversariantes_hoje": [{"id": m.id, "nome": m.nome_completo, "data_nascimento": str(m.data_nascimento)} for m in aniv_hoje],
        "evolucao_mensal": evolucao,
        "status_membros": [{"status": s, "total": c} for s, c in status_membros],
        "mes_atual": mes_atual,
    }


# ════════════════════════════════════════════════════════════════════════════════
# FESTAS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/festas", response_model=List[schemas.FestaResponse])
def list_festas(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    return db.query(models.Festa).order_by(models.Festa.data_festa.desc()).all()

@app.post("/api/festas", response_model=schemas.FestaResponse)
def create_festa(req: schemas.FestaCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = models.Festa(id=str(uuid.uuid4()), user_id=current_user.id, created_at=datetime.utcnow(), updated_at=datetime.utcnow(), **req.dict())
    db.add(f)
    db.commit()
    db.refresh(f)
    return f

@app.get("/api/festas/{festa_id}", response_model=schemas.FestaResponse)
def get_festa(festa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    return f

@app.put("/api/festas/{festa_id}", response_model=schemas.FestaResponse)
def update_festa(festa_id: str, req: schemas.FestaUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    for k, v in req.dict(exclude_none=True).items():
        setattr(f, k, v)
    f.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(f)
    return f

@app.delete("/api/festas/{festa_id}")
def delete_festa(festa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    db.delete(f)
    db.commit()
    return {"ok": True}

@app.get("/api/festas/{festa_id}/participantes")
def get_participantes(festa_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    parts = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.festa_id == festa_id).all()
    result = []
    for p in parts:
        pd = {
            "id": p.id, "festa_id": p.festa_id, "membro_id": p.membro_id,
            "nome_participante": p.nome_participante, "tipo_participante": p.tipo_participante,
            "custo_convite": float(p.custo_convite) if p.custo_convite else 0,
            "pago": p.pago, "data_pagamento": str(p.data_pagamento) if p.data_pagamento else None,
            "nome_dependente": p.nome_dependente, "parentesco": p.parentesco,
            "observacoes": p.observacoes, "membro_nome": None
        }
        if p.membro_id:
            m = db.query(models.Membro).filter(models.Membro.id == p.membro_id).first()
            if m:
                pd["membro_nome"] = m.nome_completo
        result.append(pd)
    return result

@app.post("/api/festas/{festa_id}/participantes", response_model=schemas.ParticipacaoResponse)
def add_participante(festa_id: str, req: schemas.ParticipacaoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    req.festa_id = festa_id
    p = models.ParticipacaoFesta(id=str(uuid.uuid4()), created_at=datetime.utcnow(), updated_at=datetime.utcnow(), **req.dict())
    db.add(p)
    db.commit()
    db.refresh(p)
    return p

@app.put("/api/participantes/{part_id}", response_model=schemas.ParticipacaoResponse)
def update_participante(part_id: str, req: schemas.ParticipacaoUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.id == part_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Participante não encontrado")
    for k, v in req.dict(exclude_none=True).items():
        setattr(p, k, v)
    p.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(p)
    return p

@app.delete("/api/participantes/{part_id}")
def delete_participante(part_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.id == part_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Participante não encontrado")
    db.delete(p)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# ANIVERSARIANTES
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/aniversariantes")
def aniversariantes(
    mes: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_filtro = mes or today.month
    membros = db.query(models.Membro).filter(
        models.Membro.data_nascimento.isnot(None),
        func.strftime('%m', models.Membro.data_nascimento) == str(mes_filtro).zfill(2)
    ).order_by(func.strftime('%d', models.Membro.data_nascimento)).all()
    
    result = []
    for m in membros:
        nascimento = m.data_nascimento
        idade = today.year - nascimento.year if nascimento else None
        is_hoje = (nascimento.month == today.month and nascimento.day == today.day) if nascimento else False
        result.append({
            "id": m.id, "nome": m.nome_completo, "data_nascimento": str(nascimento),
            "dia": nascimento.day if nascimento else None,
            "mes": nascimento.month if nascimento else None,
            "idade": idade, "email": m.email, "celular": m.celular,
            "aniversario_hoje": is_hoje
        })
    return result


# ════════════════════════════════════════════════════════════════════════════════
# CONCILIAÇÃO
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/conciliacao", response_model=List[schemas.ConciliacaoResponse])
def list_conciliacao(
    mes_referencia: Optional[str] = None,
    conciliado: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    q = db.query(models.Conciliacao)
    if mes_referencia:
        q = q.filter(models.Conciliacao.mes_referencia == mes_referencia)
    if conciliado is not None:
        q = q.filter(models.Conciliacao.conciliado == conciliado)
    return q.order_by(models.Conciliacao.data_extrato.desc()).all()

@app.post("/api/conciliacao", response_model=schemas.ConciliacaoResponse)
def create_conciliacao(req: schemas.ConciliacaoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    mes_ref = req.mes_referencia or req.data_extrato.strftime("%Y-%m")
    c = models.Conciliacao(
        id=str(uuid.uuid4()), user_id=current_user.id,
        mes_referencia=mes_ref, created_at=datetime.utcnow(), updated_at=datetime.utcnow(),
        **req.dict(exclude={"mes_referencia"})
    )
    c.mes_referencia = mes_ref
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

@app.put("/api/conciliacao/{conc_id}", response_model=schemas.ConciliacaoResponse)
def update_conciliacao(conc_id: str, req: schemas.ConciliacaoUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    c = db.query(models.Conciliacao).filter(models.Conciliacao.id == conc_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Conciliação não encontrada")
    for k, v in req.dict(exclude_none=True).items():
        setattr(c, k, v)
    c.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(c)
    return c

@app.delete("/api/conciliacao/{conc_id}")
def delete_conciliacao(conc_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    c = db.query(models.Conciliacao).filter(models.Conciliacao.id == conc_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Conciliação não encontrada")
    db.delete(c)
    db.commit()
    return {"ok": True}


# ════════════════════════════════════════════════════════════════════════════════
# FINANCEIRO / BALANCETE
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/financeiro/balancete")
def balancete(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")

    pags = db.query(models.Pagamento).filter(
        models.Pagamento.mes_referencia == mes_ref,
        models.Pagamento.status_pagamento == 'pago'
    ).all()
    desp = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_ref).all()
    rendas = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_ref).all()

    total_mensalidades = sum(float(p.valor_pago) for p in pags if p.valor_pago)
    total_outras = sum(r.valor for r in rendas if r.valor)
    total_entradas = total_mensalidades + total_outras
    total_saidas = sum(d.valor for d in desp if d.valor)
    saldo = total_entradas - total_saidas

    # Por categoria de despesa
    desp_por_cat = defaultdict(float)
    for d in desp:
        if d.valor:
            desp_por_cat[d.categoria or "Outros"] += d.valor

    # Por categoria de renda
    renda_por_cat = defaultdict(float)
    for r in rendas:
        if r.valor:
            renda_por_cat[r.categoria or "Outros"] += r.valor

    return {
        "mes_referencia": mes_ref,
        "total_mensalidades": total_mensalidades,
        "total_outras_rendas": total_outras,
        "total_entradas": total_entradas,
        "total_despesas": total_saidas,
        "saldo": saldo,
        "qtd_pagantes": len(pags),
        "despesas_por_categoria": dict(desp_por_cat),
        "rendas_por_categoria": dict(renda_por_cat),
        "pagamentos": [{
            "membro_id": p.membro_id, "valor_pago": float(p.valor_pago) if p.valor_pago else 0,
            "data_pagamento": str(p.data_pagamento) if p.data_pagamento else None,
            "forma_pagamento": p.forma_pagamento
        } for p in pags],
        "despesas": [{
            "id": d.id, "descricao": d.descricao, "categoria": d.categoria,
            "valor": d.valor, "data_despesa": str(d.data_despesa) if d.data_despesa else None
        } for d in desp],
    }


# ════════════════════════════════════════════════════════════════════════════════
# RELATÓRIOS / EXPORTAÇÃO EXCEL
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/relatorios/membros")
def exportar_membros(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    q = db.query(models.Membro)
    if status:
        q = q.filter(models.Membro.status == status)
    membros = q.order_by(models.Membro.nome_completo).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Membros"

    headers = ["Matrícula", "Nome", "CPF", "Email", "Telefone", "Celular",
               "Endereço", "Bairro", "Cidade", "Estado", "CEP",
               "Data Nascimento", "Data Associação", "Status", "Benefício", "Valor Mensalidade"]

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, m in enumerate(membros, 2):
        ws.cell(row=row, column=1, value=m.matricula)
        ws.cell(row=row, column=2, value=m.nome_completo)
        ws.cell(row=row, column=3, value=m.cpf)
        ws.cell(row=row, column=4, value=m.email)
        ws.cell(row=row, column=5, value=m.telefone)
        ws.cell(row=row, column=6, value=m.celular)
        ws.cell(row=row, column=7, value=m.endereco)
        ws.cell(row=row, column=8, value=m.bairro)
        ws.cell(row=row, column=9, value=m.cidade)
        ws.cell(row=row, column=10, value=m.estado)
        ws.cell(row=row, column=11, value=m.cep)
        ws.cell(row=row, column=12, value=str(m.data_nascimento) if m.data_nascimento else "")
        ws.cell(row=row, column=13, value=str(m.data_associacao) if m.data_associacao else "")
        ws.cell(row=row, column=14, value=m.status)
        ws.cell(row=row, column=15, value=m.beneficio)
        ws.cell(row=row, column=16, value=float(m.valor_mensalidade) if m.valor_mensalidade else 0)

    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=membros.xlsx"}
    )

@app.get("/api/relatorios/pagamentos")
def exportar_pagamentos(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")

    membros = db.query(models.Membro).filter(models.Membro.status == 'ativo').order_by(models.Membro.nome_completo).all()
    pagamentos = {p.membro_id: p for p in db.query(models.Pagamento).filter(models.Pagamento.mes_referencia == mes_ref).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Pagamentos {mes_ref}"

    headers = ["Matrícula", "Nome", "Valor Mensalidade", "Valor Pago", "Data Pagamento", "Status", "Forma Pagamento"]
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    red_fill = PatternFill("solid", fgColor="FFB3B3")
    green_fill = PatternFill("solid", fgColor="B3FFB3")

    for row, m in enumerate(membros, 2):
        p = pagamentos.get(m.id)
        status = p.status_pagamento if p else "pendente"
        fill = green_fill if status == "pago" else red_fill
        
        values = [
            m.matricula, m.nome_completo,
            float(m.valor_mensalidade) if m.valor_mensalidade else 0,
            float(p.valor_pago) if p and p.valor_pago else 0,
            str(p.data_pagamento) if p and p.data_pagamento else "",
            status,
            p.forma_pagamento if p else ""
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pagamentos_{mes_ref}.xlsx"}
    )

@app.get("/api/relatorios/aniversariantes")
def exportar_aniversariantes(
    mes: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    today = date.today()
    mes_filtro = mes or today.month
    membros = db.query(models.Membro).filter(
        models.Membro.data_nascimento.isnot(None),
        func.strftime('%m', models.Membro.data_nascimento) == str(mes_filtro).zfill(2)
    ).order_by(func.strftime('%d', models.Membro.data_nascimento)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Aniversariantes Mês {mes_filtro}"
    headers = ["Nome", "Data Nascimento", "Dia", "Idade", "Email", "Celular", "Telefone"]
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for row, m in enumerate(membros, 2):
        idade = today.year - m.data_nascimento.year
        ws.cell(row=row, column=1, value=m.nome_completo)
        ws.cell(row=row, column=2, value=str(m.data_nascimento))
        ws.cell(row=row, column=3, value=m.data_nascimento.day)
        ws.cell(row=row, column=4, value=idade)
        ws.cell(row=row, column=5, value=m.email)
        ws.cell(row=row, column=6, value=m.celular)
        ws.cell(row=row, column=7, value=m.telefone)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=aniversariantes_mes_{mes_filtro}.xlsx"}
    )

@app.get("/api/relatorios/balancete")
def exportar_balancete(
    mes_referencia: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    today = date.today()
    mes_ref = mes_referencia or today.strftime("%Y-%m")
    pags = db.query(models.Pagamento).filter(
        models.Pagamento.mes_referencia == mes_ref,
        models.Pagamento.status_pagamento == 'pago'
    ).all()
    desp = db.query(models.Despesa).filter(models.Despesa.mes_referencia == mes_ref).all()
    rendas = db.query(models.OutraRenda).filter(models.OutraRenda.mes_referencia == mes_ref).all()

    wb = openpyxl.Workbook()
    
    # Sheet 1: Resumo
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1["A1"] = f"BALANCETE MENSAL - {mes_ref}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A3"] = "ENTRADAS"
    ws1["A3"].font = Font(bold=True)
    ws1["A4"] = "Mensalidades"
    ws1["B4"] = sum(float(p.valor_pago) for p in pags if p.valor_pago)
    ws1["A5"] = "Outras Rendas"
    ws1["B5"] = sum(r.valor for r in rendas if r.valor)
    ws1["A6"] = "TOTAL ENTRADAS"
    ws1["A6"].font = Font(bold=True)
    ws1["B6"] = ws1["B4"].value + ws1["B5"].value
    ws1["B6"].font = Font(bold=True)
    ws1["A8"] = "SAÍDAS"
    ws1["A8"].font = Font(bold=True)
    ws1["A9"] = "Despesas"
    ws1["B9"] = sum(d.valor for d in desp if d.valor)
    ws1["A10"] = "TOTAL SAÍDAS"
    ws1["A10"].font = Font(bold=True)
    ws1["B10"] = ws1["B9"].value
    ws1["B10"].font = Font(bold=True)
    ws1["A12"] = "SALDO"
    ws1["A12"].font = Font(bold=True, size=12)
    ws1["B12"] = ws1["B6"].value - ws1["B10"].value
    ws1["B12"].font = Font(bold=True, size=12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=balancete_{mes_ref}.xlsx"}
    )

@app.get("/api/relatorios/festas/{festa_id}")
def exportar_festa(
    festa_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    festa = db.query(models.Festa).filter(models.Festa.id == festa_id).first()
    if not festa:
        raise HTTPException(status_code=404, detail="Festa não encontrada")
    
    parts = db.query(models.ParticipacaoFesta).filter(models.ParticipacaoFesta.festa_id == festa_id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participantes"
    ws["A1"] = f"LISTA DE PARTICIPANTES - {festa.nome_festa}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Data: {festa.data_festa} | Local: {festa.local_festa}"
    
    headers = ["Nome Participante", "Tipo", "Membro Titular", "Dependente/Convidado", "Custo Convite", "Pago"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1e3a5f")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, p in enumerate(parts, 5):
        membro_nome = ""
        if p.membro_id:
            m = db.query(models.Membro).filter(models.Membro.id == p.membro_id).first()
            if m:
                membro_nome = m.nome_completo
        ws.cell(row=row, column=1, value=p.nome_participante)
        ws.cell(row=row, column=2, value=p.tipo_participante)
        ws.cell(row=row, column=3, value=membro_nome)
        ws.cell(row=row, column=4, value=p.nome_dependente or "")
        ws.cell(row=row, column=5, value=float(p.custo_convite) if p.custo_convite else 0)
        ws.cell(row=row, column=6, value="Sim" if p.pago else "Não")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=festa_{festa_id}.xlsx"}
    )


# ════════════════════════════════════════════════════════════════════════════════
# ETIQUETAS
# ════════════════════════════════════════════════════════════════════════════════
@app.get("/api/etiquetas")
def gerar_etiquetas(
    status: Optional[str] = "ativo",
    ids: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors

    q = db.query(models.Membro)
    if ids:
        id_list = ids.split(",")
        q = q.filter(models.Membro.id.in_(id_list))
    elif status:
        q = q.filter(models.Membro.status == status)
    membros = q.order_by(models.Membro.nome_completo).all()

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # 3x8 labels per page
    label_w = width / 3
    label_h = height / 8
    cols, rows_per_page = 3, 8

    x, y = 0, height - label_h
    count = 0

    for m in membros:
        if count > 0 and count % (cols * rows_per_page) == 0:
            c.showPage()
            x, y = 0, height - label_h

        col = count % cols
        row = (count % (cols * rows_per_page)) // cols

        lx = col * label_w
        ly = height - (row + 1) * label_h

        # Draw label border
        c.setStrokeColor(colors.grey)
        c.rect(lx + 2, ly + 2, label_w - 4, label_h - 4)

        # Write content
        c.setFont("Helvetica-Bold", 9)
        c.drawString(lx + 8, ly + label_h - 20, (m.nome_completo or "")[:40])
        c.setFont("Helvetica", 8)
        endereco = f"{m.endereco or ''}, {m.numero or ''}"
        if m.complemento:
            endereco += f" - {m.complemento}"
        c.drawString(lx + 8, ly + label_h - 35, endereco[:45])
        c.drawString(lx + 8, ly + label_h - 48, f"{m.bairro or ''}")
        c.drawString(lx + 8, ly + label_h - 61, f"{m.cidade or ''} - {m.estado or ''}")
        c.drawString(lx + 8, ly + label_h - 74, f"CEP: {m.cep or ''}")
        if m.matricula:
            c.drawString(lx + 8, ly + label_h - 87, f"Matrícula: {m.matricula}")

        count += 1

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=etiquetas.pdf"}
    )


# ════════════════════════════════════════════════════════════════════════════════
# SERVE FRONTEND STATIC FILES
# ════════════════════════════════════════════════════════════════════════════════
from fastapi.staticfiles import StaticFiles

_static_dir = os.path.join(os.path.dirname(__file__), "../frontend/dist")
if os.path.exists(_static_dir):
    app.mount("/assets", StaticFiles(directory=f"{_static_dir}/assets"), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_frontend(full_path: str):
        index_file = os.path.join(_static_dir, "index.html")
        return FileResponse(index_file)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
